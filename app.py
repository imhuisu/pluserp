import sys, subprocess as _sp

# Windows 콘솔 QuickEdit 모드 비활성화 (클릭 시 프로세스 멈춤 방지)
if sys.platform == 'win32':
    try:
        import ctypes
        _k32 = ctypes.windll.kernel32
        _h = _k32.GetStdHandle(-10)
        _m = ctypes.c_ulong()
        _k32.GetConsoleMode(_h, ctypes.byref(_m))
        _k32.SetConsoleMode(_h, _m.value & ~0x0040)
    except Exception:
        pass
try:
    from fpdf import FPDF as _fpdf_check
except ImportError:
    print("[앱 시작] fpdf2 없음 → 자동 설치 중...")
    _sp.check_call([sys.executable, "-m", "pip", "install", "fpdf2"])
    print("[앱 시작] fpdf2 설치 완료")

from flask import Flask, render_template, request, jsonify, send_from_directory, session, redirect, url_for
import sqlite3
import csv
import json
from pathlib import Path
from datetime import datetime, timedelta
import os
from base64 import urlsafe_b64encode
try:
    from pywebpush import webpush, WebPushException
    from cryptography.hazmat.primitives.asymmetric import ec
    from cryptography.hazmat.primitives import serialization
    PUSH_AVAILABLE = True
except Exception:
    PUSH_AVAILABLE = False
    WebPushException = Exception
import uuid
import re
import subprocess
import shutil
from urllib.parse import quote
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from PIL import Image, ImageOps
from openpyxl import load_workbook
import xlrd

APP_DIR = Path(__file__).resolve().parent
DB_PATH = APP_DIR.parent / "plusdoor.db"
UPLOAD_DIR = APP_DIR.parent / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)
VEHICLE_DOCS_DIR = UPLOAD_DIR / "vehicle_docs"
VEHICLE_DOCS_DIR.mkdir(exist_ok=True)
EQUIPMENT_DOCS_DIR = UPLOAD_DIR / "equipment_docs"
EQUIPMENT_DOCS_DIR.mkdir(exist_ok=True)
CONSTRUCTION_DOCS_DIR = UPLOAD_DIR / "construction_docs"
CONSTRUCTION_DOCS_DIR.mkdir(exist_ok=True)

VAPID_KEYS_FILE = APP_DIR / "vapid_keys.json"
VAPID_KEYS = {}

def _init_vapid():
    global VAPID_KEYS
    if not PUSH_AVAILABLE:
        return
    if VAPID_KEYS_FILE.exists():
        with open(VAPID_KEYS_FILE) as f:
            VAPID_KEYS = json.load(f)
        return
    priv = ec.generate_private_key(ec.SECP256R1())
    private_pem = priv.private_bytes(
        serialization.Encoding.PEM,
        serialization.PrivateFormat.TraditionalOpenSSL,
        serialization.NoEncryption()
    ).decode()
    pub_bytes = priv.public_key().public_bytes(
        serialization.Encoding.X962,
        serialization.PublicFormat.UncompressedPoint
    )
    public_key = urlsafe_b64encode(pub_bytes).rstrip(b'=').decode()
    VAPID_KEYS = {"private_pem": private_pem, "public_key": public_key}
    with open(VAPID_KEYS_FILE, 'w') as f:
        json.dump(VAPID_KEYS, f, indent=2)

_init_vapid()

WORK_ORDER_TEMPLATE_DIR = APP_DIR / "work_order_templates"
WORK_ORDER_DATA_DIR = APP_DIR / "work_order_data"
WORK_ORDER_OUTPUT_DIR = APP_DIR / "work_order_output"
for _p in [WORK_ORDER_TEMPLATE_DIR, WORK_ORDER_DATA_DIR, WORK_ORDER_OUTPUT_DIR]:
    _p.mkdir(exist_ok=True)

AS_VIDEO_EXCEL_PATH = APP_DIR / "as_order_templates" / "as.xlsx"

WORK_ORDER_CSV_PATH = WORK_ORDER_DATA_DIR / "work_orders.csv"
WORK_ORDER_TEMPLATE_PATH = WORK_ORDER_TEMPLATE_DIR / "나인도어_비단열.xlsx"
WORK_ORDER_MAPPING_PATH = WORK_ORDER_TEMPLATE_DIR / "나인도어_비단열_mapping.csv"
SOFFICE_PATH = Path(r"C:\Program Files\LibreOffice\program\soffice.exe")
WORK_ORDER_PREVIEW_DIR = WORK_ORDER_OUTPUT_DIR / "preview"
WORK_ORDER_PREVIEW_DIR.mkdir(exist_ok=True)

WORK_ORDER_FIELDS = [
    "id","created_at","updated_at","writer","status",
    "workDate","productGroup","templateType","scheduleDate","releaseDate",
    "customer","siteName","model","width","height","qty","deliveryType",
    "smallDoorDirection","buryDepth","namma","hingeDirection","smallDoorOpen",
    "doorSize","doorGlass","grill","smallDoorSize","smallDoorGlass","smallDoorGrill",
    "postBar","dongBar","danBar","smallPost","smallTop","frame","reinforce","topExtra",
    "colorChange","color","panel1","panel2","panel3","keyHeight","alColor","panelColor",
    "digitalLock","keyType","content","specialNotes","memo","customerContact","output_xlsx","preview_png"
]

DEFAULT_WORK_ORDER_MAPPING_ROWS = [
    ("workDate","작성일","H1"),
    ("model","모델번호","C3"),
    ("width","가로","C4"),
    ("height","세로","C5"),
    ("buryDepth","묻힘","C6"),
    ("smallDoorDirection","소대방향","C7"),
    ("hingeDirection","경첩방향","C8"),
    ("smallDoorOpen","소대개폐","C9"),
    ("doorSize","문사이즈","C10"),
    ("namma","남마포함높이","F5"),
    ("doorGlass","판, 유리","G5"),
    ("keyType","키 종류","A40"),
    ("customer","거래처","B46"),
    ("scheduleDate","생산일정","G46"),
    ("memo","비고","B48"),
    ("deliveryType","출고구분","D48"),
    ("releaseDate","출고일자","G48"),
]

app = Flask(__name__)
app.secret_key = os.environ.get("PLUSHDOOR_SECRET_KEY", "plusdoor-change-this-secret-key")
app.config["MAX_CONTENT_LENGTH"] = 30 * 1024 * 1024
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(hours=2)

def get_conn():
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA busy_timeout=5000")
    return conn

def column_exists(conn, table, column):
    rows = conn.execute(f"PRAGMA table_info({table})").fetchall()
    return any(r["name"] == column for r in rows)

def table_exists(conn, table):
    row = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone()
    return row is not None

def reset_price_tables_if_old_schema(conn):
    # 예전 테스트/관리형 단가표 테이블이 남아있으면 현재 견적형 단가표 구조와 충돌한다.
    # 특히 price_products에 series 컬럼이 없으면 /api/price_data에서 오류가 나므로
    # 단가표 관련 테이블만 초기화하고 CSV에서 다시 가져오도록 한다.
    if table_exists(conn, "price_products") and not column_exists(conn, "price_products", "series"):
        for table in ["price_notes", "price_options", "price_types", "price_regions", "price_products"]:
            conn.execute(f"DROP TABLE IF EXISTS {table}")


PERMISSION_LABELS = {
    "price": "단가표",
    "journal_sales": "영업일지",
    "journal_consult": "상담일지",
    "journal_measure": "실측일지",
    "journal_install": "시공일지",
    "work_order": "작업지시서",
    "production": "생산스케줄",
    "completed": "제작완료",
    "as": "A/S",
    "cs_install": "시공",
    "calendar": "캘린더",
    "mail": "서식관리",
    "ledger_vehicle": "차량관리대장",
    "ledger_machine": "설비관리대장",
    "hr_attendance": "근태관리",
    "hr_leave": "휴가관리",
    "hr_org": "조직도",
    "user_manage": "사용자관리",
}
# UI용 트리 구조 — 리프 키만 PERMISSION_LABELS에 있고, 부모는 UI 그룹핑 전용
PERMISSION_TREE = [
    {"key": "price",    "label": "단가표"},
    {"label": "일지관리", "children": [
        {"key": "journal_sales",   "label": "영업일지"},
        {"key": "journal_consult", "label": "상담일지"},
        {"key": "journal_measure", "label": "실측일지"},
        {"key": "journal_install", "label": "시공일지"},
    ]},
    {"key": "work_order",  "label": "작업지시서"},
    {"key": "production",  "label": "생산스케줄"},
    {"key": "completed",   "label": "제작완료"},
    {"label": "C/S관리", "children": [
        {"key": "as",          "label": "A/S"},
        {"key": "cs_install",  "label": "시공"},
    ]},
    {"key": "calendar",    "label": "캘린더"},
    {"key": "mail",        "label": "서식관리"},
    {"label": "대장관리", "children": [
        {"key": "ledger_vehicle", "label": "차량관리대장"},
        {"key": "ledger_machine", "label": "설비관리대장"},
    ]},
    {"label": "인사관리", "children": [
        {"key": "hr_attendance", "label": "근태관리"},
        {"key": "hr_leave",      "label": "휴가관리"},
        {"key": "hr_org",        "label": "조직도"},
    ]},
    {"key": "user_manage", "label": "사용자관리"},
]
PERMISSION_LEVELS = {
    "none": "못봄",
    "read": "읽기전용",
    "write": "수정/저장",
}
MENU_ITEMS = [
    ("price", "단가표", "/price"),
    ("journal", "일지관리", "/journal"),
    ("work_order", "작업지시서", "/work_order"),
    ("production", "생산스케줄", "/"),
    ("completed", "제작완료", "/completed"),
    ("as", "C/S관리", "/as"),
    ("calendar", "캘린더", "/calendar"),
    ("mail", "서식관리", "/mail"),
    ("ledger", "대장관리", "/ledger"),
    ("hr", "인사관리", "/hr"),
]
ALL_PERMISSIONS = list(PERMISSION_LABELS.keys())
WRITE_METHODS = {"POST", "PUT", "PATCH", "DELETE"}

ORG_DATA = [
    {
        "id": "executive", "label": "경영진",
        "members": [
            {"rank": "회장",    "name": "윤병광", "phone": "010-5404-2600", "photo": "img_r003_c063.jpg"},
            {"rank": "대표이사", "name": "윤영석", "phone": "010-3434-2602", "photo": "img_r010_c064.jpg"},
            {"rank": "대표이사", "name": "남승희", "phone": "010-3434-5295", "photo": "img_r017_c064.jpg"},
        ]
    },
    {
        "id": "strategy", "label": "미래전략팀",
        "members": [
            {"rank": "차장", "team": "품질,생산,안전", "name": "윤성태", "phone": "010-4646-1081", "photo": "img_r026_c075.jpg"},
            {"rank": "차장", "team": "인사관리",       "name": "임희수", "phone": "010-5121-0927", "photo": "img_r026_c085.jpg"},
            {"rank": "과장", "team": "개발, 설계",     "name": "남승훈", "phone": "010-8815-5295", "photo": "img_r026_c095.jpg"},
            {"rank": "과장", "team": "마케팅,품질",    "name": "임진묵", "phone": "010-5242-8343", "photo": "img_r025_c105.jpg"},
        ]
    },
    {
        "id": "management", "label": "경영관리부",
        "head": {"rank": "이사", "name": "신서경", "phone": "010-9242-5383", "photo": "img_r036_c006.jpg"},
        "teams": [
            {"label": "생산관리팀", "members": [
                {"rank": "과장", "name": "박평운", "phone": "010-3659-8335", "photo": "img_r044_c007.jpg"},
                {"rank": "대리", "name": "강준상", "phone": "010-3075-1073", "photo": "img_r051_c006.jpg"},
                {"rank": "사원", "name": "김용성", "phone": "010-8253-3070", "photo": "img_r058_c007.jpg"},
            ]},
            {"label": "판매관리팀", "members": [
                {"rank": "대리", "name": "김미혜", "phone": "010-8842-4424", "photo": "img_r067_c002.jpg"},
                {"rank": "사원", "name": "차선아", "phone": "010-4562-7471", "photo": "img_r074_c002.jpg"},
            ]},
            {"label": "영업관리팀", "members": [
                {"rank": "대리", "name": "김준홍", "phone": "010-3131-7378", "photo": "img_r067_c012.jpg"},
                {"rank": "과장", "name": "고은영", "phone": "010-7359-0307", "photo": "img_r073_c012.jpg"},
                {"rank": "주임", "name": "신지원", "phone": "010-2798-0091", "photo": "img_r081_c012.jpg"},
            ]},
        ]
    },
    {
        "id": "sales", "label": "영업부",
        "head": {"rank": "과장", "name": "이세영", "phone": "010-3896-1678", "photo": "img_r036_c037.jpg"},
        "teams": [
            {"label": "영업1팀", "members": [
                {"rank": "과장", "name": "이세영",  "phone": "010-3896-1678", "photo": "img_r045_c023.jpg"},
                {"rank": "대리", "name": "임원산",  "phone": "010-4178-2188", "photo": "img_r052_c023.jpg"},
            ]},
            {"label": "영업3팀", "members": [
                {"rank": "대리", "name": "김영진",  "phone": "010-2380-1853", "photo": "img_r045_c043.jpg"},
                {"rank": "대리", "name": "이원희",  "phone": "010-5747-8784", "photo": "img_r052_c042.jpg"},
            ]},
            {"label": "영업4팀", "members": [
                {"rank": "대리", "name": "정욱기",  "phone": "010-9101-7782", "photo": "img_r045_c052.jpg"},
            ]},
        ]
    },
    {
        "id": "production", "label": "생산부",
        "head": {"rank": "공장장", "name": "임명택", "phone": "010-6434-4478", "photo": "img_r036_c089.jpg"},
        "teams": [
            {"label": "도어팀", "members": [
                {"rank": "이사", "name": "최준식", "phone": "010-8849-2090", "photo": "img_r045_c063.jpg"},
                {"rank": "주임", "name": "류승만", "phone": "010-9486-9843", "photo": "img_r051_c064.jpg"},
                {"rank": "사원", "name": "김동율", "phone": "010-7418-7942", "photo": "img_r059_c064.jpg"},
                {"rank": "사원", "name": "이창재", "phone": "010-4634-5644", "photo": "img_r066_c063.jpg"},
                {"rank": "사원", "name": "이훈",   "phone": "010-2573-5627", "photo": "img_r073_c064.jpg"},
                {"rank": "사원", "name": "민승찬", "phone": "010-8144-7947", "photo": "img_r080_c064.jpg"},
            ]},
            {"label": "폴딩·시스템팀", "members": [
                {"rank": "대리", "name": "박준영", "phone": "010-2414-1824", "photo": "img_r045_c073.jpg"},
                {"rank": "주임", "name": "이종준", "phone": "010-2863-2803", "photo": "img_r052_c074.jpg"},
                {"rank": "이사", "name": "박규석", "phone": "010-2286-1285", "photo": "img_r044_c085.jpg"},
                {"rank": "과장", "name": "박종덕", "phone": "010-8147-5825", "photo": "img_r052_c084.jpg"},
                {"rank": "주임", "name": "이영돈", "phone": "010-5462-4278", "photo": "img_r059_c084.jpg"},
                {"rank": "사원", "name": "김시현", "phone": "010-8837-0438", "photo": "img_r045_c094.jpg"},
                {"rank": "사원", "name": "양성열", "phone": "010-2316-9378", "photo": "img_r052_c094.jpg"},
                {"rank": "주임", "name": "김범식", "phone": "010-2334-6711", "photo": "img_r065_c084.jpg"},
                {"rank": "사원", "name": "박성준", "phone": "010-6343-0949", "photo": "img_r066_c074.jpg"},
                {"rank": "사원", "name": "최민수", "phone": "010-4172-4236", "photo": "img_r073_c074.jpg"},
                {"rank": "사원", "name": "이환인", "phone": "010-4798-6879", "photo": None},
            ]},
            {"label": "중문·방충망팀", "members": [
                {"rank": "대리", "name": "윤소라", "phone": "010-2365-1771", "photo": "img_r045_c106.jpg"},
                {"rank": "사원", "name": "이선근", "phone": "010-9336-2888", "photo": "img_r051_c106.jpg"},
            ]},
        ]
    },
    {
        "id": "cs", "label": "C/S관리부",
        "head": {"rank": "대리", "name": "홍대권", "phone": "010-9632-6515", "photo": "img_r036_c121.jpg"},
        "teams": [
            {"label": "C/S팀", "members": [
                {"rank": "과장", "name": "정종식", "phone": "010-9748-1123", "photo": None},
                {"rank": "과장", "name": "이정석", "phone": "010-9980-3164", "photo": "img_r052_c116.jpg"},
                {"rank": "대리", "name": "황기연", "phone": "010-5852-0382", "photo": "img_r059_c116.jpg"},
                {"rank": "대리", "name": "노왕우", "phone": "010-6487-0929", "photo": "img_r065_c117.jpg"},
                {"rank": "대리", "name": "이문상", "phone": "010-5479-6253", "photo": "img_r073_c117.jpg"},
            ]},
            {"label": "시공팀", "members": [
                {"rank": "과장", "name": "김진학", "phone": "010-3922-9155", "photo": "img_r045_c127.jpg"},
                {"rank": "대리", "name": "오병래", "phone": "010-3412-3927", "photo": "img_r052_c127.jpg"},
                {"rank": "주임", "name": "남상규", "phone": "010-4611-9782", "photo": "img_r059_c128.jpg"},
                {"rank": "주임", "name": "김현우", "phone": "010-3757-2523", "photo": "img_r066_c127.jpg"},
                {"rank": "사원", "name": "이선용", "phone": "010-9426-0816", "photo": "img_r073_c128.jpg"},
            ]},
        ]
    },
]


def blank_permission_map():
    return {k: "none" for k in ALL_PERMISSIONS}


def all_write_permission_map():
    return {k: "write" for k in ALL_PERMISSIONS}


def normalize_permissions(value):
    """권한 문자열을 {menu: none/read/write} 형태로 변환한다.
    v56/v57의 기존 콤마 권한은 모두 수정/저장 권한으로 자동 변환된다.
    """
    result = blank_permission_map()
    if not value:
        return result
    if isinstance(value, dict):
        source = value
    else:
        text = str(value).strip()
        if not text:
            return result
        if text.startswith("{"):
            try:
                source = json.loads(text)
            except Exception:
                source = {}
        else:
            # 예전 형식: price,journal,production → 모두 write로 취급
            for key in [x.strip() for x in text.split(",") if x.strip()]:
                if key in result:
                    result[key] = "write"
            return result
    for key, level in (source or {}).items():
        if key in result and level in PERMISSION_LEVELS:
            result[key] = level
    return result


def permissions_to_db(value, role="일반"):
    if role == "관리자":
        data = all_write_permission_map()
    else:
        data = normalize_permissions(value)
    return json.dumps(data, ensure_ascii=False)


def permission_level(perm):
    if not session.get("user_id"):
        return "none"
    if session.get("role") == "관리자":
        return "write"
    return normalize_permissions(session.get("permissions", "")).get(perm, "none")


def user_has_perm(perm):
    return permission_level(perm) in ("read", "write")


def user_can_write(perm):
    return permission_level(perm) == "write"


def require_login(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login_page", next=request.path))
        return fn(*args, **kwargs)
    return wrapper


def require_perm(perm):
    def deco(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not session.get("user_id"):
                return redirect(url_for("login_page", next=request.path))
            if not user_has_perm(perm):
                return "권한이 없습니다.", 403
            return fn(*args, **kwargs)
        return wrapper
    return deco


def require_write_perm(perm):
    def deco(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not session.get("user_id"):
                return redirect(url_for("login_page", next=request.path))
            if not user_can_write(perm):
                return "수정/저장 권한이 없습니다.", 403
            return fn(*args, **kwargs)
        return wrapper
    return deco


def send_push_notifications(user_ids, title, body, url="/as"):
    if not PUSH_AVAILABLE or not user_ids or not VAPID_KEYS.get("private_pem"):
        return
    try:
        conn = get_conn()
        ph = ",".join("?" * len(user_ids))
        subs = conn.execute(
            f"SELECT * FROM push_subscriptions WHERE user_id IN ({ph})", user_ids
        ).fetchall()
        conn.close()
        dead = []
        for sub in subs:
            try:
                webpush(
                    subscription_info={
                        "endpoint": sub["endpoint"],
                        "keys": {"p256dh": sub["p256dh"], "auth": sub["auth"]}
                    },
                    data=json.dumps({"title": title, "body": body, "url": url}),
                    vapid_private_key=VAPID_KEYS["private_pem"],
                    vapid_claims={"sub": "mailto:info@plusdoor.com"}
                )
            except WebPushException as e:
                if e.response and e.response.status_code in (404, 410):
                    dead.append(sub["id"])
            except Exception:
                pass
        if dead:
            conn2 = get_conn()
            conn2.execute(f"DELETE FROM push_subscriptions WHERE id IN ({','.join('?'*len(dead))})", dead)
            conn2.commit()
            conn2.close()
    except Exception:
        pass


def render_top_menu(active_key=""):
    css = (
        '<style id="pd-topbar">'
        ':root{--pdn:#1a3a5c;--pda:#e85d04}'
        'body{padding-top:0!important}'
        '.top-menu{display:none!important}'
        '.current-page{display:none!important}'
        '.pd-bar{display:flex!important;position:sticky!important;top:0!important;'
            'z-index:99999!important;background:#1a3a5c!important;height:48px!important;'
            'align-items:center!important;gap:4px!important;padding:0 14px!important;'
            'box-shadow:0 2px 8px rgba(0,0,0,.22)!important;flex-wrap:nowrap!important}'
        '.pd-bar .logo{font-size:14px!important;font-weight:700!important;color:#fff!important;'
            'white-space:nowrap!important;margin-right:6px!important;display:inline-block!important}'
        '.pd-bar .sep{width:1px!important;height:22px!important;'
            'background:rgba(255,255,255,.2)!important;margin:0 2px!important;flex-shrink:0!important}'
        '.pd-bar .pd-nav{display:flex!important;align-items:center!important;'
            'gap:4px!important;flex:1!important;overflow-x:auto!important;overflow-y:hidden!important;'
            'scrollbar-width:none!important}'
        '.pd-bar .pd-nav::-webkit-scrollbar{display:none!important}'
        '.pd-bar .nb{padding:5px 11px!important;border:1px solid rgba(255,255,255,.25)!important;'
            'background:transparent!important;color:rgba(255,255,255,.75)!important;'
            'cursor:pointer!important;border-radius:4px!important;font-size:12px!important;'
            'font-weight:600!important;white-space:nowrap!important;'
            'text-decoration:none!important;display:inline-flex!important;align-items:center!important}'
        '.pd-bar .nb:hover{background:rgba(255,255,255,.18)!important;color:#fff!important}'
        '.pd-bar .nb.on{background:#fff!important;color:#1a3a5c!important;'
            'font-weight:700!important;border-color:#fff!important}'
        '.pd-bar .nr{position:relative!important;padding:5px 11px!important;'
            'border:1px solid rgba(255,255,255,.25)!important;background:transparent!important;'
            'color:rgba(255,255,255,.8)!important;cursor:pointer!important;border-radius:4px!important;'
            'font-size:12px!important}'
        '.pd-bar .nr:hover{background:rgba(255,255,255,.18)!important}'
        '.pd-bar .nr .bx{position:absolute!important;top:-5px!important;right:-5px!important;'
            'background:#e85d04!important;color:#fff!important;border-radius:10px!important;'
            'padding:1px 5px!important;font-size:10px!important;font-weight:700!important;'
            'min-width:18px!important;text-align:center!important;pointer-events:none!important}'
        '.pd-bar .ml{margin-left:auto!important;display:flex!important;'
            'align-items:center!important;gap:8px!important;flex-shrink:0!important}'
        '.pd-bar .un{font-size:12px!important;color:rgba(255,255,255,.65)!important;'
            'white-space:nowrap!important}'
        '.pd-bar .lo{padding:4px 10px!important;border:1px solid rgba(255,255,255,.2)!important;'
            'background:transparent!important;color:rgba(255,255,255,.55)!important;'
            'cursor:pointer!important;border-radius:4px!important;font-size:11px!important}'
        '.pd-bar .lo:hover{color:#fff!important}'
        '@media(max-width:767px){'
            '.pd-bar{height:auto!important;flex-wrap:wrap!important;padding:4px 10px 0!important}'
            '.pd-bar .pd-nav{display:none!important}'
            '.pd-bar .sep.nav-sep{display:none!important}'
            '.pd-bar .un{display:none!important}'
            '.pd-menu-btn{display:flex!important;order:10!important;width:100%!important;'
                'justify-content:center!important;margin:4px 0 6px!important;'
                'border-radius:8px!important;padding:8px!important;font-size:13px!important}'
        '}'
        '.pd-menu-btn{display:none;align-items:center;gap:6px;padding:5px 12px;'
            'border:1px solid rgba(255,255,255,.35);background:transparent;color:#fff;'
            'cursor:pointer;border-radius:6px;font-size:13px;font-weight:700;flex-shrink:0}'
        '.pd-menu-btn:active{background:rgba(255,255,255,.15)}'
        '.pd-nav-sheet-bg{display:none;position:fixed;inset:0;background:rgba(0,0,0,.35);'
            'z-index:999999;align-items:flex-start;justify-content:stretch}'
        '.pd-nav-sheet-bg.open{display:flex}'
        '.pd-nav-sheet{background:#fff;width:100%;border-radius:0 0 16px 16px;'
            'box-shadow:0 6px 24px rgba(0,0,0,.22);max-height:80vh;overflow-y:auto}'
        '.pd-nav-sheet-hd{display:flex;align-items:center;justify-content:space-between;'
            'padding:12px 18px 8px;border-bottom:1px solid #eee}'
        '.pd-nav-sheet-hd span{font-size:15px;font-weight:800;color:#1a3a5c}'
        '.pd-nav-sheet-hd button{background:none;border:none;font-size:22px;cursor:pointer;'
            'color:#888;line-height:1;padding:0 2px}'
        '.pd-nav-sheet-body{display:grid;grid-template-columns:1fr 1fr;gap:0;padding:8px}'
        '.pd-si{display:flex;align-items:center;padding:14px 16px;border-radius:10px;'
            'font-size:14px;font-weight:700;color:#1a3a5c;cursor:pointer;'
            'text-decoration:none!important;width:100%;text-align:left;gap:6px}'
        '.pd-si:active{background:#f0f4ff}'
        '.pd-si.on{color:#e85d04}'
        '.pd-nav-sheet-logout{margin:6px 8px 8px;padding:13px;border-radius:10px;'
            'background:#fee2e2;color:#991b1b;font-size:13px;font-weight:700;'
            'border:none;cursor:pointer;width:calc(100% - 16px)}'
        '</style>'
    )
    parts = [css, '<div class="pd-bar">',
             '<span class="logo">PLUSDOOR</span>',
             '<span class="sep"></span>',
             '<div class="pd-nav">']
    JOURNAL_KEYS = ["journal_sales", "journal_consult", "journal_measure", "journal_install"]
    LEDGER_KEYS  = ["ledger_vehicle", "ledger_machine"]
    HR_KEYS      = ["hr_attendance", "hr_leave", "hr_org"]
    CS_KEYS      = ["as", "cs_install"]
    sheet_items = []
    for key, label, href in MENU_ITEMS:
        if key == "journal":
            visible = any(user_has_perm(k) for k in JOURNAL_KEYS)
        elif key == "ledger":
            visible = any(user_has_perm(k) for k in LEDGER_KEYS)
        elif key == "hr":
            visible = any(user_has_perm(k) for k in HR_KEYS)
        elif key == "as":
            visible = any(user_has_perm(k) for k in CS_KEYS)
        else:
            visible = user_has_perm(key)
        if visible:
            if key == active_key:
                parts.append(f'<a class="nb on" href="{href}">{label}</a>')
                sheet_items.append(f'<a class="pd-si on" href="{href}" onclick="pdCloseNavSheet()">{label}</a>')
            else:
                parts.append(f'<a class="nb" href="{href}">{label}</a>')
                sheet_items.append(f'<a class="pd-si" href="{href}">{label}</a>')
    if user_has_perm("user_manage"):
        if active_key == "user_manage":
            parts.append('<a class="nb on" href="/users">사용자관리</a>')
            sheet_items.append('<a class="pd-si on" href="/users" onclick="pdCloseNavSheet()">사용자관리</a>')
        else:
            parts.append('<a class="nb" href="/users">사용자관리</a>')
            sheet_items.append('<a class="pd-si" href="/users">사용자관리</a>')
    parts.append('</div>')
    parts.append('<button class="pd-menu-btn" onclick="pdOpenNavSheet()">&#9776; 메뉴</button>')
    user_name = session.get("name") or session.get("username") or ""
    parts.append('<span class="sep nav-sep"></span>')
    parts.append(
        '<button class="nr" id="checkRequestTopBtn"' +
        ' onclick="if(window.showCheckRequestInbox){showCheckRequestInbox()}' +
        'else{alert(\'\uc0dd\uc0b0\uc2a4\ucf00\uc904 \ud654\uba74\uc5d0\uc11c ' +
        '\ud655\uc778\uc694\uccad\ud568\uc744 \ud655\uc778\ud558\uc138\uc694.\')}">&#128276; ' +
        '\ud655\uc778\uc694\uccad<span class="bx" id="checkRequestBadge">0</span></button>'
    )
    parts.append('<div class="ml">')
    if user_name:
        parts.append(f'<span class="un">&#128100; {user_name}</span>')
    parts.append(
        '<button class="lo" onclick="location.href=\'/logout\'">\ub85c\uadf8\uc544\uc6c3</button>'
    )
    parts.append('</div></div>')
    sheet_body = ''.join(sheet_items)
    parts.append(f'''<div class="pd-nav-sheet-bg" id="pdNavSheetBg" onclick="if(event.target===this)pdCloseNavSheet()">
  <div class="pd-nav-sheet">
    <div class="pd-nav-sheet-hd"><span>&#9776; 메뉴</span><button onclick="pdCloseNavSheet()">&#10005;</button></div>
    <div class="pd-nav-sheet-body">{sheet_body}</div>
    <button class="pd-nav-sheet-logout" onclick="location.href='/logout'">로그아웃</button>
  </div>
</div>
<script id="pd-nav-sheet-js">
function pdOpenNavSheet(){{
  var bar=document.querySelector('.pd-bar');
  var bg=document.getElementById('pdNavSheetBg');
  if(bar)bg.style.paddingTop=bar.getBoundingClientRect().bottom+'px';
  bg.classList.add('open');
  document.body.style.overflow='hidden';
}}
function pdCloseNavSheet(){{document.getElementById('pdNavSheetBg').classList.remove('open');document.body.style.overflow='';}}
window.addEventListener('resize',function(){{if(window.innerWidth>767)pdCloseNavSheet();}});
window.addEventListener('orientationchange',function(){{setTimeout(function(){{if(window.innerWidth>767)pdCloseNavSheet();}},150);}});
</script>''')
    parts.append('''<div id="pdCheckInboxModal" onclick="if(event.target===this)pdHideCheckInbox()" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.45);z-index:999999;align-items:center;justify-content:center;padding:16px">
  <div style="background:#fff;border-radius:10px;width:min(600px,98vw);max-height:88vh;display:flex;flex-direction:column;box-shadow:0 20px 60px rgba(0,0,0,.3);overflow:hidden">
    <div style="background:#1a3a5c;color:#fff;padding:12px 16px;display:flex;align-items:center;justify-content:space-between;flex-shrink:0">
      <span style="font-weight:700;font-size:14px">&#128276; 확인요청함</span>
      <div style="display:flex;gap:8px;align-items:center">
        <button id="pdTabInbox" onclick="pdLoadCheckInbox(\'inbox\')" style="background:#fff;color:#1a3a5c;border:none;border-radius:4px;padding:4px 12px;font-size:12px;font-weight:700;cursor:pointer">받은 요청</button>
        <button id="pdTabSent" onclick="pdLoadCheckInbox(\'sent\')" style="background:rgba(255,255,255,.2);color:#fff;border:none;border-radius:4px;padding:4px 12px;font-size:12px;cursor:pointer">보낸 요청</button>
        <button onclick="pdHideCheckInbox()" style="background:transparent;border:none;color:rgba(255,255,255,.8);font-size:20px;cursor:pointer;line-height:1;padding:0 4px">&#10005;</button>
      </div>
    </div>
    <div id="pdCheckInboxList" style="overflow-y:auto;flex:1;padding:12px 16px;min-height:100px"></div>
  </div>
</div>
<script id="pd-check-inbox">
(function(){
  var currentMode="inbox";
  function esc(v){return String(v||"").replace(/[&<>"\']/g,function(m){return{"&":"&amp;","<":"&lt;",">":"&gt;",\'"\':"&quot;","\'":"&#39;"}[m];});}
  async function updateBadge(){
    try{
      var r=await fetch("/api/check_requests/count");
      var d=await r.json();
      var b=document.getElementById("checkRequestBadge");
      if(b){var cnt=d.count||0;b.textContent=cnt;b.style.display=cnt>0?"":"none";}
    }catch(e){}
  }
  updateBadge();
  setInterval(updateBadge,30000);
  window.showCheckRequestInbox=function(){
    var m=document.getElementById("pdCheckInboxModal");
    if(m)m.style.display="flex";
    pdLoadCheckInbox("inbox");
  };
  window.pdHideCheckInbox=function(){
    var m=document.getElementById("pdCheckInboxModal");
    if(m)m.style.display="none";
  };
  window.pdLoadCheckInbox=async function(mode){
    currentMode=mode;
    var ti=document.getElementById("pdTabInbox");
    var ts=document.getElementById("pdTabSent");
    var onStyle="background:#fff;color:#1a3a5c;border:none;border-radius:4px;padding:4px 12px;font-size:12px;font-weight:700;cursor:pointer";
    var offStyle="background:rgba(255,255,255,.2);color:#fff;border:none;border-radius:4px;padding:4px 12px;font-size:12px;cursor:pointer";
    if(ti)ti.style.cssText=mode==="inbox"?onStyle:offStyle;
    if(ts)ts.style.cssText=mode==="sent"?onStyle:offStyle;
    var list=document.getElementById("pdCheckInboxList");
    if(!list)return;
    list.innerHTML="<div style=\'padding:20px;text-align:center;color:#999\'>불러오는 중...</div>";
    try{
      var r=await fetch("/api/check_requests?mode="+mode);
      var rows=await r.json();
      if(!rows.length){list.innerHTML="<div style=\'padding:20px;text-align:center;color:#999\'>"+(mode==="inbox"?"받은 요청이 없습니다":"보낸 요청이 없습니다")+"</div>";return;}
      var html=rows.map(function(req){
        var isNew=req.status==="미확인";
        var jobParts=[req.customer,req.product_group,req.model].filter(Boolean);
        var jobInfo=jobParts.join(" ");
        var detail=req.detail_content?"("+esc(req.detail_content)+")":"";
        var who=mode==="inbox"?"발신: "+esc(req.from_user_name||""):"→ "+esc(req.to_user_name||req.to_group||"");
        var statusBadge=isNew?"<span style=\'background:#e85d04;color:#fff;border-radius:3px;padding:1px 6px;font-size:10px;font-weight:700;margin-left:4px\'>미확인</span>":"<span style=\'background:#16a34a;color:#fff;border-radius:3px;padding:1px 6px;font-size:10px;margin-left:4px\'>완료</span>";
        var doneInfo=req.status==="확인완료"?"<div style=\'font-size:11px;color:#16a34a;margin-top:4px\'>&#10003; 확인완료: "+esc(req.completed_by||"")+" ("+esc((req.completed_at||"").slice(0,16))+")</div>":"";
        var isLeave=req.source_type==="leave"&&req.leave_request_id;
        var completeBtn="";
        if(isNew&&mode==="inbox"){
          if(isLeave){
            completeBtn="<div style=\'display:flex;gap:8px;justify-content:flex-end;margin-top:8px\'>"
              +"<button onclick=\'pdApproveLeave("+req.leave_request_id+",&#39;승인&#39;)\' style=\'background:#16a34a;color:#fff;border:none;border-radius:4px;padding:5px 18px;font-size:12px;font-weight:700;cursor:pointer\'>승인</button>"
              +"<button onclick=\'pdApproveLeave("+req.leave_request_id+",&#39;반려&#39;)\' style=\'background:#dc2626;color:#fff;border:none;border-radius:4px;padding:5px 18px;font-size:12px;font-weight:700;cursor:pointer\'>반려</button>"
              +"</div>";
          } else {
            completeBtn="<div style=\'text-align:right;margin-top:8px\'><button onclick=\'pdCompleteCheckReq("+req.id+")\' style=\'background:#1a3a5c;color:#fff;border:none;border-radius:4px;padding:5px 16px;font-size:12px;font-weight:700;cursor:pointer\'>확인완료</button></div>";
          }
        }
        return "<div style=\'border:1px solid "+(isNew?"#f4a86a":"#e0e0e0")+";border-radius:6px;padding:10px 14px;margin-bottom:8px;background:"+(isNew?"#fff8f0":"#fafafa")+"\'>"
          +"<div style=\'display:flex;align-items:center;justify-content:space-between;margin-bottom:6px\'>"
          +"<span style=\'font-weight:700;font-size:13px;color:#1a3a5c\'>"+who+statusBadge+"</span>"
          +"<span style=\'font-size:11px;color:#888\'>"+esc((req.created_at||"").slice(0,16))+"</span>"
          +"</div>"
          +(jobInfo?"<div style=\'font-size:12px;color:#555;margin-bottom:6px\'>"+esc(jobInfo)+" "+detail+"</div>":"")
          +"<div style=\'font-size:13px;color:#222;white-space:pre-wrap;border-left:3px solid "+(isNew?"#e85d04":"#ccc")+";padding-left:8px;margin-bottom:2px\'>"+esc(req.message||"")+"</div>"
          +doneInfo+completeBtn
          +"</div>";
      }).join("");
      list.innerHTML=html;
    }catch(e){list.innerHTML="<div style=\'padding:20px;text-align:center;color:#dc2626\'>불러오기 실패</div>";}
  };
  window.pdCompleteCheckReq=async function(id){
    try{
      await fetch("/api/check_requests/"+id+"/complete",{method:"PUT"});
      pdLoadCheckInbox(currentMode);
      updateBadge();
    }catch(e){alert("처리 실패");}
  };
  window.pdApproveLeave=async function(leaveId,action){
    if(!confirm(action==="승인"?"승인하시겠습니까?":"반려하시겠습니까?"))return;
    try{
      var r=await fetch("/api/hr/leave/requests/"+leaveId+"/approve",{method:"POST",headers:{"Content-Type":"application/json"},body:JSON.stringify({action:action})});
      var d=await r.json();
      if(!r.ok||d.error){alert(d.error||"처리 실패");return;}
      alert(action+"되었습니다.");
      pdLoadCheckInbox(currentMode);
      updateBadge();
    }catch(e){alert("처리 실패");}
  };
})();
</script>''')
    parts.append('''<style id="pd-noscroll-style">body.pd-noscroll{overflow:hidden!important;touch-action:none!important}</style>''')
    parts.append('''<script id="pd-modal-scroll-lock">(function(){
  var MODAL_SEL=[
    '.modal','.detail-modal','.check-request-v14-modal',
    '.completed-edit-v7-modal','.as-check-request-v14-modal',
    '.schedule-form-modal','.plusdoor-calendar-pop',
    '.as-mobile-detail-modal','[id$="Modal"]:not(#checkRequestBadge)'
  ].join(',');
  function sync(){
    var open=false;
    try{
      var els=document.querySelectorAll(MODAL_SEL);
      for(var i=0;i<els.length;i++){
        var d=els[i].style.display;
        if(d&&d!=='none'){open=true;break;}
      }
    }catch(e){}
    document.body.classList.toggle('pd-noscroll',open);
  }
  new MutationObserver(sync).observe(document.documentElement,{subtree:true,attributes:true,attributeFilter:['style','class']});
})();</script>''')
    parts.append('<script id="pd-nav-scroll"></script>')
    parts.append('''<script id="pd-push-sub">(function(){
  async function initPush(){
    if(!('PushManager' in window)||!('serviceWorker' in navigator))return;
    var reg=await navigator.serviceWorker.ready;
    var existing=await reg.pushManager.getSubscription();
    if(existing)return;
    var perm=Notification.permission;
    if(perm==='denied')return;
    if(perm==='default')perm=await Notification.requestPermission();
    if(perm!=='granted')return;
    try{
      var r=await fetch('/api/push/vapid-public-key');
      var d=await r.json();
      function toUint8(b64){
        var pad='='.repeat((4-b64.length%4)%4);
        var raw=atob((b64+pad).replace(/-/g,'+').replace(/_/g,'/'));
        var arr=new Uint8Array(raw.length);
        for(var i=0;i<raw.length;i++)arr[i]=raw.charCodeAt(i);
        return arr;
      }
      var sub=await reg.pushManager.subscribe({userVisibleOnly:true,applicationServerKey:toUint8(d.public_key)});
      await fetch('/api/push/subscribe',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(sub)});
    }catch(e){}
  }
  if('serviceWorker' in navigator){navigator.serviceWorker.ready.then(function(){setTimeout(initPush,2000);});}
})();</script>''')
    parts.append('''<script id="pd-pwa">(function(){
  if(!document.querySelector('link[rel="manifest"]')){
    var ml=document.createElement('link');ml.rel='manifest';ml.href='/static/manifest.json';document.head.appendChild(ml);}
  [['mobile-web-app-capable','yes'],['apple-mobile-web-app-capable','yes'],
   ['apple-mobile-web-app-status-bar-style','black-translucent'],
   ['apple-mobile-web-app-title','PLUSDOOR'],['theme-color','#1a3a5c']
  ].forEach(function(a){
    if(!document.querySelector('meta[name="'+a[0]+'"]')){
      var m=document.createElement('meta');m.name=a[0];m.content=a[1];document.head.appendChild(m);}
  });
  if(!document.querySelector('link[rel="apple-touch-icon"]')){
    var ai=document.createElement('link');ai.rel='apple-touch-icon';ai.href='/static/img/logo.png';document.head.appendChild(ai);}
  if('serviceWorker' in navigator){
    navigator.serviceWorker.register('/sw.js',{scope:'/'}).catch(function(){});}
})();</script>''')
    return ''.join(parts)


@app.context_processor
def inject_auth_helpers():
    return {
        "top_menu": render_top_menu,
        "has_perm": user_has_perm,
        "has_write_perm": user_can_write,
        "permission_labels": PERMISSION_LABELS,
        "permission_levels": PERMISSION_LEVELS,
        "permission_tree": PERMISSION_TREE,
    }


def page_permission_for_path(path):
    if path == "/":
        return "production"
    return {
        "/price": "price",
        "/work_order": "work_order",
        "/completed": "completed",
        "/as": "as",
        "/calendar": "calendar",
        "/mail": "mail",
        "/users": "user_manage",
    }.get(path)


def api_permission_for_path(path):
    # 확인요청은 상세보기에서 조회권한 사용자도 요청할 수 있도록 별도 엔드포인트에서 직접 검사한다.
    if path.startswith("/api/check_requests") or path.startswith("/api/check_request_targets"):
        return None
    if path.startswith("/api/users"):
        return "user_manage"
    if path.startswith("/api/price"):
        return "price"

    if path.startswith("/api/work_orders"):
        return "work_order"
    if path.startswith("/api/calendar_events") or path.startswith("/api/delivery_people"):
        return "calendar"
    if path.startswith("/api/completed_search") or path.startswith("/api/photos") or (path.startswith("/api/schedules/") and "/photos" in path) or (path.startswith("/api/schedules/") and path.endswith("/complete")):
        return "completed"
    if path.startswith("/api/customers") or path.startswith("/api/customer_") or path.startswith("/api/schedules") or path.startswith("/api/active_summary"):
        return "production"
    if path.startswith("/api/as_ref"):
        return "as"
    if path.startswith("/api/mail"):
        return "mail"
    return None


@app.before_request
def enforce_login_and_timeout():
    public_paths = ["/login", "/api/login", "/api/change_password", "/static/", "/.well-known/"]
    if request.path.startswith(tuple(public_paths)):
        return None

    if not session.get("user_id"):
        if request.path.startswith("/api/"):
            return jsonify({"ok": False, "error": "로그인이 필요합니다."}), 401
        return redirect(url_for("login_page", next=request.path))

    now_ts = datetime.now().timestamp()
    last = float(session.get("last_activity", now_ts))
    if now_ts - last > 2 * 60 * 60:
        session.clear()
        if request.path.startswith("/api/"):
            return jsonify({"ok": False, "error": "로그인 시간이 만료되었습니다."}), 401
        return redirect(url_for("login_page", expired="1"))
    session["last_activity"] = now_ts

    path = request.path
    page_perm = page_permission_for_path(path)
    if page_perm and not user_has_perm(page_perm):
        return "권한이 없습니다.", 403

    api_perm = api_permission_for_path(path)
    if api_perm:
        if request.method in WRITE_METHODS:
            if not user_can_write(api_perm):
                return jsonify({"ok": False, "error": "수정/저장 권한이 없습니다."}), 403
        elif not user_has_perm(api_perm):
            return jsonify({"ok": False, "error": "권한이 없습니다."}), 403
    return None


def _migrate_sales_customers(conn):
    """sales_customers 테이블 데이터를 customers 테이블로 일회성 마이그레이션"""
    try:
        sc_rows = conn.execute("SELECT * FROM sales_customers").fetchall()
    except Exception:
        return  # 테이블 없으면 스킵
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for sc in sc_rows:
        name = sc["customer"] or ""
        if not name:
            continue
        existing = conn.execute("SELECT id FROM customers WHERE name=?", (name,)).fetchone()
        if existing:
            cid = existing["id"]
            conn.execute("""
                UPDATE customers SET
                    writer           = CASE WHEN COALESCE(writer,'')='' THEN ? ELSE writer END,
                    address          = CASE WHEN COALESCE(address,'')='' THEN ? ELSE address END,
                    interest_items   = CASE WHEN COALESCE(interest_items,'')='' THEN ? ELSE interest_items END,
                    existing_supplier= CASE WHEN COALESCE(existing_supplier,'')='' THEN ? ELSE existing_supplier END,
                    sales_status     = CASE WHEN COALESCE(sales_status,'영업중')='영업중' THEN ? ELSE sales_status END,
                    sales_notes      = CASE WHEN COALESCE(sales_notes,'')='' THEN ? ELSE sales_notes END,
                    default_region   = CASE WHEN COALESCE(default_region,'')='' THEN ? ELSE default_region END
                WHERE id=?
            """, (sc["writer"] or "", sc["address"] or "", sc["interest_items"] or "",
                  sc["existing_supplier"] or "", sc["status"] or "영업중",
                  sc["notes"] or "", sc["region"] or "", cid))
        else:
            cur = conn.execute("""
                INSERT OR IGNORE INTO customers
                    (name, default_region, writer, address, interest_items,
                     existing_supplier, sales_status, sales_notes, active, created_at)
                VALUES (?,?,?,?,?,?,?,?,1,?)
            """, (name, sc["region"] or "", sc["writer"] or "", sc["address"] or "",
                  sc["interest_items"] or "", sc["existing_supplier"] or "",
                  sc["status"] or "영업중", sc["notes"] or "",
                  sc["created_at"] or now))
            cid = cur.lastrowid
        if cid:
            conn.execute(
                "UPDATE sales_visits SET customer_id=? WHERE customer_id=?",
                (cid, sc["id"])
            )
    conn.commit()


def _upgrade_trading_customers(conn):
    """생산 이력이 있는 업체는 sales_status를 거래중으로 (아직 영업중인 경우만)"""
    conn.execute("""
        UPDATE customers SET sales_status='거래중'
        WHERE (sales_status='영업중' OR sales_status='' OR sales_status IS NULL)
          AND name IN (SELECT DISTINCT customer FROM schedules WHERE customer IS NOT NULL AND customer != '')
    """)
    conn.execute("""
        UPDATE sales_leads SET sales_status='거래중'
        WHERE (sales_status='영업중' OR sales_status='' OR sales_status IS NULL)
          AND linked_customer_id IN (
              SELECT c.id FROM customers c
              WHERE c.name IN (SELECT DISTINCT customer FROM schedules WHERE customer IS NOT NULL AND customer != '')
          )
    """)
    conn.commit()


def _migrate_all_customers_to_leads(conn):
    """customers 테이블 전체를 sales_leads로 이전 (아직 연결 안 된 업체만)"""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    all_customers = conn.execute("SELECT * FROM customers WHERE active=1").fetchall()
    for c in all_customers:
        existing = conn.execute(
            "SELECT id FROM sales_leads WHERE linked_customer_id=?", (c["id"],)
        ).fetchone()
        if existing:
            lid = existing["id"]
        else:
            cur = conn.execute("""
                INSERT INTO sales_leads (writer, name, phone, region, address, interest_items,
                    existing_supplier, sales_status, notes, linked_customer_id, active, created_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,1,?)
            """, (c["writer"] or "", c["name"] or "", c["contact_phone"] or "",
                  c["default_region"] or "", c["address"] or "",
                  c["interest_items"] or "", c["existing_supplier"] or "",
                  c["sales_status"] or "영업중", c["sales_notes"] or "",
                  c["id"], c["created_at"] or now))
            lid = cur.lastrowid
        conn.execute(
            "UPDATE sales_visits SET lead_id=? WHERE customer_id=? AND (lead_id IS NULL OR lead_id=0)",
            (lid, c["id"])
        )
    conn.commit()


def _migrate_sales_visits_to_leads(conn):
    """기존 customer_id 기반 방문기록을 sales_leads 기반으로 일회성 이전"""
    orphans = conn.execute(
        "SELECT DISTINCT customer_id FROM sales_visits WHERE customer_id IS NOT NULL AND (lead_id IS NULL OR lead_id=0)"
    ).fetchall()
    if not orphans:
        return
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for row in orphans:
        cid = row[0]
        existing = conn.execute("SELECT id FROM sales_leads WHERE linked_customer_id=?", (cid,)).fetchone()
        if existing:
            lid = existing[0]
        else:
            c = conn.execute("SELECT * FROM customers WHERE id=?", (cid,)).fetchone()
            if not c:
                continue
            cur = conn.execute("""
                INSERT INTO sales_leads (writer, name, phone, region, address, interest_items,
                    existing_supplier, sales_status, notes, linked_customer_id, active, created_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,1,?)
            """, (c["writer"] or "", c["name"] or "", c["contact_phone"] or "",
                  c["default_region"] or "", c["address"] or "",
                  c["interest_items"] or "", c["existing_supplier"] or "",
                  c["sales_status"] or "영업중", c["sales_notes"] or "",
                  cid, c["created_at"] or now))
            lid = cur.lastrowid
        conn.execute(
            "UPDATE sales_visits SET lead_id=? WHERE customer_id=? AND (lead_id IS NULL OR lead_id=0)",
            (lid, cid)
        )


def init_db():
    conn = get_conn()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE,
            password_hash TEXT,
            name TEXT DEFAULT '',
            role TEXT DEFAULT '일반',
            permissions TEXT DEFAULT '',
            active INTEGER DEFAULT 1,
            last_login TEXT DEFAULT '',
            created_at TEXT,
            updated_at TEXT
        )
    """)
    # 사용자 그룹: 확인요청 알림을 개인 또는 그룹(사무실/생산팀 등)으로 보낼 때 사용
    if not column_exists(conn, "users", "user_group"):
        conn.execute("ALTER TABLE users ADD COLUMN user_group TEXT DEFAULT '사무실'")

    # 확인요청 대상 그룹 관리. 총괄은 상위 그룹으로 모든 그룹 요청을 볼 수 있다.
    conn.execute("""
        CREATE TABLE IF NOT EXISTS check_groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_name TEXT UNIQUE,
            memo TEXT DEFAULT '',
            sort_order INTEGER DEFAULT 0,
            active INTEGER DEFAULT 1,
            created_at TEXT,
            updated_at TEXT
        )
    """)

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    admin = conn.execute("SELECT id FROM users WHERE username=?", ("관리자",)).fetchone()
    if not admin:
        conn.execute("""
            INSERT INTO users(username, password_hash, name, user_group, role, permissions, active, created_at, updated_at)
            VALUES (?, ?, ?, '관리자', '관리자', ?, 1, ?, ?)
        """, ("관리자", generate_password_hash("1234"), "관리자", permissions_to_db({}, "관리자"), now, now))
    conn.execute("""
        CREATE TABLE IF NOT EXISTS schedules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_date TEXT,
            due_date TEXT,
            customer TEXT,
            site_name TEXT,
            product_group TEXT,
            model TEXT,
            qty INTEGER DEFAULT 1,
            status TEXT DEFAULT '접수',
            memo TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    """)

    for col, ddl in [
        ("delivery_type", "ALTER TABLE schedules ADD COLUMN delivery_type TEXT DEFAULT '납품'"),
        ("delivery_date", "ALTER TABLE schedules ADD COLUMN delivery_date TEXT DEFAULT ''"),
        ("calendar_status", "ALTER TABLE schedules ADD COLUMN calendar_status TEXT DEFAULT '예정'"),
        ("completion_label", "ALTER TABLE schedules ADD COLUMN completion_label TEXT DEFAULT ''"),
        ("completion_memo", "ALTER TABLE schedules ADD COLUMN completion_memo TEXT DEFAULT ''"),
        ("completed_at", "ALTER TABLE schedules ADD COLUMN completed_at TEXT DEFAULT ''"),
        ("detail_content", "ALTER TABLE schedules ADD COLUMN detail_content TEXT DEFAULT ''"),
        ("customer_contact", "ALTER TABLE schedules ADD COLUMN customer_contact TEXT DEFAULT ''")
    ]:
        if not column_exists(conn, "schedules", col):
            conn.execute(ddl)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE,
            created_at TEXT
        )
    """)

    # 예전 DB에 customers 테이블이 이미 있으면 새 컬럼이 자동으로 안 생기므로 보강
    for col, ddl in [
        ("memo",              "ALTER TABLE customers ADD COLUMN memo TEXT DEFAULT ''"),
        ("payment_note",      "ALTER TABLE customers ADD COLUMN payment_note TEXT DEFAULT ''"),
        ("contact_name",      "ALTER TABLE customers ADD COLUMN contact_name TEXT DEFAULT ''"),
        ("contact_phone",     "ALTER TABLE customers ADD COLUMN contact_phone TEXT DEFAULT ''"),
        ("contact_memo",      "ALTER TABLE customers ADD COLUMN contact_memo TEXT DEFAULT ''"),
        ("active",            "ALTER TABLE customers ADD COLUMN active INTEGER DEFAULT 1"),
        ("default_region",    "ALTER TABLE customers ADD COLUMN default_region TEXT DEFAULT ''"),
        # 영업일지 통합 컬럼
        ("writer",            "ALTER TABLE customers ADD COLUMN writer TEXT DEFAULT ''"),
        ("address",           "ALTER TABLE customers ADD COLUMN address TEXT DEFAULT ''"),
        ("interest_items",    "ALTER TABLE customers ADD COLUMN interest_items TEXT DEFAULT ''"),
        ("existing_supplier", "ALTER TABLE customers ADD COLUMN existing_supplier TEXT DEFAULT ''"),
        ("sales_status",      "ALTER TABLE customers ADD COLUMN sales_status TEXT DEFAULT '영업중'"),
        ("sales_notes",       "ALTER TABLE customers ADD COLUMN sales_notes TEXT DEFAULT ''"),
    ]:
        if not column_exists(conn, "customers", col):
            conn.execute(ddl)

    # sales_customers → customers 일회성 마이그레이션
    _migrate_sales_customers(conn)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS calendar_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            schedule_id INTEGER,
            event_date TEXT,
            event_type TEXT,
            title TEXT,
            status TEXT DEFAULT '예정',
            created_at TEXT,
            updated_at TEXT
        )
    """)

    # 예전 DB에 calendar_events 테이블이 이미 있으면 새 컬럼이 자동으로 안 생기므로 보강
    for col, ddl in [
        ("status", "ALTER TABLE calendar_events ADD COLUMN status TEXT DEFAULT '예정'"),
        ("updated_at", "ALTER TABLE calendar_events ADD COLUMN updated_at TEXT DEFAULT ''")
    ]:
        if not column_exists(conn, "calendar_events", col):
            conn.execute(ddl)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS schedule_photos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            schedule_id INTEGER,
            filename TEXT,
            original_name TEXT,
            memo TEXT,
            created_at TEXT
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS delivery_people (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT UNIQUE,
            phone TEXT DEFAULT '',
            vehicle_no TEXT DEFAULT '',
            memo TEXT DEFAULT '',
            created_at TEXT,
            updated_at TEXT
        )
    """)

    for col, ddl in [
        ("delivery_person_id", "ALTER TABLE calendar_events ADD COLUMN delivery_person_id INTEGER"),
        ("delivery_person_name", "ALTER TABLE calendar_events ADD COLUMN delivery_person_name TEXT DEFAULT ''"),
        ("delivery_phone", "ALTER TABLE calendar_events ADD COLUMN delivery_phone TEXT DEFAULT ''"),
        ("delivery_vehicle_no", "ALTER TABLE calendar_events ADD COLUMN delivery_vehicle_no TEXT DEFAULT ''"),
        ("delivery_status", "ALTER TABLE calendar_events ADD COLUMN delivery_status TEXT DEFAULT '예정'"),
        ("delivery_start", "ALTER TABLE calendar_events ADD COLUMN delivery_start TEXT DEFAULT ''"),
        ("delivery_end", "ALTER TABLE calendar_events ADD COLUMN delivery_end TEXT DEFAULT ''"),
        ("delivery_memo", "ALTER TABLE calendar_events ADD COLUMN delivery_memo TEXT DEFAULT ''"),
        ("manual_customer", "ALTER TABLE calendar_events ADD COLUMN manual_customer TEXT DEFAULT ''"),
        ("manual_site_name", "ALTER TABLE calendar_events ADD COLUMN manual_site_name TEXT DEFAULT ''"),
        ("manual_product", "ALTER TABLE calendar_events ADD COLUMN manual_product TEXT DEFAULT ''"),
        ("manual_qty", "ALTER TABLE calendar_events ADD COLUMN manual_qty INTEGER DEFAULT 0"),
        ("manual_memo", "ALTER TABLE calendar_events ADD COLUMN manual_memo TEXT DEFAULT ''"),
    ]:
        if not column_exists(conn, "calendar_events", col):
            conn.execute(ddl)


    reset_price_tables_if_old_schema(conn)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS price_products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            series TEXT,
            model TEXT,
            image_key TEXT DEFAULT '',
            active INTEGER DEFAULT 1,
            created_at TEXT,
            updated_at TEXT,
            UNIQUE(series, model)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS price_types (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER,
            type_name TEXT,
            factory_price INTEGER DEFAULT 0,
            consumer_price INTEGER DEFAULT 0,
            sort_order INTEGER DEFAULT 0
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS price_options (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER,
            option_name TEXT,
            option_price INTEGER DEFAULT 0,
            sort_order INTEGER DEFAULT 0
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS price_notes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER,
            note TEXT,
            sort_order INTEGER DEFAULT 0
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS price_regions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            region_name TEXT,
            category TEXT DEFAULT '',
            travel_time TEXT DEFAULT '',
            region_grade TEXT DEFAULT '',
            region_price INTEGER DEFAULT 0
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS journal_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            journal_date TEXT,
            journal_type TEXT DEFAULT '상담일지',
            channel TEXT DEFAULT '',
            customer TEXT DEFAULT '',
            contact_name TEXT DEFAULT '',
            contact_phone TEXT DEFAULT '',
            title TEXT DEFAULT '',
            content TEXT DEFAULT '',
            follow_up TEXT DEFAULT '',
            status TEXT DEFAULT '진행중',
            writer TEXT DEFAULT '',
            created_at TEXT,
            updated_at TEXT
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS sales_customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            writer TEXT DEFAULT '',
            customer TEXT DEFAULT '',
            region TEXT DEFAULT '',
            address TEXT DEFAULT '',
            interest_items TEXT DEFAULT '',
            existing_supplier TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            status TEXT DEFAULT '영업중',
            created_at TEXT DEFAULT '',
            updated_at TEXT DEFAULT ''
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS sales_visits (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER,
            visit_date TEXT DEFAULT '',
            visitor TEXT DEFAULT '',
            sales_items TEXT DEFAULT '',
            content TEXT DEFAULT '',
            created_at TEXT DEFAULT ''
        )
    """)
    if not column_exists(conn, "sales_visits", "visitor"):
        conn.execute("ALTER TABLE sales_visits ADD COLUMN visitor TEXT DEFAULT ''")

    conn.execute("""
        CREATE TABLE IF NOT EXISTS sales_leads (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            writer TEXT DEFAULT '',
            name TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            region TEXT DEFAULT '',
            address TEXT DEFAULT '',
            interest_items TEXT DEFAULT '',
            existing_supplier TEXT DEFAULT '',
            sales_status TEXT DEFAULT '영업중',
            notes TEXT DEFAULT '',
            linked_customer_id INTEGER DEFAULT NULL,
            active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT ''
        )
    """)
    if not column_exists(conn, "sales_visits", "lead_id"):
        conn.execute("ALTER TABLE sales_visits ADD COLUMN lead_id INTEGER")
    _migrate_sales_visits_to_leads(conn)
    _migrate_all_customers_to_leads(conn)
    _upgrade_trading_customers(conn)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS consult_clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            writer TEXT DEFAULT '',
            name TEXT DEFAULT '',
            first_consult_date TEXT DEFAULT '',
            channel TEXT DEFAULT '경로확인',
            region TEXT DEFAULT '',
            address TEXT DEFAULT '',
            interest_items TEXT DEFAULT '',
            notes TEXT DEFAULT '',
            linked_customer_id INTEGER DEFAULT NULL,
            active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT ''
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS consult_visits (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id INTEGER,
            visit_date TEXT DEFAULT '',
            sales_items TEXT DEFAULT '',
            content TEXT DEFAULT '',
            consult_type TEXT DEFAULT '',
            delivery_type TEXT DEFAULT '',
            created_at TEXT DEFAULT ''
        )
    """)

    for col, ddl in [
        ("phone",         "ALTER TABLE consult_clients ADD COLUMN phone TEXT DEFAULT ''"),
    ]:
        if not column_exists(conn, "consult_clients", col):
            conn.execute(ddl)

    for col, ddl in [
        ("consult_type",  "ALTER TABLE consult_visits ADD COLUMN consult_type TEXT DEFAULT ''"),
        ("delivery_type", "ALTER TABLE consult_visits ADD COLUMN delivery_type TEXT DEFAULT ''"),
    ]:
        if not column_exists(conn, "consult_visits", col):
            conn.execute(ddl)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS journal_photos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            entity_type TEXT DEFAULT '',
            entity_id INTEGER DEFAULT 0,
            filename TEXT DEFAULT '',
            original_name TEXT DEFAULT '',
            uploaded_at TEXT DEFAULT ''
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS work_orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_group TEXT DEFAULT '',
            template_type TEXT DEFAULT '비단열',
            title TEXT DEFAULT '',
            customer TEXT DEFAULT '',
            site_name TEXT DEFAULT '',
            model TEXT DEFAULT '',
            width TEXT DEFAULT '',
            height TEXT DEFAULT '',
            qty INTEGER DEFAULT 1,
            delivery_type TEXT DEFAULT '',
            schedule_date TEXT DEFAULT '',
            release_date TEXT DEFAULT '',
            status TEXT DEFAULT '작성중',
            memo TEXT DEFAULT '',
            data_json TEXT DEFAULT '',
            writer TEXT DEFAULT '',
            created_at TEXT,
            updated_at TEXT
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS check_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            schedule_id INTEGER,
            from_user_id INTEGER,
            from_user_name TEXT DEFAULT '',
            to_user_id INTEGER,
            to_user_name TEXT DEFAULT '',
            to_group TEXT DEFAULT '',
            message TEXT DEFAULT '',
            status TEXT DEFAULT '미확인',
            created_at TEXT,
            completed_at TEXT DEFAULT '',
            completed_by TEXT DEFAULT ''
        )
    """)

    import_price_csv_if_empty(conn)

    conn.execute("DELETE FROM calendar_events WHERE event_type='픽업예정'")
    sync_missing_calendar_events(conn)
    
    conn.execute("""
        CREATE TABLE IF NOT EXISTS as_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            status TEXT DEFAULT '접수',
            receipt_date TEXT DEFAULT '',
            request_date TEXT DEFAULT '',
            receiver TEXT DEFAULT '',
            customer TEXT DEFAULT '',
            company_phone TEXT DEFAULT '',
            consumer_phone TEXT DEFAULT '',
            region TEXT DEFAULT '',
            region_group TEXT DEFAULT '',
            address TEXT DEFAULT '',
            product_group TEXT DEFAULT '',
            request_content TEXT DEFAULT '',
            memo TEXT DEFAULT '',
            scheduled_date TEXT DEFAULT '',
            assigned_to TEXT DEFAULT '',
            needed_parts TEXT DEFAULT '',
            defect_cause TEXT DEFAULT '',
            process_content TEXT DEFAULT '',
            improvement TEXT DEFAULT '',
            completed_date TEXT DEFAULT '',
            completed_by TEXT DEFAULT '',
            hold_reason TEXT DEFAULT '',
            created_at TEXT,
            updated_at TEXT
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS as_photos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            as_id INTEGER,
            photo_type TEXT,
            filename TEXT,
            original_name TEXT,
            memo TEXT DEFAULT '',
            created_at TEXT
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS as_ref_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_group TEXT DEFAULT '공통',
            title TEXT DEFAULT '',
            content TEXT DEFAULT '',
            url TEXT DEFAULT '',
            url_type TEXT DEFAULT '',
            sort_order INTEGER DEFAULT 0,
            created_by TEXT DEFAULT '',
            created_at TEXT,
            updated_at TEXT
        )
    """)

    try:
        conn.execute("ALTER TABLE as_requests ADD COLUMN region_group TEXT DEFAULT ''")
    except Exception:
        pass

    
    try:
        conn.execute("ALTER TABLE check_requests ADD COLUMN as_id INTEGER")
    except Exception:
        pass
    try:
        conn.execute("ALTER TABLE check_requests ADD COLUMN source_type TEXT DEFAULT 'schedule'")
    except Exception:
        pass

    conn.execute("""
        CREATE TABLE IF NOT EXISTS regions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            created_at TEXT
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS mail_settings (
            id INTEGER PRIMARY KEY,
            smtp_host TEXT DEFAULT 'smtp.naver.com',
            smtp_port INTEGER DEFAULT 465,
            smtp_user TEXT DEFAULT '',
            smtp_password TEXT DEFAULT '',
            from_name TEXT DEFAULT '',
            updated_at TEXT DEFAULT ''
        )
    """)
    conn.execute("""
        INSERT OR IGNORE INTO mail_settings (id) VALUES (1)
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS mail_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sent_at TEXT,
            sent_by TEXT,
            to_email TEXT,
            cc_email TEXT DEFAULT '',
            subject TEXT,
            body TEXT DEFAULT '',
            attachments TEXT DEFAULT '',
            pdf_path TEXT DEFAULT '',
            status TEXT DEFAULT '성공',
            error_msg TEXT DEFAULT ''
        )
    """)
    try:
        conn.execute("ALTER TABLE mail_log ADD COLUMN pdf_path TEXT DEFAULT ''")
    except Exception:
        pass

    import os as _os
    _docs_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "documents", "납품확인서")
    _os.makedirs(_docs_dir, exist_ok=True)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS vehicle_ledger (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company TEXT DEFAULT '',
            vehicle_type TEXT DEFAULT '',
            unit_no TEXT DEFAULT '',
            vehicle_no TEXT DEFAULT '',
            hipass_digits TEXT DEFAULT '',
            fuel_type TEXT DEFAULT '',
            reg_date TEXT DEFAULT '',
            owner TEXT DEFAULT '',
            manager TEXT DEFAULT '',
            hipass_card TEXT DEFAULT '',
            vehicle_class TEXT DEFAULT '',
            mileage_json TEXT DEFAULT '[]',
            engine_oil TEXT DEFAULT '',
            tire TEXT DEFAULT '',
            repair TEXT DEFAULT '',
            urea TEXT DEFAULT '',
            violation TEXT DEFAULT '',
            car_insurance TEXT DEFAULT '',
            driver_insurance TEXT DEFAULT '',
            inspection TEXT DEFAULT '',
            tax TEXT DEFAULT '',
            memo TEXT DEFAULT '',
            driver_age TEXT DEFAULT '',
            photo TEXT DEFAULT '',
            created_at TEXT DEFAULT '',
            updated_at TEXT DEFAULT ''
        )
    """)
    for _col, _ddl in [
        ("vehicle_class",    "ALTER TABLE vehicle_ledger ADD COLUMN vehicle_class TEXT DEFAULT ''"),
        ("mileage_json",     "ALTER TABLE vehicle_ledger ADD COLUMN mileage_json TEXT DEFAULT '[]'"),
        ("engine_oil",       "ALTER TABLE vehicle_ledger ADD COLUMN engine_oil TEXT DEFAULT ''"),
        ("tire",             "ALTER TABLE vehicle_ledger ADD COLUMN tire TEXT DEFAULT ''"),
        ("repair",           "ALTER TABLE vehicle_ledger ADD COLUMN repair TEXT DEFAULT ''"),
        ("urea",             "ALTER TABLE vehicle_ledger ADD COLUMN urea TEXT DEFAULT ''"),
        ("violation",        "ALTER TABLE vehicle_ledger ADD COLUMN violation TEXT DEFAULT ''"),
        ("car_insurance",    "ALTER TABLE vehicle_ledger ADD COLUMN car_insurance TEXT DEFAULT ''"),
        ("driver_insurance", "ALTER TABLE vehicle_ledger ADD COLUMN driver_insurance TEXT DEFAULT ''"),
        ("inspection",       "ALTER TABLE vehicle_ledger ADD COLUMN inspection TEXT DEFAULT ''"),
        ("tax",              "ALTER TABLE vehicle_ledger ADD COLUMN tax TEXT DEFAULT ''"),
        ("status",           "ALTER TABLE vehicle_ledger ADD COLUMN status TEXT DEFAULT '운행'"),
        ("driver_age",       "ALTER TABLE vehicle_ledger ADD COLUMN driver_age TEXT DEFAULT ''"),
    ]:
        if not column_exists(conn, "vehicle_ledger", _col):
            conn.execute(_ddl)
    # 기존 데이터 중 차량번호에 매각/판매 포함된 것은 status='매각'으로 자동 분류
    conn.execute("""
        CREATE TABLE IF NOT EXISTS equipment_ledger (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company TEXT DEFAULT '',
            name TEXT DEFAULT '',
            manage_no TEXT DEFAULT '',
            mfg_no TEXT DEFAULT '',
            type TEXT DEFAULT '',
            grade TEXT DEFAULT '',
            category TEXT DEFAULT '',
            maker TEXT DEFAULT '',
            mfg_date TEXT DEFAULT '',
            voltage TEXT DEFAULT '',
            supplier TEXT DEFAULT '',
            install_date TEXT DEFAULT '',
            location TEXT DEFAULT '',
            parts_json TEXT DEFAULT '[]',
            repair_json TEXT DEFAULT '[]',
            memo TEXT DEFAULT '',
            photo TEXT DEFAULT '',
            maker_contact TEXT DEFAULT '',
            maker_note TEXT DEFAULT '',
            supplier_contact TEXT DEFAULT '',
            supplier_note TEXT DEFAULT '',
            created_at TEXT DEFAULT '',
            updated_at TEXT DEFAULT ''
        )
    """)

    for _col, _ddl in [
        ("maker_contact",    "ALTER TABLE equipment_ledger ADD COLUMN maker_contact TEXT DEFAULT ''"),
        ("maker_note",       "ALTER TABLE equipment_ledger ADD COLUMN maker_note TEXT DEFAULT ''"),
        ("supplier_contact", "ALTER TABLE equipment_ledger ADD COLUMN supplier_contact TEXT DEFAULT ''"),
        ("supplier_note",    "ALTER TABLE equipment_ledger ADD COLUMN supplier_note TEXT DEFAULT ''"),
    ]:
        if not column_exists(conn, "equipment_ledger", _col):
            conn.execute(_ddl)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS vehicle_docs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vehicle_id INTEGER,
            filename TEXT DEFAULT '',
            original_name TEXT DEFAULT '',
            uploaded_at TEXT DEFAULT ''
        )
    """)

    conn.execute("""
        UPDATE vehicle_ledger SET status='매각'
        WHERE (status IS NULL OR status='운행')
          AND (vehicle_no LIKE '%매각%' OR vehicle_no LIKE '%판매%'
               OR hipass_card LIKE '%매각%')
    """)

    # journal/ledger 권한 분리 마이그레이션
    users = conn.execute("SELECT id, permissions FROM users").fetchall()
    for u in users:
        try:
            perms = json.loads(u["permissions"] or "{}")
        except Exception:
            perms = {}
        changed = False
        if "journal" in perms:
            level = perms.pop("journal")
            for k in ["journal_sales", "journal_consult", "journal_measure", "journal_install"]:
                perms.setdefault(k, level)
            changed = True
        if "ledger" in perms:
            level = perms.pop("ledger")
            perms.setdefault("ledger_vehicle", level)
            perms.setdefault("ledger_machine", level)
            changed = True
        if changed:
            conn.execute("UPDATE users SET permissions=? WHERE id=?",
                         (json.dumps(perms, ensure_ascii=False), u["id"]))

    # 조직도 테이블
    conn.execute("""
        CREATE TABLE IF NOT EXISTS org_departments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            label TEXT NOT NULL,
            dept_type TEXT DEFAULT 'regular',
            sort_order INTEGER DEFAULT 0,
            created_at TEXT DEFAULT '',
            updated_at TEXT DEFAULT ''
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS org_teams (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dept_id INTEGER NOT NULL,
            label TEXT NOT NULL,
            sort_order INTEGER DEFAULT 0,
            created_at TEXT DEFAULT '',
            updated_at TEXT DEFAULT ''
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS org_members (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dept_id INTEGER NOT NULL,
            team_id INTEGER,
            name TEXT NOT NULL DEFAULT '',
            rank TEXT DEFAULT '',
            phone TEXT DEFAULT '',
            team_tag TEXT DEFAULT '',
            photo TEXT DEFAULT '',
            is_head INTEGER DEFAULT 0,
            sort_order INTEGER DEFAULT 0,
            created_at TEXT DEFAULT '',
            updated_at TEXT DEFAULT ''
        )
    """)
    if not column_exists(conn, "org_members", "hire_date"):
        conn.execute("ALTER TABLE org_members ADD COLUMN hire_date TEXT DEFAULT ''")
    if not column_exists(conn, "org_members", "is_team_leader"):
        conn.execute("ALTER TABLE org_members ADD COLUMN is_team_leader INTEGER DEFAULT 0")

    conn.execute("""
        CREATE TABLE IF NOT EXISTS hr_leave_requests (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            member_id INTEGER NOT NULL,
            leave_type TEXT DEFAULT '연차',
            start_date TEXT NOT NULL,
            end_date TEXT NOT NULL,
            days REAL DEFAULT 1,
            reason TEXT DEFAULT '',
            status TEXT DEFAULT '대기',
            applied_by TEXT DEFAULT '',
            processed_at TEXT DEFAULT '',
            processed_by TEXT DEFAULT '',
            memo TEXT DEFAULT '',
            created_at TEXT DEFAULT '',
            updated_at TEXT DEFAULT ''
        )
    """)

    for col, default in [
        ("approver1_name",   "''"), ("approver1_status", "'신청'"),
        ("approver2_name",   "''"), ("approver2_status", "'신청'"),
    ]:
        if not column_exists(conn, "hr_leave_requests", col):
            conn.execute(f"ALTER TABLE hr_leave_requests ADD COLUMN {col} TEXT DEFAULT {default}")

    if not column_exists(conn, "check_requests", "leave_request_id"):
        conn.execute("ALTER TABLE check_requests ADD COLUMN leave_request_id INTEGER")

    if not conn.execute("SELECT id FROM org_departments LIMIT 1").fetchone():
        _migrate_org_from_const(conn)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS push_subscriptions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            endpoint TEXT NOT NULL,
            p256dh TEXT NOT NULL,
            auth TEXT NOT NULL,
            created_at TEXT DEFAULT ''
        )
    """)
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_push_endpoint ON push_subscriptions(endpoint)")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS construction_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            type TEXT DEFAULT '실측',
            status TEXT DEFAULT '예정',
            company TEXT DEFAULT '',
            scheduled_date TEXT DEFAULT '',
            customer TEXT DEFAULT '',
            address TEXT DEFAULT '',
            manager TEXT DEFAULT '',
            content TEXT DEFAULT '',
            memo TEXT DEFAULT '',
            created_by TEXT DEFAULT '',
            created_at TEXT DEFAULT '',
            updated_at TEXT DEFAULT ''
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS construction_docs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT DEFAULT '',
            filename TEXT DEFAULT '',
            original_name TEXT DEFAULT '',
            uploaded_by TEXT DEFAULT '',
            created_at TEXT DEFAULT ''
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS att_saved (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT DEFAULT '',
            month TEXT DEFAULT '',
            data TEXT DEFAULT '',
            created_at TEXT DEFAULT ''
        )
    """)

    conn.commit()

    # att_saved 마이그레이션: updated_at 컬럼 추가
    try:
        conn.execute("ALTER TABLE att_saved ADD COLUMN updated_at TEXT DEFAULT ''")
        conn.commit()
    except Exception:
        pass

    conn.close()


def _migrate_org_from_const(conn):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for i, dept in enumerate(ORG_DATA):
        dept_type = "special" if "members" in dept and "teams" not in dept else "regular"
        conn.execute(
            "INSERT INTO org_departments (label, dept_type, sort_order, created_at, updated_at) VALUES (?,?,?,?,?)",
            (dept["label"], dept_type, i * 10, now, now)
        )
        dept_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        if dept.get("head"):
            h = dept["head"]
            conn.execute(
                "INSERT INTO org_members (dept_id, team_id, name, rank, phone, photo, is_head, sort_order, created_at, updated_at) VALUES (?,NULL,?,?,?,?,1,0,?,?)",
                (dept_id, h["name"], h["rank"], h.get("phone", ""), h.get("photo", ""), now, now)
            )
        for j, m in enumerate(dept.get("members", [])):
            conn.execute(
                "INSERT INTO org_members (dept_id, team_id, name, rank, phone, team_tag, photo, is_head, sort_order, created_at, updated_at) VALUES (?,NULL,?,?,?,?,?,0,?,?,?)",
                (dept_id, m["name"], m["rank"], m.get("phone", ""), m.get("team", ""), m.get("photo", ""), j * 10, now, now)
            )
        for j, team in enumerate(dept.get("teams", [])):
            conn.execute(
                "INSERT INTO org_teams (dept_id, label, sort_order, created_at, updated_at) VALUES (?,?,?,?,?)",
                (dept_id, team["label"], j * 10, now, now)
            )
            team_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
            for k, m in enumerate(team.get("members", [])):
                conn.execute(
                    "INSERT INTO org_members (dept_id, team_id, name, rank, phone, photo, is_head, sort_order, created_at, updated_at) VALUES (?,?,?,?,?,?,0,?,?,?)",
                    (dept_id, team_id, m["name"], m["rank"], m.get("phone", ""), m.get("photo", ""), k * 10, now, now)
                )


def upsert_calendar_event(conn, schedule_id, data):
    delivery_type = data.get("delivery_type") or "납품"

    # 직접픽업은 업체가 가져가는 건이라 납품/시공 달력에서는 제외한다.
    # 기존에 생성된 픽업 일정이 있으면 삭제해서 달력에 보이지 않게 한다.
    if delivery_type == "직접픽업":
        conn.execute("DELETE FROM calendar_events WHERE schedule_id=?", (schedule_id,))
        return

    event_type_map = {
        "시공": "시공예정",
        "납품": "납품예정",
        "화물": "화물예정",
        "용차": "용차예정",
        "택배": "택배예정",
    }
    event_type = event_type_map.get(delivery_type, "납품예정")

    status = data.get("status") or "접수"
    calendar_status = "출고가능" if status == "완료" else "예정"
    title = f"{data.get('customer','')} / {data.get('product_group','')} {data.get('model','')} / {data.get('qty',1)}개"
    event_date = data.get("delivery_date") or data.get("due_date", "")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    exists = conn.execute("SELECT id FROM calendar_events WHERE schedule_id=?", (schedule_id,)).fetchone()
    if exists:
        conn.execute("""
            UPDATE calendar_events SET event_date=?, event_type=?, title=?, status=?, updated_at=?
            WHERE schedule_id=?
        """, (event_date, event_type, title, calendar_status, now, schedule_id))
    else:
        conn.execute("""
            INSERT INTO calendar_events(schedule_id, event_date, event_type, title, status, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (schedule_id, event_date, event_type, title, calendar_status, now, now))


def sync_missing_calendar_events(conn):
    rows = conn.execute("SELECT * FROM schedules").fetchall()
    for row in rows:
        upsert_calendar_event(conn, row["id"], row_to_dict(row))


def row_to_dict(r):
    return dict(r)

def safe_filename_part(value, default="미입력"):
    text = str(value or "").strip() or default
    text = re.sub(r'[\\/:*?"<>|]+', '_', text)
    text = re.sub(r'\s+', '', text)
    return text[:50] or default

def make_photo_filename(item_id, original_filename, row=None):
    # 저장 파일명: 연도-날짜-업체명-제품-라벨번호 순서
    # 실제 파일은 용량 절감을 위해 jpg로 변환 저장한다.
    now_date = datetime.now().strftime('%Y-%m-%d')
    now_time = datetime.now().strftime('%H%M%S')
    customer = safe_filename_part(row["customer"] if row and "customer" in row.keys() else "거래처")
    product = safe_filename_part(row["product_group"] if row and "product_group" in row.keys() else "제품")
    model = safe_filename_part(row["model"] if row and "model" in row.keys() else "")
    label = safe_filename_part(row["completion_label"] if row and "completion_label" in row.keys() else "라벨없음")
    product_name = product if not model or model == "미입력" else f"{product}-{model}"
    return f"{now_date}_{customer}_{product_name}_{label}_{now_time}_{uuid.uuid4().hex[:6]}.jpg"

def save_resized_photo(file_storage, save_path, target_bytes=1024*1024, max_side=1600):
    """휴대폰 원본 사진을 약 1MB 이하가 되도록 리사이즈/압축 저장한다."""
    img = Image.open(file_storage.stream)
    img = ImageOps.exif_transpose(img)

    if img.mode not in ("RGB", "L"):
        img = img.convert("RGB")
    elif img.mode == "L":
        img = img.convert("RGB")

    w, h = img.size
    if max(w, h) > max_side:
        scale = max_side / max(w, h)
        img = img.resize((int(w * scale), int(h * scale)), Image.LANCZOS)

    # 우선 85부터 저장 후 1MB를 넘으면 품질을 낮춰 재저장한다.
    for quality in [85, 78, 72, 66, 60, 55, 50]:
        img.save(save_path, format="JPEG", quality=quality, optimize=True)
        if save_path.stat().st_size <= target_bytes:
            return

    # 그래도 크면 한번 더 축소 후 저장한다.
    w, h = img.size
    img = img.resize((int(w * 0.8), int(h * 0.8)), Image.LANCZOS)
    img.save(save_path, format="JPEG", quality=50, optimize=True)


def clean_value(v):
    if v is None:
        return ""
    s = str(v).strip()
    if not s or s.lower() == "nan":
        return ""
    return s

def price_to_int(v):
    s = re.sub(r"[^0-9.-]", "", clean_value(v))
    if not s:
        return 0
    try:
        return int(float(s))
    except Exception:
        return 0

def find_price_csv_dir():
    candidates = [
        APP_DIR.parent / "price" / "price_csv",
        APP_DIR.parent / "price_csv",
        APP_DIR / "price_csv",
    ]
    for p in candidates:
        if p.exists() and any(p.glob("*.csv")):
            return p
    return None

def import_price_csv_if_empty(conn):
    existing = conn.execute("SELECT COUNT(*) AS cnt FROM price_products").fetchone()["cnt"]
    if existing:
        return
    csv_dir = find_price_csv_dir()
    if not csv_dir:
        return

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    series_order = ["엔토브", "로이도어", "리젠도어", "나인도어", "클래식도어", "방화문", "대문"]
    for series in series_order:
        file_path = csv_dir / f"{series}.csv"
        if not file_path.exists():
            continue
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                product_series = clean_value(row.get("제품군")) or series
                model = clean_value(row.get("모델명"))
                if not model:
                    continue
                image_key = clean_value(row.get("이미지키")) or model
                conn.execute("""
                    INSERT OR IGNORE INTO price_products(series, model, image_key, active, created_at, updated_at)
                    VALUES (?, ?, ?, 1, ?, ?)
                """, (product_series, model, image_key, now, now))
                p_row = conn.execute("SELECT id FROM price_products WHERE series=? AND model=?", (product_series, model)).fetchone()
                if not p_row:
                    continue
                product_id = p_row["id"]
                conn.execute("DELETE FROM price_types WHERE product_id=?", (product_id,))
                conn.execute("DELETE FROM price_options WHERE product_id=?", (product_id,))
                conn.execute("DELETE FROM price_notes WHERE product_id=?", (product_id,))
                for i in range(1, 31):
                    type_name = clean_value(row.get(f"타입{i}명"))
                    if type_name:
                        conn.execute("""
                            INSERT INTO price_types(product_id, type_name, factory_price, consumer_price, sort_order)
                            VALUES (?, ?, ?, ?, ?)
                        """, (product_id, type_name, price_to_int(row.get(f"타입{i}_공장도")), price_to_int(row.get(f"타입{i}_소비자")), i))
                    opt_name = clean_value(row.get(f"선택사양{i}"))
                    if opt_name:
                        conn.execute("""
                            INSERT INTO price_options(product_id, option_name, option_price, sort_order)
                            VALUES (?, ?, ?, ?)
                        """, (product_id, opt_name, price_to_int(row.get(f"선택사양{i}금액")), i))
                    note = clean_value(row.get(f"비고{i}"))
                    if note:
                        conn.execute("""
                            INSERT INTO price_notes(product_id, note, sort_order)
                            VALUES (?, ?, ?)
                        """, (product_id, note, i))

    region_file = csv_dir / "지역구분.csv"
    if region_file.exists():
        conn.execute("DELETE FROM price_regions")
        with open(region_file, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                name = clean_value(row.get("지역명"))
                grade = clean_value(row.get("지역등급"))
                if not name and not grade:
                    continue
                conn.execute("""
                    INSERT INTO price_regions(region_name, category, travel_time, region_grade, region_price)
                    VALUES (?, ?, ?, ?, ?)
                """, (name, clean_value(row.get("대분류")), clean_value(row.get("예상소요시간")), grade, price_to_int(row.get("지역금액"))))

@app.route('/.well-known/acme-challenge/<path:token>')
def acme_challenge(token):
    challenge_dir = r"D:\Plusdoor Web\nginx-1.30.2\html\.well-known\acme-challenge"
    return send_from_directory(challenge_dir, token)


@app.route("/")
@require_perm("production")
def index():
    return render_template("index.html")

@app.route("/completed")
@require_perm("completed")
def completed_page():
    return render_template("completed.html")

@app.route("/calendar")
@require_perm("calendar")
def calendar_page():
    return render_template("calendar.html")


def generate_delivery_pdf(data):
    from fpdf import FPDF
    import os as _os

    font_dir = r'C:\Windows\Fonts'
    regular = _os.path.join(font_dir, 'malgun.ttf')
    bold    = _os.path.join(font_dir, 'malgunbd.ttf')

    pdf = FPDF('P', 'mm', 'A4')
    pdf.add_font('K', '',  regular)
    pdf.add_font('K', 'B', bold)
    pdf.set_auto_page_break(auto=False)
    pdf.add_page()

    lm, rm, tm = 18, 18, 20
    pw = 210 - lm - rm   # 174 mm

    # ── 제목 ──
    pdf.set_xy(lm, tm)
    pdf.set_font('K', 'B', 22)
    pdf.cell(pw, 14, '납  품  확  인  서', align='C')
    y = tm + 14
    pdf.set_line_width(0.6)
    pdf.line(lm, y, lm + pw, y)
    y += 7
    pdf.set_line_width(0.2)

    # ── 필드 ──
    lw, fh = 38, 8.5
    vw = pw - lw
    fields = [
        ('현  장  주  소', data.get('address', '')),
        ('현    장    명', data.get('site_name', '')),
        ('건    축    주', data.get('client', '')),
        ('납  품  일  자', data.get('delivery_date', '')),
    ]
    for label, value in fields:
        pdf.set_fill_color(248, 248, 248)
        pdf.rect(lm, y, pw, fh)
        pdf.line(lm + lw, y, lm + lw, y + fh)
        pdf.set_font('K', 'B', 10)
        pdf.set_xy(lm, y)
        pdf.cell(lw, fh, label, align='C', fill=True)
        pdf.set_font('K', '', 10)
        pdf.set_xy(lm + lw + 2, y)
        pdf.cell(vw - 2, fh, value, align='L')
        y += fh

    y += 5

    # ── 납품내역 헤더 ──
    pdf.set_fill_color(242, 242, 242)
    pdf.rect(lm, y, pw, 8, 'FD')
    pdf.set_font('K', 'B', 11)
    pdf.set_xy(lm, y)
    pdf.cell(pw, 8, '납  품  내  역', align='C')
    y += 8

    # ── 품목 컬럼 헤더 ──
    c1, c2, c3 = pw * 0.58, pw * 0.27, pw * 0.15
    pdf.set_fill_color(245, 245, 245)
    pdf.set_font('K', 'B', 10)
    for col, label in [(c1, '품     명'), (c2, '규 격 (mm)'), (c3, '수 량')]:
        x = lm if label == '품     명' else lm + c1 if label == '규 격 (mm)' else lm + c1 + c2
        pdf.rect(x, y, col, 7.5, 'FD')
        pdf.set_xy(x, y)
        pdf.cell(col, 7.5, label, align='C')
    y += 7.5

    # ── 품목 행 ──
    items = [i for i in data.get('items', []) if str(i.get('name', '')).strip() or str(i.get('spec', '')).strip()]
    pdf.set_font('K', '', 10)
    for item in items:
        pdf.rect(lm,        y, c1, 7.5)
        pdf.rect(lm + c1,   y, c2, 7.5)
        pdf.rect(lm+c1+c2,  y, c3, 7.5)
        pdf.set_xy(lm + 2, y)
        pdf.cell(c1 - 2, 7.5, str(item.get('name', '')), align='L')
        pdf.set_xy(lm + c1, y)
        pdf.cell(c2, 7.5, str(item.get('spec', '')), align='C')
        pdf.set_xy(lm + c1 + c2, y)
        pdf.cell(c3, 7.5, str(item.get('qty', '')), align='C')
        y += 7.5

    y += 10

    # ── 확인 텍스트 ──
    pdf.set_font('K', '', 11)
    pdf.set_xy(lm, y)
    pdf.cell(pw, 8, '상기 물품을 아무 이상 없이 당 현장에 납품하였음을 확인하는 바입니다.', align='C')
    y += 18

    # ── 공급자 ──
    sup_y = y
    pdf.set_font('K', 'B', 10)
    pdf.set_xy(lm, y)
    pdf.cell(pw - 38, 7, '- 공  급  자 -', align='L')
    y += 9
    pdf.set_font('K', '', 10)
    for label, value in [
        ('사 업 자 번 호', '457-81-00767'),
        ('상          호', '(주)청주금속플러스도어'),
        ('대  표  자', '윤 영 석'),
        ('주          소', '충북 청주시 서원구 남이면 가좌신송로 17'),
    ]:
        pdf.set_xy(lm, y)
        pdf.cell(36, 7, label, align='L')
        pdf.set_xy(lm + 36, y)
        pdf.cell(7, 7, ':', align='C')
        pdf.set_xy(lm + 43, y)
        pdf.cell(pw - 80, 7, value, align='L')
        y += 7

    # ── 도장 ──
    stamp_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'work_order_templates', '도장.jpg')
    if _os.path.exists(stamp_path):
        pdf.image(stamp_path, x=lm + pw - 32, y=sup_y - 2, w=30, h=30)

    return bytes(pdf.output())


@app.route("/mail")
@require_perm("mail")
def mail_page():
    import base64 as _b64, os as _os
    stamp_b64 = ""
    stamp_path = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "work_order_templates", "도장.jpg")
    try:
        with open(stamp_path, "rb") as f:
            stamp_b64 = _b64.b64encode(f.read()).decode()
    except Exception:
        pass
    return render_template("mail.html", stamp_b64=stamp_b64,
                           is_admin=session.get("role") == "관리자")


@app.route("/mail/send")
def mail_send_page():
    from flask import redirect
    return redirect("/mail")


@app.route("/api/mail/cert_files")
@require_perm("mail")
def list_cert_files():
    import os as _os
    cert_dir = str(APP_DIR.parent / "documents" / "Report & Certificate")
    if not _os.path.exists(cert_dir):
        return jsonify([])
    files = []
    for fname in sorted(_os.listdir(cert_dir)):
        fpath = _os.path.join(cert_dir, fname)
        if _os.path.isfile(fpath) and not fname.startswith('.') and not fname.startswith('~'):
            files.append({"name": fname, "size": _os.path.getsize(fpath)})
    return jsonify(files)


@app.route("/api/mail/download_pdf", methods=["POST"])
def download_delivery_pdf():
    from flask import Response
    data = request.json or {}
    try:
        pdf_bytes = generate_delivery_pdf(data)
    except Exception as e:
        return jsonify({"ok": False, "error": f"PDF 생성 실패: {e}"})

    now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    client = (data.get("client") or "").replace(" ", "")[:10]
    filename = f"납품확인서_{now_str}_{client}.pdf"

    docs_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "documents", "납품확인서")
    os.makedirs(docs_dir, exist_ok=True)
    with open(os.path.join(docs_dir, filename), "wb") as f:
        f.write(pdf_bytes)

    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


@app.route("/api/mail/send_delivery", methods=["POST"])
def send_delivery_email():
    import smtplib, ssl, os as _os
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders

    import json as _json
    if request.content_type and 'application/json' in request.content_type:
        data = request.json or {}
        local_upload_files = []
    else:
        try:
            items = _json.loads(request.form.get("items", "[]"))
        except Exception:
            items = []
        data = {
            "address":       request.form.get("address", ""),
            "site_name":     request.form.get("site_name", ""),
            "client":        request.form.get("client", ""),
            "delivery_date": request.form.get("delivery_date", ""),
            "items":         items,
            "body":          request.form.get("body", ""),
            "extra_files":   request.form.getlist("extra_files"),
        }
        local_upload_files = request.files.getlist("local_files")

    to_email = (request.form.get("to_email") or data.get("to_email", "")).strip()
    cc_email = (request.form.get("cc_email") or data.get("cc_email", "")).strip()
    subject  = (request.form.get("subject") or data.get("subject", "납품확인서")).strip()

    if not to_email:
        return jsonify({"ok": False, "error": "받는 사람 이메일을 입력해주세요."})

    conn = get_conn()
    s = conn.execute("SELECT * FROM mail_settings WHERE id=1").fetchone()
    conn.close()
    if not s or not s["smtp_user"] or not s["smtp_password"]:
        return jsonify({"ok": False, "error": "SMTP 설정이 없습니다. 이메일발송 메뉴에서 먼저 설정해주세요."})

    # PDF 생성
    try:
        pdf_bytes = generate_delivery_pdf(data)
    except Exception as e:
        import traceback as _tb
        print("[PDF 생성 오류]", _tb.format_exc())
        return jsonify({"ok": False, "error": f"PDF 생성 실패: {e}"})

    # PDF 저장
    now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    client = (data.get("client") or "").replace(" ", "").replace("/", "")[:10]
    pdf_filename = f"납품확인서_{now_str}_{client}.pdf"
    docs_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "documents", "납품확인서")
    _os.makedirs(docs_dir, exist_ok=True)
    pdf_path = _os.path.join(docs_dir, pdf_filename)
    try:
        with open(pdf_path, "wb") as f:
            f.write(pdf_bytes)
    except Exception as e:
        return jsonify({"ok": False, "error": f"PDF 저장 실패: {e}"})

    # 이메일 전송
    smtp_host = s["smtp_host"] or "smtp.naver.com"
    smtp_port = int(s["smtp_port"] or 465)
    smtp_user = s["smtp_user"]
    smtp_pw   = s["smtp_password"]
    from_name = s["from_name"] or smtp_user

    from email.header import Header
    from email.utils import formataddr
    from_addr = formataddr((str(Header(from_name, 'utf-8')), smtp_user))

    msg = MIMEMultipart()
    msg["From"]    = from_addr
    msg["To"]      = to_email
    if cc_email:
        msg["Cc"]  = cc_email
    msg["Subject"] = subject

    body_text = data.get("body", "").strip() or "납품확인서를 첨부파일로 보내드립니다."
    msg.attach(MIMEText(body_text, "plain", "utf-8"))

    part = MIMEBase("application", "pdf")
    part.set_payload(pdf_bytes)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment", filename=("utf-8", "", pdf_filename))
    msg.attach(part)

    # 추가 서류 첨부 (Report & Certificate 폴더)
    cert_dir = str(APP_DIR.parent / "documents" / "Report & Certificate")
    for extra_name in (data.get("extra_files") or []):
        safe_name = _os.path.basename(extra_name)
        fpath = _os.path.join(cert_dir, safe_name)
        if _os.path.isfile(fpath):
            with open(fpath, "rb") as f:
                fb = f.read()
            ep = MIMEBase("application", "octet-stream")
            ep.set_payload(fb)
            encoders.encode_base64(ep)
            ep.add_header("Content-Disposition", "attachment", filename=("utf-8", "", safe_name))
            msg.attach(ep)

    # 로컬 PC에서 업로드한 파일 첨부
    for uf in local_upload_files:
        if not uf or not uf.filename:
            continue
        fb = uf.read()
        safe_name = _os.path.basename(uf.filename)
        ep = MIMEBase("application", "octet-stream")
        ep.set_payload(fb)
        encoders.encode_base64(ep)
        ep.add_header("Content-Disposition", "attachment", filename=("utf-8", "", safe_name))
        msg.attach(ep)

    status, error_msg = "성공", ""
    try:
        ctx = ssl.create_default_context()
        if smtp_port == 465:
            with smtplib.SMTP_SSL(smtp_host, smtp_port, context=ctx) as server:
                server.login(smtp_user, smtp_pw)
                recipients = [to_email] + ([cc_email] if cc_email else [])
                server.sendmail(smtp_user, recipients, msg.as_string())
        else:
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.ehlo()
                server.starttls(context=ctx)
                server.ehlo()
                server.login(smtp_user, smtp_pw)
                recipients = [to_email] + ([cc_email] if cc_email else [])
                server.sendmail(smtp_user, recipients, msg.as_string())
    except Exception as e:
        status = "실패"
        error_msg = str(e)

    sent_by = session.get("name") or session.get("username", "")
    sent_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    extra_names = [_os.path.basename(f) for f in (data.get("extra_files") or [])]
    local_names = [_os.path.basename(uf.filename) for uf in local_upload_files if uf and uf.filename]
    all_attachments = ", ".join([pdf_filename] + extra_names + local_names)
    conn = get_conn()
    conn.execute("""
        INSERT INTO mail_log (sent_at, sent_by, to_email, cc_email, subject, body, attachments, pdf_path, status, error_msg)
        VALUES (?,?,?,?,?,?,?,?,?,?)
    """, (sent_at, sent_by, to_email, cc_email, subject, body_text, all_attachments, pdf_path, status, error_msg))
    conn.commit()
    conn.close()

    if status == "실패":
        return jsonify({"ok": False, "error": f"발송 실패: {error_msg}"})
    return jsonify({"ok": True, "pdf": pdf_filename})


@app.route("/api/mail/settings", methods=["GET"])
def get_mail_settings():
    conn = get_conn()
    row = conn.execute("SELECT * FROM mail_settings WHERE id=1").fetchone()
    conn.close()
    if not row:
        return jsonify({})
    d = row_to_dict(row)
    d["smtp_password"] = "****" if d.get("smtp_password") else ""
    return jsonify(d)


@app.route("/api/mail/settings", methods=["POST"])
def save_mail_settings():
    if session.get("role") != "관리자":
        return jsonify({"ok": False, "error": "관리자 권한이 필요합니다."}), 403
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    cur = conn.execute("SELECT smtp_password FROM mail_settings WHERE id=1").fetchone()
    old_pw = cur["smtp_password"] if cur else ""
    new_pw = data.get("smtp_password", "")
    if new_pw == "****":
        new_pw = old_pw
    conn.execute("""
        UPDATE mail_settings SET
            smtp_host=?, smtp_port=?, smtp_user=?, smtp_password=?, from_name=?, updated_at=?
        WHERE id=1
    """, (
        data.get("smtp_host", "smtp.naver.com"),
        int(data.get("smtp_port", 465)),
        data.get("smtp_user", ""),
        new_pw,
        data.get("from_name", ""),
        now,
    ))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/mail/send", methods=["POST"])
def send_mail_api():
    import smtplib
    import ssl
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders

    to_email = request.form.get("to_email", "").strip()
    cc_email = request.form.get("cc_email", "").strip()
    subject = request.form.get("subject", "").strip()
    body = request.form.get("body", "").strip()
    files = request.files.getlist("attachments")

    if not to_email or not subject:
        return jsonify({"ok": False, "error": "받는 사람과 제목은 필수입니다."})

    conn = get_conn()
    s = conn.execute("SELECT * FROM mail_settings WHERE id=1").fetchone()
    conn.close()
    if not s or not s["smtp_user"] or not s["smtp_password"]:
        return jsonify({"ok": False, "error": "SMTP 설정이 되어 있지 않습니다. 설정을 먼저 저장해주세요."})

    smtp_host = s["smtp_host"] or "smtp.naver.com"
    smtp_port = int(s["smtp_port"] or 465)
    smtp_user = s["smtp_user"]
    smtp_pw = s["smtp_password"]
    from_name = s["from_name"] or smtp_user

    from email.header import Header
    from email.utils import formataddr
    from_addr = formataddr((str(Header(from_name, 'utf-8')), smtp_user))

    msg = MIMEMultipart()
    msg["From"] = from_addr
    msg["To"] = to_email
    if cc_email:
        msg["Cc"] = cc_email
    msg["Subject"] = subject
    body_type = request.form.get("body_type", "plain")
    msg.attach(MIMEText(body, body_type, "utf-8"))

    attachment_names = []
    for f in files:
        if not f or not f.filename:
            continue
        data = f.read()
        part = MIMEBase("application", "octet-stream")
        part.set_payload(data)
        encoders.encode_base64(part)
        fname = f.filename
        part.add_header("Content-Disposition", f'attachment; filename="{fname}"')
        msg.attach(part)
        attachment_names.append(fname)

    status = "성공"
    error_msg = ""
    try:
        ctx = ssl.create_default_context()
        if smtp_port == 465:
            with smtplib.SMTP_SSL(smtp_host, smtp_port, context=ctx) as server:
                server.login(smtp_user, smtp_pw)
                recipients = [to_email] + ([cc_email] if cc_email else [])
                server.sendmail(smtp_user, recipients, msg.as_string())
        else:
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.ehlo()
                server.starttls(context=ctx)
                server.ehlo()
                server.login(smtp_user, smtp_pw)
                recipients = [to_email] + ([cc_email] if cc_email else [])
                server.sendmail(smtp_user, recipients, msg.as_string())
    except Exception as e:
        status = "실패"
        error_msg = str(e)

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sent_by = session.get("name") or session.get("username", "")
    conn = get_conn()
    conn.execute("""
        INSERT INTO mail_log (sent_at, sent_by, to_email, cc_email, subject, body, attachments, status, error_msg)
        VALUES (?,?,?,?,?,?,?,?,?)
    """, (now, sent_by, to_email, cc_email, subject, body, ", ".join(attachment_names), status, error_msg))
    conn.commit()
    conn.close()

    if status == "실패":
        return jsonify({"ok": False, "error": f"발송 실패: {error_msg}"})
    return jsonify({"ok": True})


@app.route("/api/mail/log", methods=["GET"])
def get_mail_log():
    q = request.args.get("q", "").strip()
    conn = get_conn()
    if q:
        like = f"%{q}%"
        rows = conn.execute(
            "SELECT * FROM mail_log WHERE to_email LIKE ? OR subject LIKE ? OR sent_by LIKE ? OR attachments LIKE ? ORDER BY id DESC LIMIT 200",
            (like, like, like, like)
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM mail_log ORDER BY id DESC LIMIT 200").fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])


@app.route("/api/mail/log/<int:log_id>", methods=["DELETE"])
def delete_mail_log(log_id):
    conn = get_conn()
    row = conn.execute("SELECT pdf_path FROM mail_log WHERE id=?", (log_id,)).fetchone()
    if row and row["pdf_path"]:
        try:
            os.remove(row["pdf_path"])
        except Exception:
            pass
    conn.execute("DELETE FROM mail_log WHERE id=?", (log_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/mail/delivery_files")
@require_perm("mail")
def list_delivery_files():
    import os as _os
    q = request.args.get("q", "").strip().lower()
    docs_dir = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "documents", "납품확인서")
    if not _os.path.exists(docs_dir):
        return jsonify([])
    files = []
    for fname in sorted(_os.listdir(docs_dir), reverse=True):
        if not fname.endswith(".pdf"):
            continue
        if q and q not in fname.lower():
            continue
        fpath = _os.path.join(docs_dir, fname)
        files.append({"name": fname, "size": _os.path.getsize(fpath), "mtime": _os.path.getmtime(fpath)})
    return jsonify(files)


@app.route("/api/mail/cert_download/<path:filename>")
@require_perm("mail")
def download_cert_file(filename):
    safe_name = os.path.basename(filename)
    cert_dir = str(APP_DIR.parent / "documents" / "Report & Certificate")
    return send_from_directory(cert_dir, safe_name, as_attachment=True)


@app.route("/api/mail/reg_files")
@require_perm("mail")
def list_reg_files():
    import os as _os
    reg_dir = str(APP_DIR.parent / "documents" / "기타서류")
    if not _os.path.exists(reg_dir):
        return jsonify([])
    files = []
    for fname in sorted(_os.listdir(reg_dir)):
        fpath = _os.path.join(reg_dir, fname)
        if _os.path.isfile(fpath) and not fname.startswith('.') and not fname.startswith('~'):
            files.append({"name": fname, "size": _os.path.getsize(fpath)})
    return jsonify(files)


@app.route("/api/mail/reg_download/<path:filename>")
@require_perm("mail")
def download_reg_file(filename):
    safe_name = os.path.basename(filename)
    reg_dir = str(APP_DIR.parent / "documents" / "기타서류")
    return send_from_directory(reg_dir, safe_name, as_attachment=True)


@app.route("/ledger")
def ledger_page():
    if not session.get("user_id"):
        return redirect(url_for("login_page", next="/ledger"))
    if not (user_has_perm("ledger_vehicle") or user_has_perm("ledger_machine")):
        return "권한이 없습니다.", 403
    return render_template("ledger.html")


@app.route("/hr")
def hr_page():
    if not session.get("user_id"):
        return redirect(url_for("login_page", next="/hr"))
    if not any(user_has_perm(k) for k in ["hr_attendance", "hr_leave", "hr_org"]):
        return "권한이 없습니다.", 403
    my_name = session.get("name", "")
    lv_scope = {"type": "member"}
    if session.get("role") == "관리자":
        lv_scope = {"type": "admin"}
    else:
        conn = get_conn()
        head_row = conn.execute(
            "SELECT dept_id FROM org_members WHERE name=? AND is_head=1 LIMIT 1", (my_name,)
        ).fetchone()
        if head_row:
            lv_scope = {"type": "dept_head", "dept_id": head_row["dept_id"]}
        else:
            leader_rows = conn.execute(
                "SELECT team_id FROM org_members WHERE name=? AND is_team_leader=1", (my_name,)
            ).fetchall()
            if leader_rows:
                lv_scope = {"type": "team_leader", "team_ids": [r["team_id"] for r in leader_rows]}
        conn.close()
    return render_template("hr.html",
        current_user_name=my_name,
        current_user_role=session.get("role", "일반"),
        lv_can_write=user_can_write("hr_leave"),
        lv_user_scope=lv_scope,
    )


@app.route("/api/org")
@require_perm("hr_org")
def api_org():
    conn = get_conn()
    depts = conn.execute("SELECT * FROM org_departments ORDER BY sort_order, id").fetchall()
    result = []
    for d in depts:
        dept = dict(d)
        head = conn.execute("SELECT * FROM org_members WHERE dept_id=? AND is_head=1 ORDER BY sort_order LIMIT 1", (d["id"],)).fetchone()
        dept["head"] = dict(head) if head else None
        direct = conn.execute("SELECT * FROM org_members WHERE dept_id=? AND team_id IS NULL AND is_head=0 ORDER BY sort_order", (d["id"],)).fetchall()
        dept["members"] = [dict(m) for m in direct]
        teams = conn.execute("SELECT * FROM org_teams WHERE dept_id=? ORDER BY sort_order, id", (d["id"],)).fetchall()
        dept["teams"] = []
        for t in teams:
            team = dict(t)
            members = conn.execute("SELECT * FROM org_members WHERE team_id=? ORDER BY is_team_leader DESC, sort_order", (t["id"],)).fetchall()
            team["members"] = [dict(m) for m in members]
            dept["teams"].append(team)
        result.append(dept)
    conn.close()
    return jsonify(result)


@app.route("/api/org/departments", methods=["POST"])
@require_write_perm("hr_org")
def api_org_add_dept():
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    conn.execute(
        "INSERT INTO org_departments (label, dept_type, sort_order, created_at, updated_at) VALUES (?,?,?,?,?)",
        (data.get("label", "새부서"), data.get("dept_type", "regular"), data.get("sort_order", 990), now, now)
    )
    new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.commit(); conn.close()
    return jsonify({"id": new_id})


@app.route("/api/org/departments/<int:did>", methods=["PUT"])
@require_write_perm("hr_org")
def api_org_update_dept(did):
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    conn.execute("UPDATE org_departments SET label=?, updated_at=? WHERE id=?", (data.get("label", ""), now, did))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/org/departments/<int:did>", methods=["DELETE"])
@require_write_perm("hr_org")
def api_org_delete_dept(did):
    conn = get_conn()
    conn.execute("DELETE FROM org_members WHERE dept_id=?", (did,))
    conn.execute("DELETE FROM org_teams WHERE dept_id=?", (did,))
    conn.execute("DELETE FROM org_departments WHERE id=?", (did,))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/org/teams", methods=["POST"])
@require_write_perm("hr_org")
def api_org_add_team():
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    conn.execute(
        "INSERT INTO org_teams (dept_id, label, sort_order, created_at, updated_at) VALUES (?,?,?,?,?)",
        (data["dept_id"], data.get("label", "새팀"), data.get("sort_order", 990), now, now)
    )
    new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.commit(); conn.close()
    return jsonify({"id": new_id})


@app.route("/api/org/teams/<int:tid>", methods=["PUT"])
@require_write_perm("hr_org")
def api_org_update_team(tid):
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    conn.execute("UPDATE org_teams SET label=?, sort_order=?, updated_at=? WHERE id=?",
                 (data.get("label", ""), data.get("sort_order", 0), now, tid))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/org/teams/<int:tid>", methods=["DELETE"])
@require_write_perm("hr_org")
def api_org_delete_team(tid):
    conn = get_conn()
    conn.execute("UPDATE org_members SET team_id=NULL WHERE team_id=?", (tid,))
    conn.execute("DELETE FROM org_teams WHERE id=?", (tid,))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/org/members", methods=["POST"])
@require_write_perm("hr_org")
def api_org_add_member():
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    name = data.get("name", "")
    # 동명인 기존 행에서 정보 자동 복사 (rank/phone/hire_date/photo)
    existing = conn.execute(
        "SELECT rank, phone, hire_date, photo FROM org_members WHERE name=? LIMIT 1", (name,)
    ).fetchone()
    rank      = data.get("rank")     or (existing["rank"]      if existing else "")
    phone     = data.get("phone")    or (existing["phone"]     if existing else "")
    hire_date = data.get("hire_date")or (existing["hire_date"] if existing else "")
    photo     = data.get("photo")    or (existing["photo"]     if existing else "")
    conn.execute(
        "INSERT INTO org_members (dept_id, team_id, name, rank, phone, team_tag, photo, hire_date, is_head, is_team_leader, sort_order, created_at, updated_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (data["dept_id"], data.get("team_id") or None, name, rank,
         phone, data.get("team_tag", ""), photo, hire_date,
         1 if data.get("is_head") else 0,
         1 if data.get("is_team_leader") else 0, data.get("sort_order", 0), now, now)
    )
    new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.commit(); conn.close()
    return jsonify({"id": new_id})


@app.route("/api/org/members/<int:mid>", methods=["PUT"])
@require_write_perm("hr_org")
def api_org_update_member(mid):
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    conn.execute(
        "UPDATE org_members SET dept_id=?, team_id=?, name=?, rank=?, phone=?, team_tag=?, hire_date=?, is_head=?, is_team_leader=?, sort_order=?, updated_at=? WHERE id=?",
        (data["dept_id"], data.get("team_id") or None, data.get("name", ""), data.get("rank", ""),
         data.get("phone", ""), data.get("team_tag", ""), data.get("hire_date", ""),
         1 if data.get("is_head") else 0, 1 if data.get("is_team_leader") else 0,
         data.get("sort_order", 0), now, mid)
    )
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/org/members/<int:mid>", methods=["DELETE"])
@require_write_perm("hr_org")
def api_org_delete_member(mid):
    conn = get_conn()
    conn.execute("DELETE FROM org_members WHERE id=?", (mid,))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/org/members/<int:mid>/photo", methods=["POST"])
@require_write_perm("hr_org")
def api_org_member_photo(mid):
    if "photo" not in request.files:
        return jsonify({"error": "no file"}), 400
    f = request.files["photo"]
    ext = os.path.splitext(f.filename)[1].lower() or ".jpg"
    fname = f"member_{mid}{ext}"
    photo_dir = APP_DIR / "static" / "org_photos"
    photo_dir.mkdir(exist_ok=True)
    f.save(photo_dir / fname)
    try:
        img = Image.open(photo_dir / fname)
        img = ImageOps.exif_transpose(img)
        img.thumbnail((300, 300))
        img.save(photo_dir / fname)
    except Exception:
        pass
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    conn.execute("UPDATE org_members SET photo=?, updated_at=? WHERE id=?", (fname, now, mid))
    conn.commit(); conn.close()
    return jsonify({"photo": fname})


@app.route("/api/org/members/<int:mid>/move", methods=["POST"])
@require_write_perm("hr_org")
def api_org_move_member(mid):
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    conn.execute(
        "UPDATE org_members SET dept_id=?, team_id=?, sort_order=?, is_head=?, updated_at=? WHERE id=?",
        (data.get("dept_id"), data.get("team_id"), data.get("sort_order", 0),
         1 if data.get("is_head") else 0, now, mid)
    )
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/org/departments/reorder", methods=["POST"])
@require_write_perm("hr_org")
def api_org_reorder_depts():
    ids = (request.json or {}).get("ids", [])
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    for i, did in enumerate(ids):
        conn.execute("UPDATE org_departments SET sort_order=?, updated_at=? WHERE id=?", (i * 10, now, did))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/org/members/reorder", methods=["POST"])
@require_write_perm("hr_org")
def api_org_reorder_members():
    items = (request.json or {}).get("items", [])
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    for i, item in enumerate(items):
        conn.execute("UPDATE org_members SET sort_order=?, updated_at=? WHERE id=?",
                     (i * 10, now, item["id"]))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


# ── 휴가관리 API ──

def _current_fiscal_year():
    """현재 회계연도 (3/1 기준: 3월 이상이면 올해, 1~2월이면 작년)"""
    from datetime import date
    today = date.today()
    return today.year if today.month >= 3 else today.year - 1

def _fiscal_year_range(fy):
    """회계연도 날짜 범위 문자열 반환: ('{fy}-03-01', '{fy+1}-02-28 or 29')"""
    import calendar
    last_day = calendar.monthrange(fy + 1, 2)[1]
    return f"{fy}-03-01", f"{fy + 1}-02-{last_day:02d}"

def _calc_leave_entitlement(hire_date_str, fy=None):
    """회사 규정: 3/1 회계연도 기준
    - 입사 후 첫 3/1 이전 (첫 부분 회계연도): 11일 선지급
    - 첫 3/1 이후: 15일 (3년차부터 2년마다 +1일, 최대 25일)
    """
    if not hire_date_str:
        return 0
    try:
        from datetime import date
        hire = date.fromisoformat(hire_date_str[:10])
        if fy is None:
            fy = _current_fiscal_year()
        fy_start = date(fy, 3, 1)
        # 입사 후 첫 정식 회계연도 시작일 (입사 월이 3월 이상이면 다음해 3/1, 1~2월이면 당해 3/1)
        first_full_fy = date(hire.year + 1, 3, 1) if hire.month >= 3 else date(hire.year, 3, 1)
        if fy_start < first_full_fy:
            # 첫 부분 회계연도: 11일 선지급
            return 11
        else:
            # 정식 회계연도 몇 번째인지 (0-indexed)
            full_years = fy_start.year - first_full_fy.year
            return min(15 + max(full_years - 1, 0) // 2, 25)
    except Exception:
        return 0


def _get_approver_pairs(conn):
    """부서별 1차(팀장)/2차(총괄) 결재자 쌍 반환.
    {dept_id: {"ap1": name|None, "ap1_role": str, "ap2": name, "ap2_role": str}}
    - 미래전략: ap1=None, ap2=대표이사
    - 일반: ap1=is_team_leader, ap2=is_head
    """
    ceo = conn.execute("""
        SELECT m.name FROM org_members m
        JOIN org_departments d ON m.dept_id = d.id
        WHERE d.dept_type='special' AND d.label='경영진'
        AND (m.is_head=1 OR m.rank='대표이사')
        ORDER BY m.is_head DESC, m.id ASC
        LIMIT 1
    """).fetchone()
    ceo_name = ceo["name"] if ceo else ""

    # team_id → 팀장 이름 (팀별 1차 결재)
    team_leaders_by_team = {}
    for r in conn.execute("SELECT team_id, name FROM org_members WHERE is_team_leader=1 AND team_id IS NOT NULL"):
        team_leaders_by_team[r["team_id"]] = r["name"]
    # dept_id → 총괄 이름 (부서별 2차 결재)
    heads = {}
    for r in conn.execute("SELECT dept_id, name FROM org_members WHERE is_head=1"):
        heads[r["dept_id"]] = r["name"]

    depts = conn.execute("SELECT id, label FROM org_departments").fetchall()
    dept_is_strategy = {d["id"]: ("미래전략" in (d["label"] or "")) for d in depts}

    # team_id → {ap1, ap1_role, ap2, ap2_role}
    teams = conn.execute("SELECT id, dept_id FROM org_teams").fetchall()
    result = {}
    for t in teams:
        dept_id = t["dept_id"]
        if dept_is_strategy.get(dept_id):
            result[t["id"]] = {"ap1": None, "ap1_role": "", "ap2": ceo_name, "ap2_role": "대표이사"}
        else:
            result[t["id"]] = {
                "ap1": team_leaders_by_team.get(t["id"]),
                "ap1_role": "팀장",
                "ap2": heads.get(dept_id, "-"),
                "ap2_role": "총괄",
            }
    # dept_id 기반 fallback (팀 없이 부서 직속인 경우)
    dept_result = {}
    for d in depts:
        if dept_is_strategy.get(d["id"]):
            dept_result[d["id"]] = {"ap1": None, "ap1_role": "", "ap2": ceo_name, "ap2_role": "대표이사"}
        else:
            dept_result[d["id"]] = {"ap1": None, "ap1_role": "", "ap2": heads.get(d["id"], "-"), "ap2_role": "총괄"}
    return result, dept_result


@app.route("/api/hr/leave/members")
@require_perm("hr_leave")
def api_hr_leave_members():
    conn = get_conn()
    # 이름 중복 제거: 같은 이름은 첫 번째 행(MIN id)만 사용
    members = conn.execute("""
        SELECT m.id, m.name, m.rank, m.hire_date, m.phone, m.dept_id, m.team_id, m.is_head,
               d.label as dept_label, t.label as team_label
        FROM org_members m
        LEFT JOIN org_departments d ON m.dept_id = d.id
        LEFT JOIN org_teams t ON m.team_id = t.id
        WHERE d.label != '경영진'
          AND m.id = (SELECT MIN(m2.id) FROM org_members m2 WHERE m2.name = m.name)
        ORDER BY d.sort_order, m.is_head DESC, m.sort_order
    """).fetchall()
    fy = _current_fiscal_year()
    fy_start, fy_end = _fiscal_year_range(fy)
    team_pairs, dept_pairs = _get_approver_pairs(conn)
    result = []
    for m in members:
        mid = m["id"]
        # 사용 연차: 현재 회계연도(3/1~익년2/28) 내 승인된 일수 합산
        used = conn.execute("""
            SELECT COALESCE(SUM(r.days), 0) as total FROM hr_leave_requests r
            JOIN org_members om ON r.member_id = om.id
            WHERE om.name=? AND r.status='승인'
              AND r.start_date >= ? AND r.start_date <= ?
        """, (m["name"], fy_start, fy_end)).fetchone()["total"]
        entitled = _calc_leave_entitlement(m["hire_date"], fy)
        pair = (team_pairs.get(m["team_id"]) if m["team_id"] else None) or dept_pairs.get(m["dept_id"], {})
        ap1 = pair.get("ap1")
        if ap1 == m["name"]:
            ap1 = None
        result.append({
            "id": mid, "name": m["name"], "rank": m["rank"],
            "hire_date": m["hire_date"], "phone": m["phone"],
            "dept_id": m["dept_id"], "team_id": m["team_id"],
            "dept_label": m["dept_label"] or "", "team_label": m["team_label"] or "", "is_head": m["is_head"],
            "entitled": entitled, "used": float(used),
            "remaining": round(entitled - float(used), 1),
            "approver1_name": ap1 or "",
            "approver1_role": pair.get("ap1_role", "팀장") if ap1 else "",
            "approver2_name": pair.get("ap2", "-"),
            "approver2_role": pair.get("ap2_role", "총괄"),
        })
    conn.close()
    return jsonify(result)


@app.route("/api/hr/leave/settle")
@require_perm("hr_leave")
def api_hr_leave_settle():
    mid = request.args.get("member_id", type=int)
    leave_date_str = request.args.get("leave_date", "")
    if not mid or not leave_date_str:
        return jsonify({"error": "missing params"}), 400

    conn = get_conn()
    m = conn.execute(
        "SELECT m.*, d.label as dept_label FROM org_members m "
        "LEFT JOIN org_departments d ON m.dept_id=d.id WHERE m.id=?", (mid,)
    ).fetchone()
    if not m:
        conn.close()
        return jsonify({"error": "직원을 찾을 수 없습니다"}), 404

    hire_date_str = m["hire_date"]
    if not hire_date_str:
        conn.close()
        return jsonify({"error": "입사일이 입력되지 않은 직원입니다"}), 400

    try:
        from datetime import date
        import calendar as cal_mod
        hire = date.fromisoformat(hire_date_str[:10])
        leave = date.fromisoformat(leave_date_str[:10])
    except Exception:
        conn.close()
        return jsonify({"error": "날짜 형식 오류"}), 400

    if leave < hire:
        conn.close()
        return jsonify({"error": "퇴직일이 입사일보다 빠릅니다"}), 400

    from datetime import timedelta

    def full_months(start, end):
        n = (end.year - start.year) * 12 + (end.month - start.month)
        if end.day < start.day:
            n -= 1
        return max(n, 0)

    def add_years(d, y):
        try:
            return date(d.year + y, d.month, d.day)
        except ValueError:
            return date(d.year + y, d.month, d.day - 1)

    def yearly_days(n):
        return min(15 + (n - 1) // 2, 25)

    def used_in(name, s, e):
        row = conn.execute("""
            SELECT COALESCE(SUM(r.days),0) as t FROM hr_leave_requests r
            JOIN org_members om ON r.member_id=om.id
            WHERE om.name=? AND r.status='승인'
              AND r.start_date>=? AND r.start_date<=?
        """, (name, str(s), str(e))).fetchone()
        return float(row["t"])

    # 입사일 기준 (근로기준법 제60조)
    one_year = add_years(hire, 1)
    periods = []

    if leave < one_year:
        # 1년 미만 퇴직: 월차만
        months = full_months(hire, leave)
        entitled = min(months, 11)
        used = used_in(m["name"], hire, leave)
        periods.append({
            "label": f"{hire.strftime('%Y.%m.%d')}~{leave.strftime('%Y.%m.%d')} (1년 미만 월차)",
            "months": months, "entitled": entitled,
            "pre_given": 11, "used": used,
            "balance": round(entitled - used, 1)
        })
    else:
        # 1년 미만 월차 기간
        p_end = one_year - timedelta(days=1)
        months = full_months(hire, p_end)
        entitled = min(months, 11)
        used = used_in(m["name"], hire, p_end)
        periods.append({
            "label": f"{hire.strftime('%Y.%m.%d')}~{p_end.strftime('%Y.%m.%d')} (1년 미만 월차)",
            "months": months, "entitled": entitled,
            "pre_given": 11, "used": used,
            "balance": round(entitled - used, 1)
        })
        # 1주년마다 연차 발생
        n = 1
        p_start = one_year
        while p_start <= leave:
            yd = yearly_days(n)
            next_start = add_years(one_year, n)
            p_end = next_start - timedelta(days=1)
            actual_end = min(leave, p_end)
            used = used_in(m["name"], p_start, actual_end)
            if leave <= p_end:
                # 만근일에 연차 이미 발생 → 비례 없이 전체 부여
                m_in_year = full_months(p_start, leave)
                actual_entitled = yd
                label = f"{p_start.strftime('%Y.%m.%d')}~{leave.strftime('%Y.%m.%d')} ({n}년 만근 후, 퇴직 정산)"
            else:
                m_in_year = 12
                actual_entitled = yd
                label = f"{p_start.strftime('%Y.%m.%d')}~{p_end.strftime('%Y.%m.%d')} ({n}년 만근 후)"
            periods.append({
                "label": label,
                "months": m_in_year, "entitled": actual_entitled,
                "pre_given": yd, "used": used,
                "balance": round(actual_entitled - used, 1)
            })
            if leave <= p_end:
                break
            p_start = next_start
            n += 1

    conn.close()
    total_entitled = sum(p["entitled"] for p in periods)
    total_pre_given = sum(p["pre_given"] for p in periods)
    total_used = round(sum(p["used"] for p in periods), 1)
    last_balance = round(periods[-1]["balance"], 1) if periods else 0
    return jsonify({
        "name": m["name"], "hire_date": hire_date_str, "leave_date": leave_date_str,
        "periods": periods,
        "total_entitled": total_entitled, "total_pre_given": total_pre_given,
        "total_used": total_used, "total_balance": round(total_entitled - total_used, 1),
        "last_balance": last_balance
    })


@app.route("/api/hr/leave/requests")
@require_perm("hr_leave")
def api_hr_leave_list():
    year = request.args.get("year", str(datetime.now().year))
    month = request.args.get("month", "")
    conn = get_conn()
    if month:
        ym = f"{year}-{month.zfill(2)}"
        rows = conn.execute("""
            SELECT r.*, m.name, m.rank, m.dept_id, d.label as dept_label
            FROM hr_leave_requests r
            JOIN org_members m ON r.member_id = m.id
            LEFT JOIN org_departments d ON m.dept_id = d.id
            WHERE (substr(r.start_date,1,7)=? OR substr(r.end_date,1,7)=?)
            ORDER BY r.start_date
        """, (ym, ym)).fetchall()
    else:
        rows = conn.execute("""
            SELECT r.*, m.name, m.rank, m.dept_id, d.label as dept_label
            FROM hr_leave_requests r
            JOIN org_members m ON r.member_id = m.id
            LEFT JOIN org_departments d ON m.dept_id = d.id
            WHERE substr(r.start_date,1,4)=?
            ORDER BY r.start_date
        """, (year,)).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        result.append(d)
    conn.close()
    return jsonify(result)


@app.route("/api/hr/leave/requests", methods=["POST"])
@require_perm("hr_leave")
def api_hr_leave_create():
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    is_admin = user_can_write("hr_leave")
    my_name = session.get("name", "")
    # 일반 사용자: member_id를 본인 이름으로 강제, status는 신청으로 고정
    member_id = data.get("member_id")
    if not is_admin:
        me = conn.execute("SELECT id FROM org_members WHERE name=? LIMIT 1", (my_name,)).fetchone()
        if not me:
            conn.close(); return jsonify({"error": "조직도에 등록된 직원 정보가 없습니다."}), 403
        member_id = me["id"]
    status = data.get("status", "신청") if is_admin else "신청"
    # 결재자 자동 세팅
    m_row = conn.execute("SELECT dept_id, team_id, name FROM org_members WHERE id=? LIMIT 1", (member_id,)).fetchone()
    team_pairs, dept_pairs = _get_approver_pairs(conn)
    pair = (team_pairs.get(m_row["team_id"]) if m_row and m_row["team_id"] else None) \
           or dept_pairs.get(m_row["dept_id"] if m_row else 0, {})
    ap1 = pair.get("ap1") or ""
    if ap1 == (m_row["name"] if m_row else ""):
        ap1 = ""
    ap2 = pair.get("ap2") or ""
    conn.execute("""
        INSERT INTO hr_leave_requests
        (member_id, leave_type, start_date, end_date, days, reason, status,
         approver1_name, approver1_status, approver2_name, approver2_status,
         applied_by, created_at, updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (member_id, data.get("leave_type", "연차"),
          data["start_date"], data["end_date"], data.get("days", 1),
          data.get("reason", ""), status,
          ap1, "신청", ap2, "신청",
          my_name, now, now))
    new_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    # 1차 결재자(없으면 2차)에게 확인요청 알림 생성
    first_ap = ap1 or ap2
    if first_ap:
        ap_user = conn.execute("SELECT id FROM users WHERE name=? LIMIT 1", (first_ap,)).fetchone()
        msg = f"[연차신청] {my_name} / {data.get('leave_type','연차')} {data['start_date']}~{data['end_date']} ({data.get('days',1)}일)"
        conn.execute("""
            INSERT INTO check_requests
            (from_user_name, to_user_id, to_user_name, message, status, source_type, leave_request_id, created_at)
            VALUES (?,?,?,?,'미확인','leave',?,?)
        """, (my_name, ap_user["id"] if ap_user else None, first_ap, msg, new_id, now))
    conn.commit(); conn.close()
    return jsonify({"ok": True, "id": new_id})


@app.route("/api/hr/leave/requests/<int:rid>/approve", methods=["POST"])
@require_login
def api_hr_leave_approve(rid):
    data = request.json or {}
    action = data.get("action")  # "승인" or "반려"
    if action not in ("승인", "반려"):
        return jsonify({"error": "잘못된 요청"}), 400
    my_name = session.get("name", "")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    req = conn.execute("SELECT * FROM hr_leave_requests WHERE id=?", (rid,)).fetchone()
    if not req:
        conn.close(); return jsonify({"error": "신청 없음"}), 404

    # 결재 순서 판단 (ap1이 없으면 ap1 단계는 건너뜀)
    ap1_done = (not req["approver1_name"]) or req["approver1_status"] in ("", "승인")
    is_ap1 = req["approver1_name"] == my_name and req["approver1_status"] == "신청"
    is_ap2 = req["approver2_name"] == my_name and ap1_done and req["approver2_status"] == "신청"
    # 관리자는 어느 단계든 처리 가능
    if user_can_write("hr_leave"):
        is_ap1 = is_ap1 or (req["approver1_name"] and req["approver1_status"] == "신청")
        is_ap2 = is_ap2 or (req["approver2_status"] == "신청" and ap1_done)

    if not is_ap1 and not is_ap2:
        conn.close(); return jsonify({"error": "결재 권한 없음"}), 403

    if action == "반려":
        if is_ap1:
            conn.execute("UPDATE hr_leave_requests SET approver1_status='반려', status='반려', processed_at=?, processed_by=?, updated_at=? WHERE id=?",
                         (now, my_name, now, rid))
        else:
            conn.execute("UPDATE hr_leave_requests SET approver2_status='반려', status='반려', processed_at=?, processed_by=?, updated_at=? WHERE id=?",
                         (now, my_name, now, rid))
    else:  # 승인
        if is_ap1:
            conn.execute("UPDATE hr_leave_requests SET approver1_status='승인', updated_at=? WHERE id=?", (now, rid))
            # 2차 결재자에게 알림
            ap2_name = req["approver2_name"]
            if ap2_name:
                ap2_user = conn.execute("SELECT id FROM users WHERE name=? LIMIT 1", (ap2_name,)).fetchone()
                m_name = conn.execute("SELECT om.name FROM hr_leave_requests r JOIN org_members om ON r.member_id=om.id WHERE r.id=? LIMIT 1", (rid,)).fetchone()
                msg = f"[연차신청 1차승인] {m_name['name'] if m_name else ''} / {req['leave_type']} {req['start_date']}~{req['end_date']} ({req['days']}일) - 2차 결재 요청"
                conn.execute("""
                    INSERT INTO check_requests
                    (from_user_name, to_user_id, to_user_name, message, status, source_type, leave_request_id, created_at)
                    VALUES (?,?,?,?,'미확인','leave',?,?)
                """, (my_name, ap2_user["id"] if ap2_user else None, ap2_name, msg, rid, now))
        else:  # 2차 승인 → 최종 승인
            conn.execute("UPDATE hr_leave_requests SET approver2_status='승인', status='승인', processed_at=?, processed_by=?, updated_at=? WHERE id=?",
                         (now, my_name, now, rid))
    # 1차 승인이면 현재 결재자 것만 완료, 반려/2차승인이면 남은 것 전부 완료
    if action == "승인" and is_ap1:
        conn.execute("UPDATE check_requests SET status='확인완료', completed_at=?, completed_by=? WHERE leave_request_id=? AND status='미확인' AND to_user_id=?",
                     (now, my_name, rid, session.get("user_id")))
    else:
        conn.execute("UPDATE check_requests SET status='확인완료', completed_at=?, completed_by=? WHERE leave_request_id=? AND status='미확인'",
                     (now, my_name, rid))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/hr/leave/requests/<int:rid>", methods=["PUT"])
@require_perm("hr_leave")
def api_hr_leave_update(rid):
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    is_admin = user_can_write("hr_leave")
    my_name = session.get("name", "")
    # 일반 사용자: 본인 신청건만, status 변경 불가
    if not is_admin:
        row = conn.execute("SELECT m.name FROM hr_leave_requests r JOIN org_members m ON r.member_id=m.id WHERE r.id=?", (rid,)).fetchone()
        if not row or row["name"] != my_name:
            conn.close(); return jsonify({"error": "권한 없음"}), 403
    fields, vals = [], []
    for f in ["leave_type", "start_date", "end_date", "days", "reason", "memo"]:
        if f in data:
            fields.append(f"{f}=?"); vals.append(data[f])
    if is_admin:
        if "status" in data:
            fields += ["status=?", "processed_at=?", "processed_by=?"]
            vals += [data["status"], now, my_name]
        for f in ["approver1_status", "approver2_status"]:
            if f in data:
                fields.append(f"{f}=?"); vals.append(data[f])
    fields.append("updated_at=?"); vals.append(now); vals.append(rid)
    conn.execute(f"UPDATE hr_leave_requests SET {','.join(fields)} WHERE id=?", vals)
    conn.commit(); conn.close()
    return jsonify({"ok": True})


@app.route("/api/hr/leave/requests/<int:rid>", methods=["DELETE"])
@require_perm("hr_leave")
def api_hr_leave_delete(rid):
    conn = get_conn()
    is_admin = user_can_write("hr_leave")
    my_name = session.get("name", "")
    if not is_admin:
        row = conn.execute("SELECT m.name, r.status FROM hr_leave_requests r JOIN org_members m ON r.member_id=m.id WHERE r.id=?", (rid,)).fetchone()
        if not row or row["name"] != my_name or row["status"] == "승인":
            conn.close(); return jsonify({"error": "권한 없음"}), 403
    conn.execute("DELETE FROM hr_leave_requests WHERE id=?", (rid,))
    conn.commit(); conn.close()
    return jsonify({"ok": True})


# ── 근태관리: 엑셀 업로드 파싱 ──────────────────────────────────────────────
@app.route("/api/hr/attendance/parse", methods=["POST"])
@require_perm("hr_attendance")
def api_hr_attendance_parse():
    f = request.files.get("file")
    if not f:
        return jsonify({"ok": False, "error": "파일 없음"}), 400

    fname = f.filename.lower()
    tmp = APP_DIR / "static" / "tmp_attendance"
    tmp.mkdir(exist_ok=True)
    tmp_path = tmp / f.filename
    f.save(str(tmp_path))

    try:
        if fname.endswith(".xls"):
            wb = xlrd.open_workbook(str(tmp_path))
            ws = wb.sheet_by_index(0)
            rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]
        else:
            wb2 = load_workbook(str(tmp_path), data_only=True)
            ws2 = wb2.active
            rows = [[cell.value for cell in row] for row in ws2.iter_rows()]
    finally:
        try: tmp_path.unlink()
        except: pass

    # 헤더 행 찾기 (번호, 사용자ID, 이름, 근무일자...)
    header_idx = None
    for i, row in enumerate(rows):
        if row and str(row[0]).strip() == "번호":
            header_idx = i
            break

    if header_idx is None:
        return jsonify({"ok": False, "error": "파일 형식이 맞지 않습니다."}), 400

    # 연도/월 추출 (첫 번째 날짜 행에서)
    month_str = ""
    for row in rows[header_idx+2:]:
        date_val = str(row[3]).strip() if len(row) > 3 else ""
        if date_val and "/" in date_val:
            parts = date_val.split("/")
            if len(parts) >= 2:
                month_str = f"{parts[0]}년 {int(parts[1])}월"
            break

    # 직원별 데이터 수집
    employees = {}
    WEEKDAY_KO = ["월", "화", "수", "목", "금", "토", "일"]

    for row in rows[header_idx+2:]:
        if not row or not row[2]:
            continue
        name = str(row[2]).strip()
        if not name:
            continue
        date_str = str(row[3]).strip() if len(row) > 3 else ""
        if not date_str or "/" not in date_str:
            continue

        def _to_hhmm(val):
            """xlrd float / datetime.time / 문자열 → HH:MM"""
            if val is None or val == "":
                return ""
            import datetime as _dt
            if isinstance(val, _dt.time):
                return f"{val.hour:02d}:{val.minute:02d}"
            if isinstance(val, _dt.datetime):
                return f"{val.hour:02d}:{val.minute:02d}"
            if isinstance(val, float) and 0 <= val < 2:
                total_min = round(val * 24 * 60)
                h, m = divmod(total_min, 60)
                return f"{h:02d}:{m:02d}"
            s = str(val).strip()
            # "HH:MM:SS" → "HH:MM"
            import re as _re
            mm = _re.match(r'^(\d{1,2}:\d{2}):\d{2}$', s)
            if mm:
                return mm.group(1)
            return s

        day_type   = str(row[4]).strip() if len(row) > 4 else ""
        checkin    = _to_hhmm(row[5]) if len(row) > 5 else ""
        checkout   = _to_hhmm(row[6]) if len(row) > 6 else ""
        go_out     = _to_hhmm(row[7]) if len(row) > 7 else ""
        return_in  = _to_hhmm(row[8]) if len(row) > 8 else ""
        overtime_d = row[12] if len(row) > 12 else 0

        # 근무시간 계산 (출퇴근 기반)
        work_str = ""
        try:
            from datetime import datetime as dt
            ci = dt.strptime(checkin, "%H:%M")
            co = dt.strptime(checkout, "%H:%M")
            diff = co - ci
            total_min = int(diff.total_seconds() / 60)
            if total_min > 0:
                h, m = divmod(total_min, 60)
                work_str = f"{h}시간 {m}분"
        except:
            pass

        # 연장근로 시간 변환
        ot_str = ""
        try:
            ot_min = int(float(overtime_d) * 24 * 60)
            if ot_min > 0:
                h, m = divmod(ot_min, 60)
                ot_str = f"{h}시간 {m}분"
        except:
            pass

        # 날짜에서 요일 추출
        try:
            d = datetime.strptime(date_str, "%Y/%m/%d")
            day_ko = WEEKDAY_KO[d.weekday()]
            date_fmt = f"{d.month}/{d.day}({day_ko})"
            is_weekend = d.weekday() >= 5
        except:
            date_fmt = date_str
            is_weekend = "토요" in day_type or "일요" in day_type

        if name not in employees:
            employees[name] = []

        employees[name].append({
            "date": date_fmt,
            "day_type": day_type,
            "checkin": checkin,
            "checkout": checkout,
            "go_out": go_out,
            "return_in": return_in,
            "work_hours": work_str,
            "overtime": ot_str,
            "is_weekend": is_weekend,
        })

    return jsonify({
        "ok": True,
        "month": month_str,
        "employees": sorted(employees.keys()),
        "data": employees,
    })


@app.route("/api/ledger/vehicles", methods=["GET"])
@require_perm("ledger_vehicle")
def list_vehicles():
    conn = get_conn()
    rows = conn.execute("""
        SELECT * FROM vehicle_ledger
        ORDER BY
            CASE WHEN unit_no IS NULL OR TRIM(unit_no)='' THEN 1 ELSE 0 END,
            CAST(REPLACE(REPLACE(TRIM(unit_no),'호',''),' ','') AS INTEGER),
            company,
            vehicle_type
    """).fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])


@app.route("/api/ledger/vehicles", methods=["POST"])
@require_write_perm("ledger_vehicle")
def create_vehicle():
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    cur = conn.execute("""
        INSERT INTO vehicle_ledger
            (company, vehicle_type, unit_no, vehicle_no, hipass_digits, fuel_type,
             reg_date, owner, manager, hipass_card, vehicle_class, mileage_json,
             engine_oil, tire, repair, urea, violation, car_insurance, driver_insurance,
             inspection, tax, memo, created_at, updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        data.get("company", "플러스도어"),
        data.get("vehicle_type", ""),
        data.get("unit_no", ""),
        data.get("vehicle_no", ""),
        data.get("hipass_digits", ""),
        data.get("fuel_type", ""),
        data.get("reg_date", ""),
        data.get("owner", ""),
        data.get("manager", ""),
        data.get("hipass_card", ""),
        data.get("vehicle_class", ""),
        data.get("mileage_json", "[]"),
        data.get("engine_oil", ""),
        data.get("tire", ""),
        data.get("repair", ""),
        data.get("urea", ""),
        data.get("violation", ""),
        data.get("car_insurance", ""),
        data.get("driver_insurance", ""),
        data.get("inspection", ""),
        data.get("tax", ""),
        data.get("memo", ""),
        now, now
    ))
    new_id = cur.lastrowid
    conn.commit()
    row = conn.execute("SELECT * FROM vehicle_ledger WHERE id=?", (new_id,)).fetchone()
    conn.close()
    return jsonify({"ok": True, "row": row_to_dict(row)})


@app.route("/api/ledger/vehicles/<int:vid>", methods=["PUT"])
@require_write_perm("ledger_vehicle")
def update_vehicle(vid):
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    conn.execute("""
        UPDATE vehicle_ledger SET
            company=?, vehicle_type=?, unit_no=?, vehicle_no=?, hipass_digits=?,
            fuel_type=?, reg_date=?, owner=?, manager=?, hipass_card=?,
            vehicle_class=?, mileage_json=?, engine_oil=?, tire=?, repair=?,
            urea=?, violation=?, car_insurance=?, driver_insurance=?,
            inspection=?, tax=?, memo=?, driver_age=?, updated_at=?
        WHERE id=?
    """, (
        data.get("company", "플러스도어"),
        data.get("vehicle_type", ""),
        data.get("unit_no", ""),
        data.get("vehicle_no", ""),
        data.get("hipass_digits", ""),
        data.get("fuel_type", ""),
        data.get("reg_date", ""),
        data.get("owner", ""),
        data.get("manager", ""),
        data.get("hipass_card", ""),
        data.get("vehicle_class", ""),
        data.get("mileage_json", "[]"),
        data.get("engine_oil", ""),
        data.get("tire", ""),
        data.get("repair", ""),
        data.get("urea", ""),
        data.get("violation", ""),
        data.get("car_insurance", ""),
        data.get("driver_insurance", ""),
        data.get("inspection", ""),
        data.get("tax", ""),
        data.get("memo", ""),
        data.get("driver_age", ""),
        now, vid
    ))
    conn.commit()
    row = conn.execute("SELECT * FROM vehicle_ledger WHERE id=?", (vid,)).fetchone()
    conn.close()
    return jsonify({"ok": True, "row": row_to_dict(row)})


@app.route("/api/ledger/vehicles/<int:vid>/status", methods=["PUT"])
@require_write_perm("ledger_vehicle")
def update_vehicle_status(vid):
    data = request.json or {}
    status = data.get("status", "운행")
    if status not in ("운행", "매각"):
        return jsonify({"ok": False, "error": "유효하지 않은 상태입니다."})
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    conn.execute("UPDATE vehicle_ledger SET status=?, updated_at=? WHERE id=?", (status, now, vid))
    conn.commit()
    row = conn.execute("SELECT * FROM vehicle_ledger WHERE id=?", (vid,)).fetchone()
    conn.close()
    return jsonify({"ok": True, "row": row_to_dict(row)})


@app.route("/api/ledger/vehicles/<int:vid>", methods=["DELETE"])
@require_write_perm("ledger_vehicle")
def delete_vehicle(vid):
    conn = get_conn()
    conn.execute("DELETE FROM vehicle_ledger WHERE id=?", (vid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ── 설비관리대장 ──────────────────────────────────────────────────────────────

@app.route("/api/ledger/equipment", methods=["GET"])
@require_perm("ledger_machine")
def list_equipment():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM equipment_ledger ORDER BY company, manage_no, name").fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])


@app.route("/api/ledger/equipment", methods=["POST"])
@require_write_perm("ledger_machine")
def create_equipment():
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    cur = conn.execute("""
        INSERT INTO equipment_ledger
            (company, name, manage_no, mfg_no, type, grade, category,
             maker, mfg_date, voltage, supplier, install_date, location,
             parts_json, repair_json, memo,
             maker_contact, maker_note, supplier_contact, supplier_note,
             created_at, updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        data.get("company","플러스도어"), data.get("name",""), data.get("manage_no",""),
        data.get("mfg_no",""), data.get("type",""), data.get("grade",""), data.get("category",""),
        data.get("maker",""), data.get("mfg_date",""), data.get("voltage",""),
        data.get("supplier",""), data.get("install_date",""), data.get("location",""),
        data.get("parts_json","[]"), data.get("repair_json","[]"), data.get("memo",""),
        data.get("maker_contact",""), data.get("maker_note",""),
        data.get("supplier_contact",""), data.get("supplier_note",""),
        now, now
    ))
    new_id = cur.lastrowid
    conn.commit()
    row = conn.execute("SELECT * FROM equipment_ledger WHERE id=?", (new_id,)).fetchone()
    conn.close()
    return jsonify({"ok": True, "row": row_to_dict(row)})


@app.route("/api/ledger/equipment/<int:eid>", methods=["PUT"])
@require_write_perm("ledger_machine")
def update_equipment(eid):
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    existing = conn.execute("SELECT * FROM equipment_ledger WHERE id=?", (eid,)).fetchone()
    ex = row_to_dict(existing) if existing else {}
    conn.execute("""
        UPDATE equipment_ledger SET
            company=?, name=?, manage_no=?, mfg_no=?, type=?, grade=?, category=?,
            maker=?, mfg_date=?, voltage=?, supplier=?, install_date=?, location=?,
            parts_json=?, repair_json=?, memo=?,
            maker_contact=?, maker_note=?, supplier_contact=?, supplier_note=?,
            updated_at=?
        WHERE id=?
    """, (
        data.get("company", ex.get("company","")), data.get("name", ex.get("name","")),
        data.get("manage_no", ex.get("manage_no","")), data.get("mfg_no", ex.get("mfg_no","")),
        data.get("type", ex.get("type","")), data.get("grade", ex.get("grade","")),
        data.get("category", ex.get("category","")), data.get("maker", ex.get("maker","")),
        data.get("mfg_date", ex.get("mfg_date","")), data.get("voltage", ex.get("voltage","")),
        data.get("supplier", ex.get("supplier","")), data.get("install_date", ex.get("install_date","")),
        data.get("location", ex.get("location","")),
        data.get("parts_json", ex.get("parts_json","[]")),
        data.get("repair_json", ex.get("repair_json","[]")),
        data.get("memo", ex.get("memo","")),
        data.get("maker_contact", ex.get("maker_contact","")),
        data.get("maker_note", ex.get("maker_note","")),
        data.get("supplier_contact", ex.get("supplier_contact","")),
        data.get("supplier_note", ex.get("supplier_note","")),
        now, eid
    ))
    conn.commit()
    row = conn.execute("SELECT * FROM equipment_ledger WHERE id=?", (eid,)).fetchone()
    conn.close()
    return jsonify({"ok": True, "row": row_to_dict(row)})


@app.route("/api/ledger/equipment/<int:eid>", methods=["DELETE"])
@require_write_perm("ledger_machine")
def delete_equipment(eid):
    conn = get_conn()
    conn.execute("DELETE FROM equipment_ledger WHERE id=?", (eid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/ledger/equipment/<int:eid>/photo", methods=["POST"])
@require_write_perm("ledger_machine")
def upload_equipment_photo(eid):
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "파일이 없습니다."}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".heic", ".heif"}:
        return jsonify({"ok": False, "error": "지원하지 않는 파일 형식입니다."}), 400
    safe_name = f"eq_{eid}_{datetime.now().strftime('%Y%m%d%H%M%S')}{ext}"
    save_path = EQUIPMENT_DOCS_DIR / safe_name
    try:
        save_resized_photo(f, save_path, target_bytes=2*1024*1024, max_side=2400)
    except Exception:
        f.stream.seek(0)
        f.save(str(save_path))
    conn = get_conn()
    conn.execute("UPDATE equipment_ledger SET photo=?, updated_at=? WHERE id=?",
                 (safe_name, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), eid))
    conn.commit()
    row = conn.execute("SELECT * FROM equipment_ledger WHERE id=?", (eid,)).fetchone()
    conn.close()
    return jsonify({"ok": True, "row": row_to_dict(row)})


@app.route("/api/ledger/vehicles/<int:vid>/docs", methods=["GET"])
@require_perm("ledger_vehicle")
def list_vehicle_docs(vid):
    conn = get_conn()
    rows = conn.execute("SELECT * FROM vehicle_docs WHERE vehicle_id=? ORDER BY id DESC", (vid,)).fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])


@app.route("/api/ledger/vehicles/<int:vid>/docs", methods=["POST"])
@require_write_perm("ledger_vehicle")
def upload_vehicle_doc(vid):
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"ok": False, "error": "파일이 없습니다."}), 400
    ext = os.path.splitext(f.filename)[1].lower()
    if ext not in {".jpg", ".jpeg", ".png", ".gif", ".webp", ".heic", ".heif", ".pdf"}:
        return jsonify({"ok": False, "error": "지원하지 않는 파일 형식입니다."}), 400
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    safe_name = f"vehicle_{vid}_{datetime.now().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex[:6]}{ext}"
    save_path = VEHICLE_DOCS_DIR / safe_name
    if ext in {".jpg", ".jpeg", ".png", ".gif", ".webp", ".heic", ".heif"}:
        try:
            save_resized_photo(f, save_path, target_bytes=2*1024*1024, max_side=2400)
        except Exception:
            f.stream.seek(0)
            f.save(str(save_path))
    else:
        f.save(str(save_path))
    conn = get_conn()
    cur = conn.execute(
        "INSERT INTO vehicle_docs(vehicle_id, filename, original_name, uploaded_at) VALUES (?,?,?,?)",
        (vid, f"vehicle_docs/{safe_name}", f.filename, now)
    )
    conn.commit()
    row = conn.execute("SELECT * FROM vehicle_docs WHERE id=?", (cur.lastrowid,)).fetchone()
    conn.close()
    return jsonify({"ok": True, "doc": row_to_dict(row)})


@app.route("/api/ledger/vehicles/<int:vid>/docs/<int:doc_id>", methods=["DELETE"])
@require_write_perm("ledger_vehicle")
def delete_vehicle_doc(vid, doc_id):
    conn = get_conn()
    row = conn.execute("SELECT * FROM vehicle_docs WHERE id=? AND vehicle_id=?", (doc_id, vid)).fetchone()
    if row:
        try:
            p = UPLOAD_DIR / row["filename"]
            if p.exists():
                p.unlink()
        except Exception:
            pass
        conn.execute("DELETE FROM vehicle_docs WHERE id=?", (doc_id,))
        conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/mail/pdf/<path:filename>")
@require_perm("mail")
def serve_delivery_pdf(filename):
    safe_name = os.path.basename(filename)
    docs_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "documents", "납품확인서")
    as_attachment = request.args.get("dl") == "1"
    return send_from_directory(docs_dir, safe_name, as_attachment=as_attachment)


@app.route("/price")
@require_perm("price")
def price_page():
    return render_template("price.html")

@app.route("/journal")
def journal_page():
    if not session.get("user_id"):
        return redirect(url_for("login_page", next="/journal"))
    if not any(user_has_perm(k) for k in ["journal_sales", "journal_consult", "journal_measure", "journal_install"]):
        return "권한이 없습니다.", 403
    return render_template("journal.html")

@app.route("/work_order")
@require_perm("work_order")
def work_order_page():
    return render_template("work_order.html")

@app.route("/systemdoor")
@require_perm("work_order")
def systemdoor_page():
    return render_template("systemdoor.html")

@app.route("/ninedoor")
@require_perm("work_order")
def ninedoor_page():
    return render_template("ninedoor.html")

@app.route("/as")
@require_perm("as")
def as_page():
    return render_template("as.html")

@app.route("/uploads/<path:filename>")
def uploaded_file(filename):
    return send_from_directory(UPLOAD_DIR, filename)

@app.route("/work_order_output/<path:filename>")
def work_order_output_file(filename):
    return send_from_directory(WORK_ORDER_OUTPUT_DIR, filename)

@app.route("/price_images/<path:filename>")
def price_image(filename):
    price_img_dir = APP_DIR.parent / "price" / "images"
    return send_from_directory(price_img_dir, filename)

@app.route("/api/regions", methods=["GET"])
@require_perm("production")
def list_regions():
    conn = get_conn()
    rows = conn.execute("SELECT name FROM regions ORDER BY name").fetchall()
    conn.close()
    return jsonify([r["name"] for r in rows])


@app.route("/api/journal/regions", methods=["GET"])
@require_perm("journal_sales")
def journal_list_regions():
    conn = get_conn()
    rows = conn.execute("SELECT name FROM regions ORDER BY name").fetchall()
    conn.close()
    return jsonify([r["name"] for r in rows])


@app.route("/api/regions", methods=["POST"])
@require_perm("production")
def create_region():
    data = request.json or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "지역명을 입력하세요."})
    try:
        conn = get_conn()
        conn.execute("INSERT INTO regions (name, created_at) VALUES (?, ?)",
                     (name, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        conn.commit()
        conn.close()
        return jsonify({"ok": True})
    except Exception:
        return jsonify({"ok": False, "error": "이미 등록된 지역입니다."})


@app.route("/api/regions/<path:name>", methods=["DELETE"])
@require_perm("production")
def delete_region(name):
    conn = get_conn()
    conn.execute("DELETE FROM regions WHERE name=?", (name,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/customers", methods=["GET"])
def list_customers():
    include_inactive = request.args.get("include_inactive", "").strip() == "1"
    conn = get_conn()
    if include_inactive:
        rows = conn.execute("SELECT * FROM customers ORDER BY active DESC, name ASC").fetchall()
    else:
        rows = conn.execute("SELECT * FROM customers WHERE COALESCE(active,1)=1 ORDER BY name ASC").fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])

@app.route("/api/customers", methods=["POST"])
def create_customer():
    data = request.json or {}
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "거래처명을 입력하세요."}), 400
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    memo = (data.get("memo") or "").strip()
    payment_note = (data.get("payment_note") or "").strip()
    contact_name = (data.get("contact_name") or "").strip()
    contact_phone = (data.get("contact_phone") or "").strip()
    contact_memo = (data.get("contact_memo") or "").strip()
    default_region = (data.get("default_region") or "").strip()
    conn = get_conn()
    existing = conn.execute("SELECT * FROM customers WHERE name=?", (name,)).fetchone()
    if existing:
        # 이미 있는 거래처면 중복 등록하지 않고, 사용안함 상태였으면 다시 사용으로 바꾼다.
        if existing["active"] == 0:
            conn.execute("UPDATE customers SET active=1 WHERE name=?", (name,))
        if memo or payment_note or contact_name or contact_phone or contact_memo or default_region:
            conn.execute("UPDATE customers SET memo=?, payment_note=?, contact_name=?, contact_phone=?, contact_memo=?, default_region=? WHERE name=?", (memo, payment_note, contact_name, contact_phone, contact_memo, default_region, name))
        conn.commit()
        conn.close()
        return jsonify({"ok": True, "existed": True, "reactivated": existing["active"] == 0})

    cur = conn.execute(
        "INSERT INTO customers(name, memo, payment_note, contact_name, contact_phone, contact_memo, default_region, sales_status, active, created_at) VALUES(?, ?, ?, ?, ?, ?, ?, '거래중', 1, ?)",
        (name, memo, payment_note, contact_name, contact_phone, contact_memo, default_region, now)
    )
    cid = cur.lastrowid
    # 영업일지 sales_leads에도 즉시 반영 (거래중)
    existing_lead = conn.execute("SELECT id FROM sales_leads WHERE linked_customer_id=?", (cid,)).fetchone()
    if not existing_lead:
        conn.execute("""
            INSERT INTO sales_leads (name, phone, sales_status, linked_customer_id, active, created_at)
            VALUES (?,?,?,?,1,?)
        """, (name, contact_phone, '거래중', cid, now))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "existed": False})

@app.route("/api/customers/<path:name>", methods=["PUT"])
def update_customer_info(name):
    data = request.json or {}
    old_name = (name or "").strip()
    new_name = (data.get("name") or old_name).strip()
    memo = (data.get("memo") or "").strip()
    payment_note = (data.get("payment_note") or "").strip()
    contact_name = (data.get("contact_name") or "").strip()
    contact_phone = (data.get("contact_phone") or "").strip()
    contact_memo = (data.get("contact_memo") or "").strip()
    default_region = (data.get("default_region") or "").strip()
    active = 1 if int(data.get("active", 1) or 0) else 0

    if not old_name or not new_name:
        return jsonify({"ok": False, "error": "거래처명을 입력하세요."}), 400

    conn = get_conn()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute("INSERT OR IGNORE INTO customers(name, active, created_at) VALUES(?, 1, ?)", (old_name, now))

    if new_name != old_name:
        # 기존 작업내역까지 새 거래처명으로 일괄 변경해서 통계가 합쳐지게 한다.
        exists = conn.execute("SELECT * FROM customers WHERE name=?", (new_name,)).fetchone()
        if not exists:
            conn.execute("INSERT INTO customers(name, memo, payment_note, contact_name, contact_phone, contact_memo, default_region, sales_status, active, created_at) VALUES(?, ?, ?, ?, ?, ?, ?, '거래중', ?, ?)", (new_name, memo, payment_note, contact_name, contact_phone, contact_memo, default_region, active, now))
        else:
            conn.execute("UPDATE customers SET memo=?, payment_note=?, contact_name=?, contact_phone=?, contact_memo=?, default_region=?, active=? WHERE name=?", (memo, payment_note, contact_name, contact_phone, contact_memo, default_region, active, new_name))
        conn.execute("UPDATE schedules SET customer=?, updated_at=? WHERE customer=?", (new_name, now, old_name))
        conn.execute("DELETE FROM customers WHERE name=?", (old_name,))
    else:
        conn.execute("UPDATE customers SET memo=?, payment_note=?, contact_name=?, contact_phone=?, contact_memo=?, default_region=?, active=? WHERE name=?", (memo, payment_note, contact_name, contact_phone, contact_memo, default_region, active, new_name))

    conn.commit()
    conn.close()
    return jsonify({"ok": True, "name": new_name})


@app.route("/api/customers/delete", methods=["POST"])
def delete_customer_post():
    data = request.json or {}
    name = (data.get("name") or "").strip()
    force = bool(data.get("force"))
    if not name:
        return jsonify({"ok": False, "error": "거래처명을 입력하세요."}), 400

    conn = get_conn()
    exists = conn.execute("SELECT * FROM customers WHERE name=?", (name,)).fetchone()
    if not exists:
        conn.close()
        return jsonify({"ok": False, "error": "거래처를 찾을 수 없습니다."}), 404

    cnt_row = conn.execute("SELECT COUNT(*) AS cnt FROM schedules WHERE customer=?", (name,)).fetchone()
    cnt = cnt_row["cnt"] if cnt_row else 0
    if cnt > 0 and not force:
        conn.close()
        return jsonify({"ok": False, "has_records": True, "count": cnt, "error": "거래내역이 있는 거래처입니다."}), 409

    conn.execute("DELETE FROM customers WHERE name=?", (name,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/customers/<path:name>", methods=["DELETE"])
def delete_customer(name):
    name = (name or "").strip()
    force = request.args.get("force", "") == "1"
    if not name:
        return jsonify({"ok": False, "error": "거래처명을 입력하세요."}), 400

    conn = get_conn()
    exists = conn.execute("SELECT * FROM customers WHERE name=?", (name,)).fetchone()
    if not exists:
        conn.close()
        return jsonify({"ok": False, "error": "거래처를 찾을 수 없습니다."}), 404

    cnt_row = conn.execute("SELECT COUNT(*) AS cnt FROM schedules WHERE customer=?", (name,)).fetchone()
    cnt = cnt_row["cnt"] if cnt_row else 0
    if cnt > 0 and not force:
        conn.close()
        return jsonify({"ok": False, "has_records": True, "count": cnt, "error": "거래내역이 있는 거래처입니다."}), 409

    # 거래처 목록에서만 삭제한다. 기존 schedules 작업내역의 거래처명은 통계/이력 보존을 위해 유지한다.
    conn.execute("DELETE FROM customers WHERE name=?", (name,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "deleted": True, "kept_schedule_records": cnt})

@app.route("/api/customer_manage", methods=["GET"])
def customer_manage_list():
    q = request.args.get("q", "").strip()
    include_inactive = request.args.get("include_inactive", "1").strip() == "1"
    this_year = datetime.now().strftime("%Y")
    conn = get_conn()
    sql = """
        SELECT c.*,
               COALESCE((SELECT SUM(qty) FROM schedules s WHERE s.customer=c.name),0) AS total_qty,
               COALESCE((SELECT COUNT(*) FROM schedules s WHERE s.customer=c.name),0) AS total_count,
               COALESCE((SELECT SUM(qty) FROM schedules s WHERE s.customer=c.name AND substr(COALESCE(NULLIF(s.due_date,''), s.order_date),1,4)=?),0) AS year_qty,
               COALESCE((SELECT COUNT(*) FROM schedules s WHERE s.customer=c.name AND substr(COALESCE(NULLIF(s.due_date,''), s.order_date),1,4)=?),0) AS year_count,
               (SELECT MAX(COALESCE(NULLIF(s.due_date,''), s.order_date)) FROM schedules s WHERE s.customer=c.name) AS last_date
        FROM customers c
        WHERE 1=1
    """
    params = [this_year, this_year]
    if not include_inactive:
        sql += " AND COALESCE(c.active,1)=1"
    if q:
        sql += " AND (c.name LIKE ? OR c.memo LIKE ? OR c.payment_note LIKE ? OR c.contact_name LIKE ? OR c.contact_phone LIKE ? OR c.contact_memo LIKE ?)"
        like = f"%{q}%"
        params += [like, like, like, like, like, like]
    sql += " ORDER BY COALESCE(c.active,1) DESC, c.name ASC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify({"year": this_year, "rows": [row_to_dict(r) for r in rows]})

def ensure_schedules_work_order_id_column():
    conn = get_conn()
    existing = [r["name"] for r in conn.execute("PRAGMA table_info(schedules)").fetchall()]
    if "work_order_id" not in existing:
        conn.execute("ALTER TABLE schedules ADD COLUMN work_order_id TEXT DEFAULT ''")
        conn.commit()
    conn.close()

@app.route("/api/schedules", methods=["GET"])
def list_schedules():
    ensure_schedules_work_order_id_column()
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    order_from = request.args.get("order_from", "").strip()
    order_to = request.args.get("order_to", "").strip()
    work_order_id = request.args.get("work_order_id", "").strip()

    sql = "SELECT schedules.*, (SELECT COUNT(*) FROM schedule_photos WHERE schedule_photos.schedule_id=schedules.id) AS photo_count FROM schedules WHERE 1=1"
    params = []

    if q:
        sql += " AND (customer LIKE ? OR site_name LIKE ? OR model LIKE ? OR product_group LIKE ? OR memo LIKE ? OR detail_content LIKE ? OR customer_contact LIKE ? OR delivery_type LIKE ? OR completion_label LIKE ? OR completion_memo LIKE ?)"
        like = f"%{q}%"
        params += [like, like, like, like, like, like, like, like, like, like]

    if status:
        sql += " AND status = ?"
        params.append(status)

    if date_from:
        sql += " AND due_date >= ?"
        params.append(date_from)

    if date_to:
        sql += " AND due_date <= ?"
        params.append(date_to)

    if order_from:
        sql += " AND order_date >= ?"
        params.append(order_from)

    if order_to:
        sql += " AND order_date <= ?"
        params.append(order_to)

    if work_order_id:
        sql += " AND work_order_id = ?"
        params.append(work_order_id)

    sql += " ORDER BY due_date ASC, order_date ASC, product_group ASC, customer ASC, id ASC"

    conn = get_conn()
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])

@app.route("/api/schedules", methods=["POST"])
def create_schedule():
    ensure_schedules_work_order_id_column()
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = get_conn()
    cur = conn.execute("""
        INSERT INTO schedules
        (order_date, due_date, delivery_date, customer, site_name, product_group, model, detail_content, customer_contact, qty, status, delivery_type, calendar_status, memo, work_order_id, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        data.get("order_date", ""),
        data.get("due_date", ""),
        data.get("delivery_date", "") or data.get("due_date", ""),
        data.get("customer", ""),
        data.get("site_name", ""),
        data.get("product_group", ""),
        data.get("model", ""),
        data.get("detail_content", ""),
        data.get("customer_contact", ""),
        int(data.get("qty") or 1),
        data.get("status", "접수"),
        data.get("delivery_type", "납품"),
        "출고가능" if data.get("status") == "완료" else "예정",
        data.get("memo", ""),
        data.get("work_order_id", "") or "",
        now,
        now
    ))
    new_id = cur.lastrowid
    conn.execute("INSERT OR IGNORE INTO customers(name, active, created_at) VALUES(?, 1, ?)", (data.get("customer", ""), now))
    upsert_calendar_event(conn, new_id, data)
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "id": new_id})

@app.route("/api/schedules/<int:item_id>", methods=["PUT"])
def update_schedule(item_id):
    ensure_schedules_work_order_id_column()
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    calendar_status = "출고가능" if data.get("status") == "완료" else "예정"

    conn = get_conn()
    set_clause = """
        UPDATE schedules SET
            order_date=?,
            due_date=?,
            delivery_date=?,
            customer=?,
            site_name=?,
            product_group=?,
            model=?,
            detail_content=?,
            customer_contact=?,
            qty=?,
            status=?,
            delivery_type=?,
            calendar_status=?,
            memo=?,
            updated_at=?
    """
    params = [
        data.get("order_date", ""),
        data.get("due_date", ""),
        data.get("delivery_date", "") or data.get("due_date", ""),
        data.get("customer", ""),
        data.get("site_name", ""),
        data.get("product_group", ""),
        data.get("model", ""),
        data.get("detail_content", ""),
        data.get("customer_contact", ""),
        int(data.get("qty") or 1),
        data.get("status", "접수"),
        data.get("delivery_type", "납품"),
        calendar_status,
        data.get("memo", ""),
        now,
    ]
    if "work_order_id" in data:
        set_clause = set_clause.rstrip() + ",\n            work_order_id=?"
        params.append(data.get("work_order_id", "") or "")
    set_clause += "\n        WHERE id=?"
    params.append(item_id)
    conn.execute(set_clause, params)
    conn.execute("INSERT OR IGNORE INTO customers(name, active, created_at) VALUES(?, 1, ?)", (data.get("customer", ""), now))
    upsert_calendar_event(conn, item_id, data)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/schedules/<int:item_id>/status", methods=["PUT"])
def update_status(item_id):
    data = request.json or {}
    status = data.get("status", "접수")
    calendar_status = "출고가능" if status == "완료" else "예정"
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = get_conn()
    conn.execute("UPDATE schedules SET status=?, calendar_status=?, updated_at=? WHERE id=?", (status, calendar_status, now, item_id))

    row = conn.execute("SELECT * FROM schedules WHERE id=?", (item_id,)).fetchone()
    if row:
        upsert_calendar_event(conn, item_id, row_to_dict(row))

    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/schedules/<int:item_id>", methods=["DELETE"])
def delete_schedule(item_id):
    conn = get_conn()
    # 연결된 작업지시서 ID 조회 후 함께 삭제
    row = conn.execute("SELECT work_order_id FROM schedules WHERE id=?", (item_id,)).fetchone()
    wo_id = (row["work_order_id"] if row else None) or ""
    conn.execute("DELETE FROM schedules WHERE id=?", (item_id,))
    conn.execute("DELETE FROM calendar_events WHERE schedule_id=?", (item_id,))
    conn.execute("DELETE FROM schedule_photos WHERE schedule_id=?", (item_id,))
    if wo_id:
        conn.execute("DELETE FROM work_orders_web WHERE id=?", (wo_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/active_summary", methods=["GET"])
def active_summary():
    # selected_date는 기존 호환용: 납기일 기준으로 처리
    selected_date = request.args.get("selected_date", "").strip()
    selected_due_date = request.args.get("selected_due_date", "").strip() or selected_date
    selected_order_date = request.args.get("selected_order_date", "").strip()
    selected_type = request.args.get("selected_type", "").strip()  # order 또는 due
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    order_from = request.args.get("order_from", "").strip()
    order_to = request.args.get("order_to", "").strip()

    groups = ["작업요청서","엔토브","로이도어","리젠도어","나인도어","클래식도어","대문","폴딩도어","시스템도어","T/T창","중문","방충망"]
    result = {g: {"total": 0, "selected": 0, "hold": 0} for g in groups}

    conn = get_conn()

    sql = "SELECT product_group, SUM(qty) total_qty FROM schedules WHERE status IN ('접수','제작중')"
    params = []
    if date_from:
        sql += " AND due_date >= ?"
        params.append(date_from)
    if date_to:
        sql += " AND due_date <= ?"
        params.append(date_to)
    if order_from:
        sql += " AND order_date >= ?"
        params.append(order_from)
    if order_to:
        sql += " AND order_date <= ?"
        params.append(order_to)
    sql += " GROUP BY product_group"
    for r in conn.execute(sql, params).fetchall():
        if r["product_group"] in result:
            result[r["product_group"]]["total"] = r["total_qty"] or 0

    if selected_order_date or (selected_type == "order" and selected_date):
        d = selected_order_date or selected_date
        sql2 = "SELECT product_group, SUM(qty) total_qty FROM schedules WHERE status IN ('접수','제작중') AND order_date=? GROUP BY product_group"
        selected_params = (d,)
    elif selected_due_date:
        d = selected_due_date
        sql2 = "SELECT product_group, SUM(qty) total_qty FROM schedules WHERE status IN ('접수','제작중') AND due_date=? GROUP BY product_group"
        selected_params = (d,)
    else:
        sql2 = ""
        selected_params = ()

    if sql2:
        for r in conn.execute(sql2, selected_params).fetchall():
            if r["product_group"] in result:
                result[r["product_group"]]["selected"] = r["total_qty"] or 0

    hold_sql = "SELECT product_group, SUM(qty) total_qty FROM schedules WHERE status='보류'"
    hold_params = []
    if date_from:
        hold_sql += " AND due_date >= ?"
        hold_params.append(date_from)
    if date_to:
        hold_sql += " AND due_date <= ?"
        hold_params.append(date_to)
    if order_from:
        hold_sql += " AND order_date >= ?"
        hold_params.append(order_from)
    if order_to:
        hold_sql += " AND order_date <= ?"
        hold_params.append(order_to)
    hold_sql += " GROUP BY product_group"
    for r in conn.execute(hold_sql, hold_params).fetchall():
        if r["product_group"] in result:
            result[r["product_group"]]["hold"] = r["total_qty"] or 0

    conn.close()
    return jsonify(result)

@app.route("/api/schedules/<int:item_id>/complete", methods=["POST"])
def complete_schedule(item_id):
    label = (request.form.get("completion_label") or "").strip()
    completion_memo = (request.form.get("completion_memo") or "").strip()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = get_conn()
    conn.execute("""
        UPDATE schedules SET
            status='완료',
            calendar_status='출고가능',
            completion_label=?,
            completion_memo=?,
            completed_at=?,
            updated_at=?
        WHERE id=?
    """, (label, completion_memo, now, now, item_id))

    row = conn.execute("SELECT * FROM schedules WHERE id=?", (item_id,)).fetchone()
    if row:
        upsert_calendar_event(conn, item_id, row_to_dict(row))

    conn.commit()

    photos = request.files.getlist("photos")
    for f in photos:
        if not f or not f.filename:
            continue
        year_folder = datetime.now().strftime("%Y")
        year_dir = UPLOAD_DIR / year_folder
        year_dir.mkdir(exist_ok=True)
        photo_name = make_photo_filename(item_id, f.filename, row)
        save_path = year_dir / photo_name
        try:
            save_resized_photo(f, save_path)
            filename = f"{year_folder}/{photo_name}"
            conn.execute("""
                INSERT INTO schedule_photos(schedule_id, filename, original_name, memo, created_at)
                VALUES (?, ?, ?, ?, ?)
            """, (item_id, filename, f.filename, "", now))
        except Exception:
            pass
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/schedules/<int:item_id>/completion_info", methods=["PUT"])
def update_completion_info(item_id):
    data = request.json or {}
    label = (data.get("completion_label") or "").strip()
    completion_memo = (data.get("completion_memo") or "").strip()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = get_conn()
    row = conn.execute("SELECT * FROM schedules WHERE id=?", (item_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "작업을 찾을 수 없습니다."}), 404

    conn.execute("""
        UPDATE schedules SET
            completion_label=?,
            completion_memo=?,
            updated_at=?
        WHERE id=?
    """, (label, completion_memo, now, item_id))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/completed_search", methods=["GET"])
def completed_search():
    q = request.args.get("q", "").strip()
    sql = """
        SELECT schedules.*,
               (SELECT COUNT(*) FROM schedule_photos WHERE schedule_photos.schedule_id=schedules.id) AS photo_count
        FROM schedules
        WHERE status='완료'
    """
    params = []
    if q:
        sql += """
            AND (customer LIKE ? OR completion_label LIKE ? OR model LIKE ? OR site_name LIKE ?
                 OR product_group LIKE ? OR completion_memo LIKE ? OR memo LIKE ?)
        """
        like = f"%{q}%"
        params += [like, like, like, like, like, like, like]
    sql += " ORDER BY completed_at DESC, due_date ASC, customer ASC, id DESC"
    conn = get_conn()
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])

@app.route("/api/schedules/<int:item_id>/photos", methods=["GET"])
def list_photos(item_id):
    conn = get_conn()
    rows = conn.execute("SELECT * FROM schedule_photos WHERE schedule_id=? ORDER BY id DESC", (item_id,)).fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])

@app.route("/api/schedules/<int:item_id>/photos", methods=["POST"])
def upload_photo(item_id):
    f = request.files.get("photo")
    memo = request.form.get("memo", "")
    if not f:
        return jsonify({"ok": False, "error": "사진 파일이 없습니다."}), 400

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    row = conn.execute("SELECT * FROM schedules WHERE id=?", (item_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "작업을 찾을 수 없습니다."}), 404

    year_folder = datetime.now().strftime("%Y")
    year_dir = UPLOAD_DIR / year_folder
    year_dir.mkdir(exist_ok=True)

    photo_name = make_photo_filename(item_id, f.filename, row)
    save_path = year_dir / photo_name

    try:
        save_resized_photo(f, save_path)
    except Exception:
        conn.close()
        return jsonify({"ok": False, "error": "사진 리사이즈/저장 실패"}), 500

    filename = f"{year_folder}/{photo_name}"

    conn.execute("""
        INSERT INTO schedule_photos(schedule_id, filename, original_name, memo, created_at)
        VALUES (?, ?, ?, ?, ?)
    """, (item_id, filename, f.filename, memo, now))
    conn.commit()
    conn.close()

    return jsonify({"ok": True, "filename": filename})

@app.route("/api/customer_info", methods=["GET"])
def customer_info():
    name = request.args.get("name", "").strip()
    if not name:
        return jsonify({"ok": False, "error": "거래처명을 입력하세요."}), 400

    this_year = datetime.now().strftime("%Y")
    conn = get_conn()
    cust = conn.execute("SELECT * FROM customers WHERE name=?", (name,)).fetchone()
    if not cust:
        conn.close()
        return jsonify({"ok": False, "error": "등록되지 않은 거래처입니다."}), 404
    total = conn.execute("SELECT COALESCE(SUM(qty),0) AS qty, COUNT(*) AS cnt FROM schedules WHERE customer=?", (name,)).fetchone()
    year = conn.execute("SELECT COALESCE(SUM(qty),0) AS qty, COUNT(*) AS cnt FROM schedules WHERE customer=? AND substr(COALESCE(due_date, order_date),1,4)=?", (name, this_year)).fetchone()
    last = conn.execute("SELECT MAX(COALESCE(NULLIF(due_date,''), order_date)) AS last_date FROM schedules WHERE customer=?", (name,)).fetchone()
    last_site = conn.execute("SELECT site_name FROM schedules WHERE customer=? AND site_name!='' ORDER BY id DESC LIMIT 1", (name,)).fetchone()
    products = conn.execute("""
        SELECT product_group, COALESCE(SUM(qty),0) AS total_qty, COUNT(*) AS cnt
        FROM schedules
        WHERE customer=?
        GROUP BY product_group
        ORDER BY total_qty DESC, product_group ASC
    """, (name,)).fetchall()

    conn.commit()
    conn.close()
    return jsonify({
        "ok": True,
        "customer": name,
        "active": cust["active"] if cust and "active" in cust.keys() else 1,
        "memo": cust["memo"] if cust and "memo" in cust.keys() else "",
        "payment_note": cust["payment_note"] if cust and "payment_note" in cust.keys() else "",
        "contact_name": cust["contact_name"] if cust and "contact_name" in cust.keys() else "",
        "contact_phone": cust["contact_phone"] if cust and "contact_phone" in cust.keys() else "",
        "contact_memo": cust["contact_memo"] if cust and "contact_memo" in cust.keys() else "",
        "default_region": cust["default_region"] if cust and "default_region" in cust.keys() else "",
        "total_qty": total["qty"] or 0,
        "total_count": total["cnt"] or 0,
        "year": this_year,
        "year_qty": year["qty"] or 0,
        "year_count": year["cnt"] or 0,
        "last_date": last["last_date"] or "",
        "last_site_name": last_site["site_name"] if last_site else "",
        "products": [row_to_dict(r) for r in products]
    })

@app.route("/api/photos/<int:photo_id>", methods=["DELETE"])
def delete_photo(photo_id):
    conn = get_conn()
    row = conn.execute("SELECT * FROM schedule_photos WHERE id=?", (photo_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "사진을 찾을 수 없습니다."}), 404

    try:
        path = UPLOAD_DIR / row["filename"]
        if path.exists():
            path.unlink()
    except Exception:
        pass

    conn.execute("DELETE FROM schedule_photos WHERE id=?", (photo_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/delivery_people", methods=["GET"])
def list_delivery_people():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM delivery_people ORDER BY name ASC").fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])

@app.route("/api/delivery_people", methods=["POST"])
def create_delivery_person():
    data = request.json or {}
    name = (data.get("name") or "").strip()
    phone = (data.get("phone") or "").strip()
    vehicle_no = (data.get("vehicle_no") or "").strip()
    memo = (data.get("memo") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "이름을 입력하세요."}), 400
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    conn.execute("""
        INSERT INTO delivery_people(name, phone, vehicle_no, memo, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?)
        ON CONFLICT(name) DO UPDATE SET
            phone=excluded.phone,
            vehicle_no=excluded.vehicle_no,
            memo=excluded.memo,
            updated_at=excluded.updated_at
    """, (name, phone, vehicle_no, memo, now, now))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/delivery_people/<int:person_id>", methods=["DELETE"])
def delete_delivery_person(person_id):
    conn = get_conn()
    conn.execute("DELETE FROM delivery_people WHERE id=?", (person_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/calendar_events", methods=["GET"])
def list_calendar_events():
    start = request.args.get("start", "").strip()
    end = request.args.get("end", "").strip()
    event_type = request.args.get("event_type", "").strip()
    status = request.args.get("status", "").strip()

    sql = """
        SELECT calendar_events.*,
               schedules.order_date, schedules.due_date, schedules.delivery_date, schedules.customer, schedules.site_name,
               schedules.product_group, schedules.model, schedules.qty, schedules.status AS production_status,
               schedules.delivery_type, schedules.memo, schedules.completion_label, schedules.completion_memo, schedules.completed_at,
               schedules.work_order_id,
               (SELECT COUNT(*) FROM schedule_photos WHERE schedule_photos.schedule_id=schedules.id) AS photo_count
        FROM calendar_events
        LEFT JOIN schedules ON schedules.id = calendar_events.schedule_id
        WHERE (calendar_events.event_type IN ('납품예정','시공예정','화물예정','용차예정','택배예정','실측','A/S')
               OR calendar_events.schedule_id IS NULL)
    """
    params = []
    if start:
        sql += " AND calendar_events.event_date >= ?"
        params.append(start)
    if end:
        sql += " AND calendar_events.event_date <= ?"
        params.append(end)
    if event_type:
        sql += " AND calendar_events.event_type = ?"
        params.append(event_type)
    if status:
        sql += " AND calendar_events.status = ?"
        params.append(status)

    sql += " ORDER BY calendar_events.event_date ASC, COALESCE(schedules.site_name, calendar_events.manual_site_name, '') ASC, calendar_events.event_type ASC, schedules.customer ASC, calendar_events.id ASC"
    conn = get_conn()
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])

@app.route("/api/calendar_events/<int:event_id>", methods=["PUT"])
def update_calendar_event(event_id):
    data = request.json or {}
    event_date = (data.get("event_date") or "").strip()
    status = (data.get("status") or "").strip()
    delivery_status = (data.get("delivery_status") or "").strip()
    delivery_start = (data.get("delivery_start") or "").strip()
    delivery_end = (data.get("delivery_end") or "").strip()
    delivery_memo = (data.get("delivery_memo") or "").strip()
    site_name = data.get("site_name")
    person_id_raw = data.get("delivery_person_id")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = get_conn()
    row = conn.execute("SELECT * FROM calendar_events WHERE id=?", (event_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "일정을 찾을 수 없습니다."}), 404

    person_id = None
    person_name = ""
    person_phone = ""
    person_vehicle = ""
    if person_id_raw:
        try:
            person_id = int(person_id_raw)
            p = conn.execute("SELECT * FROM delivery_people WHERE id=?", (person_id,)).fetchone()
            if p:
                person_name = p["name"] or ""
                person_phone = p["phone"] or ""
                person_vehicle = p["vehicle_no"] or ""
            else:
                person_id = None
        except Exception:
            person_id = None

    sets = []
    params = []
    if event_date:
        sets.append("event_date=?")
        params.append(event_date)
    if status:
        sets.append("status=?")
        params.append(status)
    if delivery_status:
        sets.append("delivery_status=?")
        params.append(delivery_status)
    if "delivery_start" in data:
        sets.append("delivery_start=?")
        params.append(delivery_start)
    if "delivery_end" in data:
        sets.append("delivery_end=?")
        params.append(delivery_end)
    if "delivery_memo" in data:
        sets.append("delivery_memo=?")
        params.append(delivery_memo)
    if "delivery_person_id" in data:
        sets += ["delivery_person_id=?", "delivery_person_name=?", "delivery_phone=?", "delivery_vehicle_no=?"]
        params += [person_id, person_name, person_phone, person_vehicle]

    # 독립 일정(schedule_id 없음)은 manual_* 컬럼 직접 수정
    if not row["schedule_id"]:
        if site_name is not None:
            sets.append("manual_site_name=?")
            params.append((site_name or "").strip())
        if "manual_customer" in data:
            sets.append("manual_customer=?")
            params.append((data.get("manual_customer") or "").strip())
        if "manual_product" in data:
            sets.append("manual_product=?")
            params.append((data.get("manual_product") or "").strip())
        if "manual_qty" in data:
            sets.append("manual_qty=?")
            params.append(int(data.get("manual_qty") or 0))

    if sets:
        sets.append("updated_at=?")
        params.append(now)
        params.append(event_id)
        conn.execute(f"UPDATE calendar_events SET {', '.join(sets)} WHERE id=?", params)

    # 달력 날짜를 바꾸면 실제 납기일(delivery_date)도 같이 변경해서 생산일과 분리한다.
    if row["schedule_id"]:
        if event_date and site_name is not None:
            conn.execute("UPDATE schedules SET delivery_date=?, site_name=?, updated_at=? WHERE id=?", (event_date, site_name.strip(), now, row["schedule_id"]))
        elif event_date:
            conn.execute("UPDATE schedules SET delivery_date=?, updated_at=? WHERE id=?", (event_date, now, row["schedule_id"]))
        elif site_name is not None:
            conn.execute("UPDATE schedules SET site_name=?, updated_at=? WHERE id=?", (site_name.strip(), now, row["schedule_id"]))

    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/calendar_events", methods=["POST"])
@require_perm("calendar")
def create_calendar_event():
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    event_date = (data.get("event_date") or "").strip()
    if not event_date:
        return jsonify({"ok": False, "error": "일정일을 입력하세요"}), 400
    event_type = (data.get("event_type") or "기타").strip()
    manual_customer = (data.get("customer") or "").strip()
    manual_site_name = (data.get("site_name") or "").strip()
    manual_product = (data.get("product") or "").strip()
    manual_qty = int(data.get("qty") or 0)
    manual_memo = (data.get("memo") or "").strip()
    person_id_raw = data.get("delivery_person_id")

    conn = get_conn()
    person_id = None
    person_name = ""
    person_phone = ""
    person_vehicle = ""
    if person_id_raw:
        try:
            person_id = int(person_id_raw)
            p = conn.execute("SELECT * FROM delivery_people WHERE id=?", (person_id,)).fetchone()
            if p:
                person_name = p["name"] or ""
                person_phone = p["phone"] or ""
                person_vehicle = p["vehicle_no"] or ""
            else:
                person_id = None
        except Exception:
            person_id = None

    cur = conn.execute("""
        INSERT INTO calendar_events (
            event_date, event_type, status, delivery_status,
            manual_customer, manual_site_name, manual_product, manual_qty, manual_memo,
            delivery_person_id, delivery_person_name, delivery_phone, delivery_vehicle_no,
            created_at, updated_at
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        event_date, event_type, "예정", "예정",
        manual_customer, manual_site_name, manual_product, manual_qty, manual_memo,
        person_id, person_name, person_phone, person_vehicle,
        now, now
    ))
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    return jsonify({"ok": True, "id": new_id})


@app.route("/api/calendar_events/<int:event_id>", methods=["DELETE"])
@require_perm("calendar")
def delete_calendar_event(event_id):
    conn = get_conn()
    row = conn.execute("SELECT schedule_id FROM calendar_events WHERE id=?", (event_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "일정을 찾을 수 없습니다"}), 404
    if row["schedule_id"]:
        conn.close()
        return jsonify({"ok": False, "error": "생산 스케줄 연동 일정은 여기서 삭제할 수 없습니다"}), 400
    conn.execute("DELETE FROM calendar_events WHERE id=?", (event_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/price_data", methods=["GET"])
def price_data():
    conn = get_conn()
    products = {}
    rows = conn.execute("""
        SELECT * FROM price_products
        WHERE COALESCE(active,1)=1
        ORDER BY CASE series
            WHEN '엔토브' THEN 1
            WHEN '로이도어' THEN 2
            WHEN '리젠도어' THEN 3
            WHEN '나인도어' THEN 4
            WHEN '클래식도어' THEN 5
            WHEN '방화문' THEN 6
            WHEN '대문' THEN 7
            ELSE 99
        END, model ASC
    """).fetchall()
    for p in rows:
        pid = p["id"]
        prices = {}
        for t in conn.execute("SELECT * FROM price_types WHERE product_id=? ORDER BY sort_order ASC, id ASC", (pid,)).fetchall():
            prices[t["type_name"]] = {
                "factory": t["factory_price"] or 0,
                "consumer": t["consumer_price"] or 0
            }
        options = []
        for o in conn.execute("SELECT * FROM price_options WHERE product_id=? ORDER BY sort_order ASC, id ASC", (pid,)).fetchall():
            options.append({"key": f"option{o['sort_order']}", "name": o["option_name"], "price": o["option_price"] or 0})
        notes = [n["note"] for n in conn.execute("SELECT * FROM price_notes WHERE product_id=? ORDER BY sort_order ASC, id ASC", (pid,)).fetchall()]
        product = {
            "series": p["series"],
            "model": p["model"],
            "imageKey": p["image_key"] or p["model"],
            "prices": prices,
            "options": options,
            "notes": notes
        }
        products.setdefault(p["series"], []).append(product)
    regions = []
    for r in conn.execute("SELECT * FROM price_regions ORDER BY region_name ASC, id ASC").fetchall():
        regions.append({
            "name": r["region_name"],
            "category": r["category"],
            "time": r["travel_time"],
            "grade": r["region_grade"],
            "price": r["region_price"] or 0
        })
    conn.close()
    preferred = ["엔토브", "로이도어", "리젠도어", "나인도어", "클래식도어", "방화문", "대문"]
    ordered = {}
    for k in preferred:
        if k in products:
            ordered[k] = products.pop(k)
    for k in sorted(products.keys()):
        ordered[k] = products[k]
    return jsonify({"products": ordered, "regions": regions})

@app.route("/api/price_reimport", methods=["POST"])
def price_reimport():
    conn = get_conn()
    conn.execute("DELETE FROM price_notes")
    conn.execute("DELETE FROM price_options")
    conn.execute("DELETE FROM price_types")
    conn.execute("DELETE FROM price_products")
    conn.execute("DELETE FROM price_regions")
    import_price_csv_if_empty(conn)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/journals", methods=["GET"])
def list_journals():
    q = request.args.get("q", "").strip()
    journal_type = request.args.get("journal_type", "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    status = request.args.get("status", "").strip()
    sql = "SELECT * FROM journal_entries WHERE 1=1"
    params = []
    if q:
        sql += " AND (customer LIKE ? OR contact_name LIKE ? OR contact_phone LIKE ? OR title LIKE ? OR content LIKE ? OR follow_up LIKE ? OR writer LIKE ?)"
        like = f"%{q}%"
        params += [like, like, like, like, like, like, like]
    if journal_type:
        sql += " AND journal_type=?"
        params.append(journal_type)
    if date_from:
        sql += " AND journal_date>=?"
        params.append(date_from)
    if date_to:
        sql += " AND journal_date<=?"
        params.append(date_to)
    if status:
        sql += " AND status=?"
        params.append(status)
    sql += " ORDER BY journal_date DESC, id DESC"
    conn = get_conn()
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])

@app.route("/api/journals", methods=["POST"])
def create_journal():
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    cur = conn.execute("""
        INSERT INTO journal_entries
        (journal_date, journal_type, channel, customer, contact_name, contact_phone, title, content, follow_up, status, writer, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        data.get("journal_date", ""), data.get("journal_type", "상담일지"), data.get("channel", ""),
        data.get("customer", ""), data.get("contact_name", ""), data.get("contact_phone", ""),
        data.get("title", ""), data.get("content", ""), data.get("follow_up", ""),
        data.get("status", "진행중"), data.get("writer", ""), now, now
    ))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "id": cur.lastrowid})

@app.route("/api/journals/<int:item_id>", methods=["PUT"])
def update_journal(item_id):
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    conn.execute("""
        UPDATE journal_entries SET
            journal_date=?, journal_type=?, channel=?, customer=?, contact_name=?, contact_phone=?,
            title=?, content=?, follow_up=?, status=?, writer=?, updated_at=?
        WHERE id=?
    """, (
        data.get("journal_date", ""), data.get("journal_type", "상담일지"), data.get("channel", ""),
        data.get("customer", ""), data.get("contact_name", ""), data.get("contact_phone", ""),
        data.get("title", ""), data.get("content", ""), data.get("follow_up", ""),
        data.get("status", "진행중"), data.get("writer", ""), now, item_id
    ))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/journals/<int:item_id>", methods=["DELETE"])
def delete_journal(item_id):
    conn = get_conn()
    conn.execute("DELETE FROM journal_entries WHERE id=?", (item_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

# ── 영업일지 API (customers 테이블 통합) ────────────────────────────────────
_SALES_SELECT = """
    SELECT c.id,
           c.name        AS customer,
           c.writer,
           c.default_region AS region,
           c.address,
           c.interest_items,
           c.existing_supplier,
           c.sales_status   AS status,
           c.sales_notes    AS notes,
           c.contact_name,
           c.contact_phone,
           c.memo,
           c.payment_note,
           c.created_at,
           (SELECT COUNT(*) FROM sales_visits WHERE customer_id=c.id) AS visit_count,
           (SELECT MAX(visit_date) FROM sales_visits WHERE customer_id=c.id) AS last_visit
    FROM customers c
    WHERE c.active=1
"""

@app.route("/api/sales/customers", methods=["GET"])
@require_perm("journal_sales")
def sales_customers_list():
    q         = request.args.get("q",         "").strip()
    status    = request.args.get("status",    "").strip()
    writer    = request.args.get("writer",    "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to   = request.args.get("date_to",   "").strip()
    conn = get_conn()
    if date_from or date_to:
        sql = """
            SELECT DISTINCT
                c.id, c.name AS customer, c.writer, c.default_region AS region,
                c.address, c.interest_items, c.existing_supplier,
                c.sales_status AS status, c.sales_notes AS notes,
                c.contact_name, c.contact_phone, c.memo, c.payment_note, c.created_at,
                (SELECT COUNT(*) FROM sales_visits WHERE customer_id=c.id) AS visit_count,
                (SELECT MAX(visit_date) FROM sales_visits WHERE customer_id=c.id) AS last_visit
            FROM customers c
            JOIN sales_visits v ON v.customer_id=c.id
            WHERE c.active=1
        """
        params = []
        if date_from:
            sql += " AND v.visit_date >= ?"
            params.append(date_from)
        if date_to:
            sql += " AND v.visit_date <= ?"
            params.append(date_to)
        if writer:
            sql += " AND c.writer=?"
            params.append(writer)
        if status:
            sql += " AND c.sales_status=?"
            params.append(status)
        if q:
            sql += " AND (c.name LIKE ? OR c.default_region LIKE ? OR c.interest_items LIKE ?)"
            like = f"%{q}%"
            params += [like, like, like]
        sql += " ORDER BY v.visit_date DESC, c.id DESC"
    else:
        sql = _SALES_SELECT
        params = []
        if q:
            sql += " AND (c.name LIKE ? OR c.default_region LIKE ? OR c.interest_items LIKE ? OR c.writer LIKE ? OR c.sales_notes LIKE ? OR c.memo LIKE ?)"
            like = f"%{q}%"
            params += [like, like, like, like, like, like]
        if status:
            sql += " AND c.sales_status=?"
            params.append(status)
        if writer:
            sql += " AND c.writer=?"
            params.append(writer)
        sql += " ORDER BY c.id DESC"
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return jsonify(rows)

@app.route("/api/sales/customers", methods=["POST"])
@require_perm("journal_sales")
def sales_customers_create():
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    name = (data.get("customer") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "거래처명 필수"}), 400
    conn = get_conn()
    existing = conn.execute("SELECT id FROM customers WHERE name=?", (name,)).fetchone()
    if existing:
        # 이미 있으면 영업 정보만 업데이트
        cid = existing["id"]
        conn.execute("""
            UPDATE customers SET writer=?,default_region=?,address=?,contact_phone=?,interest_items=?,
                existing_supplier=?,sales_status=?,sales_notes=?,active=1
            WHERE id=?
        """, (data.get("writer",""), data.get("region",""), data.get("address",""),
              data.get("contact_phone",""), data.get("interest_items",""),
              data.get("existing_supplier",""), data.get("status","영업중"),
              data.get("notes",""), cid))
    else:
        cur = conn.execute("""
            INSERT INTO customers (name, writer, default_region, address, contact_phone, interest_items,
                existing_supplier, sales_status, sales_notes, active, created_at)
            VALUES (?,?,?,?,?,?,?,?,?,1,?)
        """, (name, data.get("writer",""), data.get("region",""), data.get("address",""),
              data.get("contact_phone",""), data.get("interest_items",""),
              data.get("existing_supplier",""), data.get("status","영업중"),
              data.get("notes",""), now))
        cid = cur.lastrowid
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "id": cid})

@app.route("/api/sales/customers/<int:cid>", methods=["PUT"])
@require_perm("journal_sales")
def sales_customers_update(cid):
    data = request.json or {}
    new_name = (data.get("customer") or "").strip()
    conn = get_conn()
    old = conn.execute("SELECT name FROM customers WHERE id=?", (cid,)).fetchone()
    update_fields = [
        ("writer",            data.get("writer","")),
        ("name",              new_name),
        ("default_region",    data.get("region","")),
        ("address",           data.get("address","")),
        ("contact_phone",     data.get("contact_phone","")),
        ("interest_items",    data.get("interest_items","")),
        ("existing_supplier", data.get("existing_supplier","")),
        ("sales_status",      data.get("status","영업중")),
        ("sales_notes",       data.get("notes","")),
    ]
    created_at = (data.get("created_at") or "").strip()
    if created_at:
        update_fields.append(("created_at", created_at))
    set_clause = ", ".join(f"{f}=?" for f, _ in update_fields) + " WHERE id=?"
    conn.execute(f"UPDATE customers SET {set_clause}",
                 [v for _, v in update_fields] + [cid])
    # 거래처명 변경 시 schedules도 연동 업데이트
    if old and new_name and old["name"] != new_name:
        conn.execute("UPDATE schedules SET customer=? WHERE customer=?", (new_name, old["name"]))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

# ── 영업일지 리드 API ─────────────────────────────────────────────────────

@app.route("/api/sales/leads", methods=["GET"])
@require_perm("journal_sales")
def sales_leads_list():
    q         = request.args.get("q",         "").strip()
    status    = request.args.get("status",    "").strip()
    writer    = request.args.get("writer",    "").strip()
    date_from = request.args.get("date_from", "").strip()
    date_to   = request.args.get("date_to",   "").strip()
    conn = get_conn()
    if date_from or date_to:
        sql = """
            SELECT DISTINCT l.*,
                (SELECT COUNT(*) FROM sales_visits WHERE lead_id=l.id) AS visit_count,
                (SELECT MAX(visit_date) FROM sales_visits WHERE lead_id=l.id) AS last_visit,
                c.name AS customer_name
            FROM sales_leads l
            LEFT JOIN customers c ON c.id=l.linked_customer_id
            JOIN sales_visits v ON v.lead_id=l.id
            WHERE l.active=1
        """
    else:
        sql = """
            SELECT l.*,
                (SELECT COUNT(*) FROM sales_visits WHERE lead_id=l.id) AS visit_count,
                (SELECT MAX(visit_date) FROM sales_visits WHERE lead_id=l.id) AS last_visit,
                c.name AS customer_name
            FROM sales_leads l
            LEFT JOIN customers c ON c.id=l.linked_customer_id
            WHERE l.active=1
        """
    params = []
    if date_from:
        sql += " AND v.visit_date >= ?"
        params.append(date_from)
    if date_to:
        sql += " AND v.visit_date <= ?"
        params.append(date_to)
    if q:
        sql += " AND (l.name LIKE ? OR l.region LIKE ? OR l.interest_items LIKE ? OR l.phone LIKE ?)"
        params += [f"%{q}%"] * 4
    if writer:
        sql += " AND l.writer LIKE ?"
        params.append(f"%{writer}%")
    if status:
        sql += " AND l.sales_status=?"
        params.append(status)
    sql += " ORDER BY last_visit DESC, l.created_at DESC, l.id DESC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])


@app.route("/api/sales/leads", methods=["POST"])
@require_perm("journal_sales")
def sales_leads_create():
    data = request.json or {}
    now  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    name = (data.get("name") or "").strip()
    if not name:
        return jsonify({"ok": False, "error": "이름 필수"}), 400
    conn = get_conn()
    cur = conn.execute("""
        INSERT INTO sales_leads (writer,name,phone,region,address,interest_items,
            existing_supplier,sales_status,notes,active,created_at)
        VALUES (?,?,?,?,?,?,?,?,?,1,?)
    """, (data.get("writer",""), name, data.get("phone",""), data.get("region",""),
          data.get("address",""), data.get("interest_items",""),
          data.get("existing_supplier",""), data.get("status","영업중"),
          data.get("notes",""), now))
    conn.commit()
    lid = cur.lastrowid
    conn.close()
    return jsonify({"ok": True, "id": lid})


@app.route("/api/sales/leads/<int:lid>", methods=["PUT"])
@require_perm("journal_sales")
def sales_leads_update(lid):
    data = request.json or {}
    fields = [
        ("writer",            data.get("writer","")),
        ("name",              (data.get("name") or "").strip()),
        ("phone",             data.get("phone","")),
        ("region",            data.get("region","")),
        ("address",           data.get("address","")),
        ("interest_items",    data.get("interest_items","")),
        ("existing_supplier", data.get("existing_supplier","")),
        ("sales_status",      data.get("status","영업중")),
        ("notes",             data.get("notes","")),
    ]
    created_at = (data.get("created_at") or "").strip()
    if created_at:
        fields.append(("created_at", created_at))
    set_clause = ", ".join(f"{f}=?" for f, _ in fields) + " WHERE id=?"
    conn = get_conn()
    conn.execute(f"UPDATE sales_leads SET {set_clause}", [v for _, v in fields] + [lid])
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/sales/leads/<int:lid>", methods=["DELETE"])
@require_perm("journal_sales")
def sales_leads_delete(lid):
    conn = get_conn()
    conn.execute("DELETE FROM sales_visits WHERE lead_id=?", (lid,))
    conn.execute("UPDATE sales_leads SET active=0 WHERE id=?", (lid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/sales/leads/<int:lid>/visits", methods=["GET"])
@require_perm("journal_sales")
def sales_lead_visits_list(lid):
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM sales_visits WHERE lead_id=? ORDER BY visit_date DESC, id DESC", (lid,)
    ).fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])


@app.route("/api/sales/leads/<int:lid>/visits", methods=["POST"])
@require_perm("journal_sales")
def sales_lead_visits_create(lid):
    data = request.json or {}
    now  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    cur = conn.execute("""
        INSERT INTO sales_visits (lead_id,visit_date,visitor,sales_items,content,created_at)
        VALUES (?,?,?,?,?,?)
    """, (lid, data.get("visit_date",""), data.get("visitor",""),
          data.get("sales_items",""), data.get("content",""), now))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "id": cur.lastrowid})


@app.route("/api/sales/leads/<int:lid>/orders", methods=["GET"])
@require_perm("journal_sales")
def sales_lead_orders(lid):
    conn = get_conn()
    lead = conn.execute("SELECT linked_customer_id FROM sales_leads WHERE id=?", (lid,)).fetchone()
    if not lead or not lead["linked_customer_id"]:
        conn.close()
        return jsonify([])
    cust = conn.execute("SELECT name FROM customers WHERE id=?", (lead["linked_customer_id"],)).fetchone()
    if not cust:
        conn.close()
        return jsonify([])
    rows = conn.execute("""
        SELECT id,order_date,due_date,delivery_date,product_group,model,qty,status,site_name
        FROM schedules WHERE customer=? ORDER BY order_date DESC, id DESC
    """, (cust["name"],)).fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])


@app.route("/api/sales/leads/<int:lid>/link", methods=["POST"])
@require_perm("journal_sales")
def sales_lead_link(lid):
    data = request.json or {}
    customer_name = (data.get("customer_name") or "").strip()
    if not customer_name:
        return jsonify({"ok": False, "error": "거래처명 필수"}), 400
    now  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    lead = conn.execute("SELECT * FROM sales_leads WHERE id=?", (lid,)).fetchone()
    if not lead:
        conn.close()
        return jsonify({"ok": False, "error": "영업 정보 없음"}), 404
    existing = conn.execute("SELECT id FROM customers WHERE name=?", (customer_name,)).fetchone()
    if existing:
        cid = existing["id"]
        conn.execute("""
            UPDATE customers SET writer=?,default_region=?,address=?,contact_phone=?,
                interest_items=?,existing_supplier=?,sales_status=?,sales_notes=?,active=1
            WHERE id=?
        """, (lead["writer"], lead["region"], lead["address"], lead["phone"],
              lead["interest_items"], lead["existing_supplier"],
              lead["sales_status"], lead["notes"], cid))
    else:
        cur = conn.execute("""
            INSERT INTO customers (name,writer,default_region,address,contact_phone,
                interest_items,existing_supplier,sales_status,sales_notes,active,created_at)
            VALUES (?,?,?,?,?,?,?,?,?,1,?)
        """, (customer_name, lead["writer"], lead["region"], lead["address"], lead["phone"],
              lead["interest_items"], lead["existing_supplier"],
              lead["sales_status"], lead["notes"], now))
        cid = cur.lastrowid
    conn.execute("UPDATE sales_leads SET linked_customer_id=? WHERE id=?", (cid, lid))
    conn.commit()
    cust = conn.execute("SELECT name FROM customers WHERE id=?", (cid,)).fetchone()
    conn.close()
    return jsonify({"ok": True, "customer_id": cid, "customer_name": cust["name"] if cust else customer_name})


@app.route("/api/sales/leads/<int:lid>/unlink", methods=["POST"])
@require_perm("journal_sales")
def sales_lead_unlink(lid):
    conn = get_conn()
    conn.execute("UPDATE sales_leads SET linked_customer_id=NULL WHERE id=?", (lid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/sales/leads/bulk_status", methods=["POST"])
@require_perm("journal_sales")
def sales_leads_bulk_status():
    data = request.json or {}
    ids    = [int(i) for i in (data.get("ids") or []) if str(i).isdigit()]
    status = (data.get("status") or "").strip()
    if not ids or not status:
        return jsonify({"ok": False, "error": "ids and status required"}), 400
    placeholders = ",".join("?" for _ in ids)
    conn = get_conn()
    conn.execute(f"UPDATE sales_leads SET sales_status=? WHERE id IN ({placeholders})", [status] + list(ids))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


# ── 일지 사진 첨부 ───────────────────────────────────────────────────────

JOURNAL_UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "static", "uploads", "journal")
os.makedirs(JOURNAL_UPLOAD_DIR, exist_ok=True)
ALLOWED_PHOTO_EXT = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".heic", ".heif"}

def _journal_photo_perm(etype):
    perm = "journal_consult" if etype == "consult" else "journal_sales"
    if not user_has_perm(perm):
        return False
    return True

@app.route("/api/journal/photos", methods=["GET"])
def journal_photos_list():
    if not session.get("user_id"):
        return jsonify({"error": "로그인 필요"}), 401
    etype = request.args.get("entity_type","").strip()
    eid   = request.args.get("entity_id","").strip()
    if not etype or not eid:
        return jsonify([])
    if not _journal_photo_perm(etype):
        return jsonify({"error": "권한 없음"}), 403
    conn = get_conn()
    rows = conn.execute(
        "SELECT * FROM journal_photos WHERE entity_type=? AND entity_id=? ORDER BY id ASC",
        (etype, int(eid))
    ).fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])


@app.route("/api/journal/photos", methods=["POST"])
def journal_photos_upload():
    if not session.get("user_id"):
        return jsonify({"ok": False, "error": "로그인 필요"}), 401
    etype = request.form.get("entity_type","").strip()
    eid   = request.form.get("entity_id","").strip()
    if not etype or not eid:
        return jsonify({"ok": False, "error": "entity_type, entity_id 필수"}), 400
    if not _journal_photo_perm(etype):
        return jsonify({"ok": False, "error": "권한 없음"}), 403
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    saved = []
    for f in request.files.getlist("photos"):
        if not f or not f.filename:
            continue
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ALLOWED_PHOTO_EXT:
            continue
        fname = uuid.uuid4().hex + ext
        f.save(os.path.join(JOURNAL_UPLOAD_DIR, fname))
        conn.execute(
            "INSERT INTO journal_photos (entity_type,entity_id,filename,original_name,uploaded_at) VALUES (?,?,?,?,?)",
            (etype, int(eid), fname, f.filename, now)
        )
        saved.append(fname)
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "saved": len(saved)})


@app.route("/api/journal/photos/<int:pid>", methods=["DELETE"])
def journal_photos_delete(pid):
    if not session.get("user_id"):
        return jsonify({"ok": False, "error": "로그인 필요"}), 401
    conn = get_conn()
    row_check = conn.execute("SELECT entity_type, filename FROM journal_photos WHERE id=?", (pid,)).fetchone()
    if not row_check:
        conn.close()
        return jsonify({"ok": True})
    if not _journal_photo_perm(row_check["entity_type"]):
        conn.close()
        return jsonify({"ok": False, "error": "권한 없음"}), 403
    row = row_check
    if row:
        fpath = os.path.join(JOURNAL_UPLOAD_DIR, row["filename"])
        if os.path.exists(fpath):
            os.remove(fpath)
        conn.execute("DELETE FROM journal_photos WHERE id=?", (pid,))
        conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/sales/customers/bulk_status", methods=["POST"])
@require_perm("journal_sales")
def sales_customers_bulk_status():
    data   = request.json or {}
    ids    = data.get("ids", [])
    status = (data.get("status") or "").strip()
    if not ids or not status:
        return jsonify({"ok": False, "error": "ids와 status가 필요합니다"}), 400
    conn = get_conn()
    placeholders = ",".join("?" for _ in ids)
    conn.execute(
        f"UPDATE customers SET sales_status=? WHERE id IN ({placeholders})",
        [status] + list(ids)
    )
    conn.commit()
    updated = conn.execute(
        f"SELECT COUNT(*) FROM customers WHERE id IN ({placeholders})",
        list(ids)
    ).fetchone()[0]
    conn.close()
    return jsonify({"ok": True, "updated": updated})

@app.route("/api/sales/customers/<int:cid>", methods=["DELETE"])
@require_perm("journal_sales")
def sales_customers_delete(cid):
    conn = get_conn()
    conn.execute("DELETE FROM sales_visits WHERE customer_id=?", (cid,))
    # 생산이력이 있으면 비활성화만, 없으면 완전 삭제
    has_orders = conn.execute(
        "SELECT 1 FROM schedules WHERE customer=(SELECT name FROM customers WHERE id=?)", (cid,)
    ).fetchone()
    if has_orders:
        conn.execute("UPDATE customers SET active=0, sales_status='거래이탈' WHERE id=?", (cid,))
    else:
        conn.execute("DELETE FROM customers WHERE id=?", (cid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/sales/customers/<int:cid>/visits", methods=["GET"])
@require_perm("journal_sales")
def sales_visits_list(cid):
    conn = get_conn()
    rows = [dict(r) for r in conn.execute(
        "SELECT * FROM sales_visits WHERE customer_id=? ORDER BY visit_date DESC, id DESC", (cid,)
    ).fetchall()]
    conn.close()
    return jsonify(rows)

@app.route("/api/sales/customers/<int:cid>/visits", methods=["POST"])
@require_perm("journal_sales")
def sales_visits_create(cid):
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    cur = conn.execute("""
        INSERT INTO sales_visits (customer_id,visit_date,visitor,sales_items,content,created_at)
        VALUES (?,?,?,?,?,?)
    """, (cid, data.get("visit_date",""), data.get("visitor",""), data.get("sales_items",""), data.get("content",""), now))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "id": cur.lastrowid})

@app.route("/api/sales/visits/<int:vid>", methods=["PUT"])
@require_perm("journal_sales")
def sales_visits_update(vid):
    data = request.json or {}
    conn = get_conn()
    conn.execute("""
        UPDATE sales_visits SET visit_date=?, visitor=?, sales_items=?, content=? WHERE id=?
    """, (data.get("visit_date",""), data.get("visitor",""), data.get("sales_items",""), data.get("content",""), vid))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/sales/visits/<int:vid>", methods=["DELETE"])
@require_perm("journal_sales")
def sales_visits_delete(vid):
    conn = get_conn()
    conn.execute("DELETE FROM sales_visits WHERE id=?", (vid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/sales/customers/<int:cid>/orders", methods=["GET"])
@require_perm("journal_sales")
def sales_customer_orders(cid):
    conn = get_conn()
    c = conn.execute("SELECT name FROM customers WHERE id=?", (cid,)).fetchone()
    if not c:
        conn.close()
        return jsonify([])
    rows = [dict(r) for r in conn.execute("""
        SELECT order_date, due_date, delivery_date, product_group, model, qty, status,
               site_name, memo, completion_label, completed_at
        FROM schedules WHERE customer=?
        ORDER BY order_date DESC, id DESC LIMIT 50
    """, (c["name"],)).fetchall()]
    conn.close()
    return jsonify(rows)


# ═══════════════════════════════════════════
# 상담일지 API
# ═══════════════════════════════════════════

@app.route("/api/consult/clients", methods=["GET"])
@require_perm("journal_consult")
def consult_clients_list():
    q         = request.args.get("q","").strip()
    writer    = request.args.get("writer","").strip()
    date_from = request.args.get("date_from","").strip()
    date_to   = request.args.get("date_to","").strip()
    conn   = get_conn()
    params = []
    if date_from or date_to:
        sql = """
            SELECT DISTINCT c.*,
                (SELECT COUNT(*) FROM consult_visits WHERE client_id=c.id) AS visit_count,
                (SELECT MAX(visit_date) FROM consult_visits WHERE client_id=c.id) AS last_visit,
                cu.name AS linked_customer_name
            FROM consult_clients c
            JOIN consult_visits v ON v.client_id=c.id
            LEFT JOIN customers cu ON cu.id=c.linked_customer_id
            WHERE c.active=1
        """
        if date_from:
            sql += " AND v.visit_date >= ?"
            params.append(date_from)
        if date_to:
            sql += " AND v.visit_date <= ?"
            params.append(date_to)
        if writer:
            sql += " AND c.writer=?"
            params.append(writer)
        if q:
            sql += " AND (c.name LIKE ? OR c.region LIKE ? OR c.interest_items LIKE ? OR c.phone LIKE ?)"
            like = f"%{q}%"
            params += [like, like, like, like]
        sql += " ORDER BY v.visit_date DESC, c.id DESC"
    else:
        sql = """
            SELECT c.*,
                (SELECT COUNT(*) FROM consult_visits WHERE client_id=c.id) AS visit_count,
                (SELECT MAX(visit_date) FROM consult_visits WHERE client_id=c.id) AS last_visit,
                cu.name AS linked_customer_name
            FROM consult_clients c
            LEFT JOIN customers cu ON cu.id=c.linked_customer_id
            WHERE c.active=1
        """
        if q:
            sql += " AND (c.name LIKE ? OR c.region LIKE ? OR c.interest_items LIKE ? OR c.phone LIKE ? OR c.writer LIKE ? OR c.notes LIKE ?)"
            like = f"%{q}%"
            params += [like, like, like, like, like, like]
        if writer:
            sql += " AND c.writer=?"
            params.append(writer)
        sql += " ORDER BY last_visit DESC, c.created_at DESC, c.id DESC"
    rows = [dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return jsonify(rows)


@app.route("/api/consult/clients", methods=["POST"])
@require_perm("journal_consult")
def consult_clients_create():
    data = request.json or {}
    now  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    cur  = conn.execute("""
        INSERT INTO consult_clients
            (writer,name,phone,first_consult_date,channel,region,address,interest_items,notes,active,created_at)
        VALUES (?,?,?,?,?,?,?,?,?,1,?)
    """, (
        data.get("writer",""), data.get("name",""), data.get("phone",""),
        data.get("first_consult_date",""), data.get("channel","경로확인"),
        data.get("region",""), data.get("address",""),
        data.get("interest_items",""), data.get("notes",""), now
    ))
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    return jsonify({"ok": True, "id": new_id})


@app.route("/api/consult/clients/<int:cid>", methods=["PUT"])
@require_perm("journal_consult")
def consult_clients_update(cid):
    data   = request.json or {}
    conn   = get_conn()
    fields = [
        ("writer",             data.get("writer","")),
        ("name",               data.get("name","")),
        ("phone",              data.get("phone","")),
        ("first_consult_date", data.get("first_consult_date","")),
        ("channel",            data.get("channel","경로확인")),
        ("region",             data.get("region","")),
        ("address",            data.get("address","")),
        ("interest_items",     data.get("interest_items","")),
        ("notes",              data.get("notes","")),
    ]
    set_clause = ", ".join(f"{f}=?" for f, _ in fields) + " WHERE id=?"
    conn.execute(f"UPDATE consult_clients SET {set_clause}", [v for _, v in fields] + [cid])
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/consult/clients/<int:cid>", methods=["DELETE"])
@require_perm("journal_consult")
def consult_clients_delete(cid):
    conn = get_conn()
    conn.execute("DELETE FROM consult_visits WHERE client_id=?", (cid,))
    conn.execute("DELETE FROM consult_clients WHERE id=?", (cid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/consult/clients/<int:cid>/visits", methods=["GET"])
@require_perm("journal_consult")
def consult_visits_list(cid):
    conn = get_conn()
    rows = [dict(r) for r in conn.execute(
        "SELECT * FROM consult_visits WHERE client_id=? ORDER BY visit_date DESC, id DESC", (cid,)
    ).fetchall()]
    conn.close()
    return jsonify(rows)


@app.route("/api/consult/clients/<int:cid>/visits", methods=["POST"])
@require_perm("journal_consult")
def consult_visits_create(cid):
    data = request.json or {}
    now  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    cur  = conn.execute("""
        INSERT INTO consult_visits (client_id,visit_date,sales_items,content,consult_type,delivery_type,created_at)
        VALUES (?,?,?,?,?,?,?)
    """, (cid, data.get("visit_date",""), data.get("sales_items",""), data.get("content",""),
          data.get("consult_type",""), data.get("delivery_type",""), now))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "id": cur.lastrowid})


@app.route("/api/consult/visits/<int:vid>", methods=["PUT"])
@require_perm("journal_consult")
def consult_visits_update(vid):
    data = request.json or {}
    conn = get_conn()
    conn.execute("""
        UPDATE consult_visits SET visit_date=?, sales_items=?, content=?, consult_type=?, delivery_type=?
        WHERE id=?
    """, (data.get("visit_date",""), data.get("sales_items",""), data.get("content",""),
          data.get("consult_type",""), data.get("delivery_type",""), vid))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/consult/visits/<int:vid>", methods=["DELETE"])
@require_perm("journal_consult")
def consult_visits_delete(vid):
    conn = get_conn()
    conn.execute("DELETE FROM consult_visits WHERE id=?", (vid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/consult/clients/<int:cid>/orders", methods=["GET"])
@require_perm("journal_consult")
def consult_client_orders(cid):
    conn = get_conn()
    c = conn.execute("""
        SELECT cc.name, cu.name AS customer_name
        FROM consult_clients cc
        LEFT JOIN customers cu ON cu.id=cc.linked_customer_id
        WHERE cc.id=?
    """, (cid,)).fetchone()
    if not c:
        conn.close()
        return jsonify([])
    search_name = c["customer_name"] or ""
    if not search_name:
        conn.close()
        return jsonify([])
    rows = [dict(r) for r in conn.execute("""
        SELECT order_date, due_date, delivery_date, product_group, model, qty, status,
               site_name, memo, completion_label, completed_at
        FROM schedules WHERE customer=?
        ORDER BY order_date DESC, id DESC LIMIT 50
    """, (search_name,)).fetchall()]
    conn.close()
    return jsonify(rows)


@app.route("/api/consult/clients/<int:cid>/link", methods=["POST"])
@require_perm("journal_consult")
def consult_client_link(cid):
    data = request.json or {}
    now  = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    customer_id = data.get("customer_id")
    if customer_id:
        conn.execute("UPDATE consult_clients SET linked_customer_id=? WHERE id=?", (customer_id, cid))
    else:
        c = conn.execute("SELECT * FROM consult_clients WHERE id=?", (cid,)).fetchone()
        if not c:
            conn.close()
            return jsonify({"ok": False, "error": "상담자를 찾을 수 없습니다"}), 404
        customer_name = (data.get("name") or c["name"] or "").strip()
        if not customer_name:
            conn.close()
            return jsonify({"ok": False, "error": "거래처명을 입력해주세요"}), 400
        existing = conn.execute("SELECT id FROM customers WHERE name=?", (customer_name,)).fetchone()
        if existing:
            customer_id = existing["id"]
        else:
            cur = conn.execute("""
                INSERT INTO customers
                    (name, writer, default_region, address, interest_items, sales_status, sales_notes, active, created_at)
                VALUES (?,?,?,?,?,'거래중','',1,?)
            """, (customer_name, c["writer"] or "", c["region"] or "",
                  c["address"] or "", c["interest_items"] or "", now))
            customer_id = cur.lastrowid
        conn.execute("UPDATE consult_clients SET linked_customer_id=? WHERE id=?", (customer_id, cid))
    conn.commit()
    cu = conn.execute("SELECT name FROM customers WHERE id=?", (customer_id,)).fetchone()
    conn.close()
    return jsonify({"ok": True, "customer_id": customer_id, "customer_name": cu["name"] if cu else ""})


@app.route("/api/consult/clients/<int:cid>/unlink", methods=["POST"])
@require_perm("journal_consult")
def consult_client_unlink(cid):
    conn = get_conn()
    conn.execute("UPDATE consult_clients SET linked_customer_id=NULL WHERE id=?", (cid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/login", methods=["GET"])
def login_page():
    return render_template("login.html")

@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.json or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    if not username or not password:
        return jsonify({"ok": False, "error": "아이디와 비밀번호를 입력하세요."}), 400
    conn = get_conn()
    user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
    if not user or not check_password_hash(user["password_hash"] or "", password):
        conn.close()
        return jsonify({"ok": False, "error": "아이디 또는 비밀번호가 맞지 않습니다."}), 401
    if int(user["active"] or 0) != 1:
        conn.close()
        return jsonify({"ok": False, "error": "사용중지된 아이디입니다."}), 403
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute("UPDATE users SET last_login=?, updated_at=? WHERE id=?", (now, now, user["id"]))
    conn.commit()
    conn.close()
    session.clear()
    session.permanent = True
    session["user_id"] = user["id"]
    session["username"] = user["username"]
    session["name"] = user["name"] or user["username"]
    session["user_group"] = user["user_group"] if "user_group" in user.keys() else ""
    session["role"] = user["role"] or "일반"
    session["permissions"] = user["permissions"] or ""
    session["last_activity"] = datetime.now().timestamp()
    next_url = data.get("next") or "/"
    return jsonify({"ok": True, "next": next_url})

@app.route("/api/change_password", methods=["POST"])
def api_change_password():
    data = request.json or {}
    username = (data.get("username") or "").strip()
    current_pw = data.get("current_password") or ""
    new_pw = data.get("new_password") or ""
    if not username or not current_pw or not new_pw:
        return jsonify({"ok": False, "error": "모든 항목을 입력하세요."}), 400
    if len(new_pw) < 4:
        return jsonify({"ok": False, "error": "새 비밀번호는 4자 이상이어야 합니다."}), 400
    conn = get_conn()
    user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
    if not user or not check_password_hash(user["password_hash"] or "", current_pw):
        conn.close()
        return jsonify({"ok": False, "error": "아이디 또는 현재 비밀번호가 맞지 않습니다."}), 401
    if int(user["active"] or 0) != 1:
        conn.close()
        return jsonify({"ok": False, "error": "사용중지된 아이디입니다."}), 403
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute("UPDATE users SET password_hash=?, updated_at=? WHERE id=?",
                 (generate_password_hash(new_pw), now, user["id"]))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login_page"))

@app.route("/sw.js")
def service_worker():
    resp = send_from_directory("static", "sw.js")
    resp.headers["Content-Type"] = "application/javascript"
    resp.headers["Service-Worker-Allowed"] = "/"
    return resp

@app.route("/api/push/vapid-public-key")
@require_login
def api_push_vapid_key():
    return jsonify({"public_key": VAPID_KEYS.get("public_key", "")})

@app.route("/api/push/subscribe", methods=["POST"])
@require_login
def api_push_subscribe():
    data = request.json or {}
    endpoint = data.get("endpoint", "")
    keys = data.get("keys", {})
    p256dh = keys.get("p256dh", "")
    auth = keys.get("auth", "")
    if not endpoint or not p256dh or not auth:
        return jsonify({"ok": False}), 400
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    conn.execute("""
        INSERT INTO push_subscriptions (user_id, endpoint, p256dh, auth, created_at)
        VALUES (?, ?, ?, ?, ?)
        ON CONFLICT(endpoint) DO UPDATE SET user_id=excluded.user_id, p256dh=excluded.p256dh, auth=excluded.auth
    """, (session.get("user_id"), endpoint, p256dh, auth, now))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/push/unsubscribe", methods=["POST"])
@require_login
def api_push_unsubscribe():
    data = request.json or {}
    endpoint = data.get("endpoint", "")
    conn = get_conn()
    conn.execute("DELETE FROM push_subscriptions WHERE endpoint=?", (endpoint,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/users")
@require_perm("user_manage")
def users_page():
    return render_template("users.html", permission_labels=PERMISSION_LABELS, permission_levels=PERMISSION_LEVELS, all_permissions=ALL_PERMISSIONS, permission_tree=PERMISSION_TREE)

@app.route("/api/org/unregistered")
@require_write_perm("user_manage")
def api_org_unregistered():
    conn = get_conn()
    members = conn.execute("""
        SELECT m.id, m.name, m.rank, d.label as dept_label
        FROM org_members m
        LEFT JOIN org_departments d ON m.dept_id = d.id
        ORDER BY d.sort_order, m.sort_order
    """).fetchall()
    existing_names = {r[0] for r in conn.execute("SELECT name FROM users WHERE name IS NOT NULL").fetchall()}
    conn.close()
    # 이름 중복 제거(겸직) + 미등록자만
    seen = set()
    result = []
    for m in members:
        if m["name"] and m["name"] not in existing_names and m["name"] not in seen:
            seen.add(m["name"])
            result.append({"id": m["id"], "name": m["name"], "rank": m["rank"], "dept_label": m["dept_label"] or ""})
    return jsonify(result)


def korean_to_eng(text):
    CHO  = ['r','R','s','e','E','f','a','q','Q','t','T','d','w','W','c','z','x','v','g']
    JUNG = ['k','o','i','O','j','p','u','P','h','hk','ho','hl','y','n','nj','np','nl','b','m','ml','l']
    JONG = ['','r','R','rt','s','sw','sg','e','f','fr','fa','fq','ft','fx','fv','fg','a','q','qt','t','T','d','w','c','z','x','v','g']
    result = []
    for ch in text:
        code = ord(ch)
        if 0xAC00 <= code <= 0xD7A3:
            code -= 0xAC00
            result.append(CHO[code // (21*28)] + JUNG[(code % (21*28)) // 28] + JONG[code % 28])
        else:
            result.append(ch)
    return ''.join(result)


@app.route("/api/users/bulk_from_org", methods=["POST"])
@require_write_perm("user_manage")
def api_users_bulk_from_org():
    data = request.json or {}
    members = data.get("members", [])  # [{name, rank, dept_label, user_group}]
    conn = get_conn()
    groups = {g["group_name"]: g for g in [dict(r) for r in conn.execute("SELECT * FROM check_groups").fetchall()]}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ok, skipped = 0, 0
    for m in members:
        name = (m.get("name") or "").strip()
        if not name:
            continue
        username = name
        password = korean_to_eng(name) + "1234"
        user_group = m.get("user_group") or ""
        g = groups.get(user_group, {})
        perms = permissions_to_db(g.get("permissions") or {}, "일반")
        try:
            conn.execute("""
                INSERT INTO users(username, password_hash, name, user_group, role, permissions, active, created_at, updated_at)
                VALUES (?,?,?,?,?,?,?,?,?)
            """, (username, generate_password_hash(password), name, user_group, "일반", perms, 1, now, now))
            ok += 1
        except Exception:
            skipped += 1
    conn.commit(); conn.close()
    return jsonify({"ok": True, "registered": ok, "skipped": skipped})


@app.route("/api/users", methods=["GET"])
@require_perm("user_manage")
def api_users_list():
    conn = get_conn()
    rows = conn.execute("""
        SELECT u.id, u.username, u.name, COALESCE(u.user_group,'') AS user_group,
               u.role, u.permissions, u.active, u.last_login, u.created_at, u.updated_at,
               COALESCE(d.label,'') AS dept_label,
               COALESCE(m.rank,'') AS rank,
               COALESCE(m.phone,'') AS phone,
               COALESCE(m.hire_date,'') AS hire_date
        FROM users u
        LEFT JOIN org_members m ON m.name = u.name
        LEFT JOIN org_departments d ON d.id = m.dept_id
        ORDER BY u.active DESC, u.username ASC
    """).fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])

@app.route("/api/users", methods=["POST"])
@require_perm("user_manage")
def api_users_create():
    data = request.json or {}
    username = (data.get("username") or "").strip()
    password = data.get("password") or ""
    name = (data.get("name") or username).strip()
    role = (data.get("role") or "일반").strip()
    user_group = (data.get("user_group") or "사무실").strip()
    permissions = data.get("permissions") or {}
    active = 1 if int(data.get("active", 1) or 0) else 0
    if not username or not password:
        return jsonify({"ok": False, "error": "아이디와 비밀번호를 입력하세요."}), 400
    perm_text = permissions_to_db(permissions, role)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    try:
        conn.execute("""
            INSERT INTO users(username, password_hash, name, user_group, role, permissions, active, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (username, generate_password_hash(password), name, user_group, role, perm_text, active, now, now))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"ok": False, "error": "이미 존재하는 아이디입니다."}), 409
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/users/<int:user_id>", methods=["PUT"])
@require_perm("user_manage")
def api_users_update(user_id):
    data = request.json or {}
    password = data.get("password") or ""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    # 비밀번호만 변경하는 경우
    if password and len(data) == 1:
        conn.execute("UPDATE users SET password_hash=?, updated_at=? WHERE id=?",
                     (generate_password_hash(password), now, user_id))
        conn.commit(); conn.close()
        return jsonify({"ok": True})
    username = (data.get("username") or "").strip()
    name = (data.get("name") or username).strip()
    role = (data.get("role") or "일반").strip()
    user_group = (data.get("user_group") or "사무실").strip()
    permissions = data.get("permissions") or {}
    active = 1 if int(data.get("active", 1) or 0) else 0
    if not username:
        conn.close()
        return jsonify({"ok": False, "error": "아이디를 입력하세요."}), 400
    perm_text = permissions_to_db(permissions, role)
    try:
        if password:
            conn.execute("""
                UPDATE users SET username=?, name=?, user_group=?, role=?, permissions=?, active=?, password_hash=?, updated_at=? WHERE id=?
            """, (username, name, user_group, role, perm_text, active, generate_password_hash(password), now, user_id))
        else:
            conn.execute("""
                UPDATE users SET username=?, name=?, user_group=?, role=?, permissions=?, active=?, updated_at=? WHERE id=?
            """, (username, name, user_group, role, perm_text, active, now, user_id))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"ok": False, "error": "이미 존재하는 아이디입니다."}), 409
    conn.close()
    return jsonify({"ok": True})



@app.route("/api/users/<int:user_id>", methods=["DELETE"])
def api_users_delete(user_id):
    if session.get("user_id") == user_id:
        return jsonify({"ok": False, "error": "현재 로그인 중인 계정은 삭제할 수 없습니다."}), 400
    conn = get_conn()
    conn.execute("DELETE FROM users WHERE id=?", (user_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/check_groups", methods=["GET"])
@require_write_perm("user_manage")
def api_check_groups_list():
    conn = get_conn()
    rows = conn.execute("""
        SELECT g.*,
               (SELECT COUNT(*) FROM users u WHERE COALESCE(u.user_group,'')=g.group_name) AS user_count
        FROM check_groups g
        ORDER BY COALESCE(g.active,1) DESC, g.sort_order ASC, g.group_name ASC
    """).fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])

@app.route("/api/check_groups", methods=["POST"])
@require_write_perm("user_manage")
def api_check_groups_create():
    data = request.json or {}
    name = (data.get("group_name") or data.get("name") or "").strip()
    memo = (data.get("memo") or "").strip()
    sort_order = int(data.get("sort_order") or 100)
    if not name:
        return jsonify({"ok": False, "error": "그룹명을 입력하세요."}), 400
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    try:
        conn.execute("""
            INSERT INTO check_groups(group_name, memo, sort_order, active, created_at, updated_at)
            VALUES (?, ?, ?, 1, ?, ?)
        """, (name, memo, sort_order, now, now))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"ok": False, "error": "이미 존재하는 그룹입니다."}), 409
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/check_groups/<int:group_id>", methods=["PUT"])
@require_write_perm("user_manage")
def api_check_groups_update(group_id):
    data = request.json or {}
    name = (data.get("group_name") or data.get("name") or "").strip()
    memo = (data.get("memo") or "").strip()
    sort_order = int(data.get("sort_order") or 100)
    active = 1 if int(data.get("active", 1) or 0) else 0
    if not name:
        return jsonify({"ok": False, "error": "그룹명을 입력하세요."}), 400
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    old = conn.execute("SELECT * FROM check_groups WHERE id=?", (group_id,)).fetchone()
    if not old:
        conn.close()
        return jsonify({"ok": False, "error": "그룹을 찾을 수 없습니다."}), 404
    old_name = old["group_name"]
    import json as _json
    permissions = data.get("permissions", None)
    perms_str = _json.dumps(permissions, ensure_ascii=False) if permissions is not None else (old["permissions"] or "")
    try:
        conn.execute("UPDATE check_groups SET group_name=?, memo=?, sort_order=?, active=?, permissions=?, updated_at=? WHERE id=?", (name, memo, sort_order, active, perms_str, now, group_id))
        if name != old_name:
            conn.execute("UPDATE users SET user_group=? WHERE user_group=?", (name, old_name))
            conn.execute("UPDATE check_requests SET to_group=? WHERE to_group=?", (name, old_name))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({"ok": False, "error": "이미 존재하는 그룹명입니다."}), 409
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/check_groups/<int:group_id>", methods=["DELETE"])
@require_write_perm("user_manage")
def api_check_groups_delete(group_id):
    force = request.args.get("force", "") == "1"
    conn = get_conn()
    row = conn.execute("SELECT * FROM check_groups WHERE id=?", (group_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "그룹을 찾을 수 없습니다."}), 404
    name = row["group_name"]
    if name == "총괄":
        conn.close()
        return jsonify({"ok": False, "error": "총괄 그룹은 삭제할 수 없습니다."}), 400
    cnt = conn.execute("SELECT COUNT(*) AS cnt FROM users WHERE COALESCE(user_group,'')=?", (name,)).fetchone()["cnt"]
    if cnt and not force:
        conn.close()
        return jsonify({"ok": False, "has_users": True, "count": cnt, "error": "이 그룹에 속한 사용자가 있습니다."}), 409
    if cnt:
        conn.execute("UPDATE users SET user_group='사무실' WHERE COALESCE(user_group,'')=?", (name,))
    conn.execute("DELETE FROM check_groups WHERE id=?", (group_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/check_request_targets", methods=["GET"])
@require_login
def api_check_request_targets():
    conn = get_conn()
    rows = conn.execute("SELECT id, username, name, COALESCE(user_group,'') AS user_group FROM users WHERE COALESCE(active,1)=1 ORDER BY user_group ASC, username ASC").fetchall()
    group_rows = conn.execute("SELECT group_name, memo FROM check_groups WHERE COALESCE(active,1)=1 ORDER BY sort_order ASC, group_name ASC").fetchall()
    conn.close()
    users = [row_to_dict(r) for r in rows]
    groups = [{"name": r["group_name"], "memo": r["memo"] or ""} for r in group_rows]
    return jsonify({"users": users, "groups": groups})


def _check_request_target_clause(prefix=""):
    uid = session.get("user_id")
    group = session.get("user_group") or ""
    # 총괄 그룹은 개인 요청 + 모든 그룹 요청을 볼 수 있다.
    if group == "총괄":
        return "((to_user_id=? AND COALESCE(to_user_id,0)<>0) OR to_group<>'')", [uid]
    return "((to_user_id=? AND COALESCE(to_user_id,0)<>0) OR (to_group<>'' AND to_group=?))", [uid, group]

@app.route("/api/check_requests/count", methods=["GET"])
@require_login
def api_check_requests_count():
    conn = get_conn()
    where, params = _check_request_target_clause()
    row = conn.execute(f"SELECT COUNT(*) AS cnt FROM check_requests WHERE status='미확인' AND {where}", params).fetchone()
    conn.close()
    return jsonify({"count": (row["cnt"] if row else 0)})

@app.route("/api/check_requests", methods=["GET"])
@require_login
def api_check_requests_list():
    mode = request.args.get("mode", "inbox").strip()
    conn = get_conn()
    if mode == "sent":
        rows = conn.execute("""
            SELECT cr.*, s.customer, s.product_group, s.model, s.detail_content, s.order_date, s.due_date, s.delivery_date,
                   ar.customer AS as_customer, ar.region AS as_region, ar.product_group AS as_product_group, ar.request_content AS as_request_content, ar.status AS as_status
            FROM check_requests cr
            LEFT JOIN schedules s ON s.id=cr.schedule_id
            LEFT JOIN as_requests ar ON ar.id=cr.as_id
            WHERE cr.from_user_id=?
            ORDER BY cr.id DESC
            LIMIT 200
        """, (session.get("user_id"),)).fetchall()
    else:
        where, params = _check_request_target_clause()
        rows = conn.execute(f"""
            SELECT cr.*, s.customer, s.product_group, s.model, s.detail_content, s.order_date, s.due_date, s.delivery_date,
                   ar.customer AS as_customer, ar.region AS as_region, ar.product_group AS as_product_group, ar.request_content AS as_request_content, ar.status AS as_status
            FROM check_requests cr
            LEFT JOIN schedules s ON s.id=cr.schedule_id
            LEFT JOIN as_requests ar ON ar.id=cr.as_id
            WHERE {where}
            ORDER BY CASE cr.status WHEN '미확인' THEN 0 ELSE 1 END, cr.id DESC
            LIMIT 200
        """, params).fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])

@app.route("/api/check_requests", methods=["POST"])
@require_login
def api_check_requests_create():
    data = request.json or {}
    schedule_id = int(data.get("schedule_id") or 0)
    message = (data.get("message") or "").strip()
    target_type = (data.get("target_type") or "group").strip()
    to_user_id = None
    to_user_name = ""
    to_group = ""
    if not schedule_id:
        return jsonify({"ok": False, "error": "작업을 선택하세요."}), 400
    if not message:
        return jsonify({"ok": False, "error": "확인요청 내용을 입력하세요."}), 400
    conn = get_conn()
    sched = conn.execute("SELECT id FROM schedules WHERE id=?", (schedule_id,)).fetchone()
    if not sched:
        conn.close()
        return jsonify({"ok": False, "error": "작업을 찾을 수 없습니다."}), 404
    if target_type == "user":
        try:
            to_user_id = int(data.get("to_user_id") or 0)
        except Exception:
            to_user_id = None
        u = conn.execute("SELECT id, username, name FROM users WHERE id=? AND COALESCE(active,1)=1", (to_user_id,)).fetchone() if to_user_id else None
        if not u:
            conn.close()
            return jsonify({"ok": False, "error": "요청 받을 사용자를 선택하세요."}), 400
        to_user_name = u["name"] or u["username"]
    else:
        to_group = (data.get("to_group") or "사무실").strip()
        if not to_group:
            conn.close()
            return jsonify({"ok": False, "error": "요청 받을 그룹을 선택하세요."}), 400
        g = conn.execute("SELECT id FROM check_groups WHERE group_name=? AND COALESCE(active,1)=1", (to_group,)).fetchone()
        if not g:
            conn.close()
            return jsonify({"ok": False, "error": "요청 받을 그룹을 찾을 수 없습니다."}), 400
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute("""
        INSERT INTO check_requests(schedule_id, from_user_id, from_user_name, to_user_id, to_user_name, to_group, message, status, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, '미확인', ?)
    """, (schedule_id, session.get("user_id"), session.get("name") or session.get("username") or "", to_user_id, to_user_name, to_group, message, now))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/check_requests/<int:req_id>/complete", methods=["PUT"])
@require_login
def api_check_requests_complete(req_id):
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    where, params = _check_request_target_clause()
    row = conn.execute(f"SELECT id FROM check_requests WHERE id=? AND {where}", [req_id] + params).fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "확인요청을 찾을 수 없거나 처리 권한이 없습니다."}), 404
    conn.execute("UPDATE check_requests SET status='확인완료', completed_at=?, completed_by=? WHERE id=?", (now, session.get("name") or session.get("username") or "", req_id))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

def ensure_work_order_mapping_file():
    if WORK_ORDER_MAPPING_PATH.exists():
        return
    with open(WORK_ORDER_MAPPING_PATH, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["field", "label", "cell"])
        writer.writerows(DEFAULT_WORK_ORDER_MAPPING_ROWS)


def read_work_order_rows():
    if not WORK_ORDER_CSV_PATH.exists():
        return []
    with open(WORK_ORDER_CSV_PATH, "r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_work_order_rows(rows):
    WORK_ORDER_DATA_DIR.mkdir(exist_ok=True)
    with open(WORK_ORDER_CSV_PATH, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=WORK_ORDER_FIELDS, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in WORK_ORDER_FIELDS})


def next_work_order_id(rows):
    nums = []
    for r in rows:
        try:
            nums.append(int(r.get("id") or 0))
        except Exception:
            pass
    return (max(nums) + 1) if nums else 1


def read_work_order_mapping():
    ensure_work_order_mapping_file()
    rows = []
    with open(WORK_ORDER_MAPPING_PATH, "r", encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            field = (row.get("field") or "").strip()
            cell = (row.get("cell") or "").strip()
            if field and cell:
                rows.append((field, cell))
    return rows


def make_work_order_output_filename(row):
    def safe(v, default="미입력"):
        text = str(v or "").strip() or default
        text = re.sub(r'[\\/:*?"<>|]+', "_", text)
        text = re.sub(r"\s+", "", text)
        return text[:40] or default
    return f"{safe(row.get('id'))}_{safe(row.get('workDate'))}_{safe(row.get('customer'))}_{safe(row.get('siteName'))}_나인도어비단열.xlsx"



def run_libreoffice_recalc(xlsx_path):
    """LibreOffice로 엑셀 수식을 재계산해서 같은 파일에 저장한다."""
    xlsx_path = Path(xlsx_path)
    if not SOFFICE_PATH.exists():
        return False, f"LibreOffice 실행파일을 찾을 수 없습니다: {SOFFICE_PATH}"

    tmp_dir = WORK_ORDER_OUTPUT_DIR / "_lo_tmp"
    tmp_dir.mkdir(exist_ok=True)
    profile_dir = WORK_ORDER_OUTPUT_DIR / "_lo_profile"
    profile_dir.mkdir(exist_ok=True)

    cmd = [
        str(SOFFICE_PATH),
        "--headless",
        "--invisible",
        "--nodefault",
        "--nofirststartwizard",
        f"-env:UserInstallation=file:///{str(profile_dir).replace(chr(92), '/')}",
        "--convert-to", "xlsx",
        "--outdir", str(tmp_dir),
        str(xlsx_path)
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=90)
        converted = tmp_dir / xlsx_path.name
        if converted.exists():
            shutil.copy2(converted, xlsx_path)
            try:
                converted.unlink()
            except Exception:
                pass
            return True, ""
        msg = (result.stderr or result.stdout or "").strip()
        return False, msg or "LibreOffice 재계산 결과 파일이 생성되지 않았습니다."
    except Exception as e:
        return False, str(e)


def path_to_file_uri(p):
    p = Path(p).resolve()
    try:
        return p.as_uri()
    except Exception:
        s = str(p).replace("\\", "/")
        if ":" in s[:3]:
            return "file:///" + quote(s)
        return "file://" + quote(s)


def write_lo_debug_log(title, cmd, result=None, error=None):
    try:
        log_path = WORK_ORDER_OUTPUT_DIR / "libreoffice_preview_debug.log"
        with open(log_path, "a", encoding="utf-8") as f:
            f.write("\n\n==== " + title + " ====\n")
            f.write("CMD: " + " ".join(map(str, cmd)) + "\n")
            if result is not None:
                f.write("RETURN: " + str(result.returncode) + "\n")
                f.write("STDOUT: " + (result.stdout or "") + "\n")
                f.write("STDERR: " + (result.stderr or "") + "\n")
            if error is not None:
                f.write("ERROR: " + str(error) + "\n")
    except Exception:
        pass


def create_work_order_preview_png(xlsx_path, item_id):
    """엑셀 A1:I50 영역을 PNG/PDF 미리보기로 생성한다."""
    xlsx_path = Path(xlsx_path)
    if not SOFFICE_PATH.exists():
        return "", ""

    WORK_ORDER_PREVIEW_DIR.mkdir(exist_ok=True)
    tmp_xlsx = WORK_ORDER_OUTPUT_DIR / f"_preview_{item_id}.xlsx"

    try:
        for old in [
            WORK_ORDER_OUTPUT_DIR / f"_preview_{item_id}.xlsx",
            WORK_ORDER_PREVIEW_DIR / f"_preview_{item_id}.pdf",
            WORK_ORDER_PREVIEW_DIR / f"{item_id}_preview.pdf",
            WORK_ORDER_PREVIEW_DIR / f"{item_id}_preview.png",
        ]:
            try:
                if old.exists():
                    old.unlink()
            except Exception:
                pass

        shutil.copy2(xlsx_path, tmp_xlsx)

        wb = load_workbook(tmp_xlsx)
        ws = wb.active
        ws.print_area = "A1:I50"
        try:
            ws.page_setup.orientation = "portrait"
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1
            ws.sheet_properties.pageSetUpPr.fitToPage = True
            ws.page_margins.left = 0.2
            ws.page_margins.right = 0.2
            ws.page_margins.top = 0.2
            ws.page_margins.bottom = 0.2
        except Exception:
            pass
        wb.save(tmp_xlsx)

        before_pdfs = {p.name for p in WORK_ORDER_PREVIEW_DIR.glob("*.pdf")}
        profile_dir = WORK_ORDER_OUTPUT_DIR / f"_lo_profile_preview_{item_id}"
        profile_dir.mkdir(exist_ok=True)
        profile_uri = path_to_file_uri(profile_dir)

        cmd = [
            str(SOFFICE_PATH),
            "--headless",
            "--invisible",
            "--nodefault",
            "--nofirststartwizard",
            f"-env:UserInstallation={profile_uri}",
            "--convert-to", "pdf",
            "--outdir", str(WORK_ORDER_PREVIEW_DIR),
            str(tmp_xlsx)
        ]
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=120, cwd=str(WORK_ORDER_OUTPUT_DIR))
        write_lo_debug_log("preview pdf convert", cmd, result)

        pdf_path = WORK_ORDER_PREVIEW_DIR / f"_preview_{item_id}.pdf"
        if not pdf_path.exists():
            new_pdfs = [p for p in WORK_ORDER_PREVIEW_DIR.glob("*.pdf") if p.name not in before_pdfs]
            if new_pdfs:
                pdf_path = sorted(new_pdfs, key=lambda p: p.stat().st_mtime, reverse=True)[0]

        if not pdf_path.exists():
            return "", ""

        final_pdf = WORK_ORDER_PREVIEW_DIR / f"{item_id}_preview.pdf"
        if final_pdf.exists():
            final_pdf.unlink()
        shutil.copy2(pdf_path, final_pdf)

        try:
            if pdf_path.name.startswith("_preview_"):
                pdf_path.unlink()
        except Exception:
            pass

        try:
            import fitz
            doc = fitz.open(final_pdf)
            page = doc.load_page(0)
            pix = page.get_pixmap(matrix=fitz.Matrix(1.6, 1.6), alpha=False)
            png_name = f"{item_id}_preview.png"
            png_path = WORK_ORDER_PREVIEW_DIR / png_name
            if png_path.exists():
                png_path.unlink()
            pix.save(png_path)
            doc.close()
            try:
                final_pdf.unlink()
            except Exception:
                pass
            return f"preview/{png_name}", ""
        except Exception as e:
            write_lo_debug_log("preview png convert failed", [], error=e)
            return f"preview/{final_pdf.name}", ""

    except Exception as e:
        write_lo_debug_log("preview exception", [], error=e)
        return "", ""
    finally:
        try:
            if tmp_xlsx.exists():
                tmp_xlsx.unlink()
        except Exception:
            pass

def create_work_order_xlsx(row):
    ensure_work_order_mapping_file()
    if not WORK_ORDER_TEMPLATE_PATH.exists():
        return "", "", "엑셀 템플릿 파일이 없습니다. work_order_templates/나인도어_비단열.xlsx 파일을 넣어주세요."

    filename = make_work_order_output_filename(row)
    output_path = WORK_ORDER_OUTPUT_DIR / filename

    wb = load_workbook(WORK_ORDER_TEMPLATE_PATH)
    ws = wb.active
    for field, cell in read_work_order_mapping():
        ws[cell] = row.get(field, "")
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass
    wb.save(output_path)

    warnings = []
    ok, msg = run_libreoffice_recalc(output_path)
    if not ok and msg:
        pass

    preview_name, preview_msg = create_work_order_preview_png(output_path, row.get("id", ""))
    if preview_msg:
        pass

    return filename, preview_name, "\n".join(warnings)


@app.route("/api/work_order_csv", methods=["GET"])
def api_work_order_csv_list():
    rows = read_work_order_rows()
    q = request.args.get("q", "").strip().lower()
    status = request.args.get("status", "").strip()
    if status:
        rows = [r for r in rows if (r.get("status") or "") == status]
    if q:
        def hit(r):
            text = " ".join([r.get("customer",""), r.get("siteName",""), r.get("model",""), r.get("memo",""), r.get("specialNotes","")]).lower()
            return q in text
        rows = [r for r in rows if hit(r)]
    rows.sort(key=lambda r: int(r.get("id") or 0), reverse=True)
    return jsonify(rows)


@app.route("/api/work_order_csv/<int:item_id>", methods=["GET"])
def api_work_order_csv_get(item_id):
    rows = read_work_order_rows()
    for r in rows:
        if int(r.get("id") or 0) == item_id:
            return jsonify(r)
    return jsonify({"ok": False, "error": "작업지시서를 찾을 수 없습니다."}), 404


@app.route("/api/work_order_csv", methods=["POST"])
def api_work_order_csv_save():
    data = request.json or {}
    rows = read_work_order_rows()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    item_id = int(data.get("id") or 0)
    is_new = item_id == 0
    if is_new:
        item_id = next_work_order_id(rows)
    row = {k: "" for k in WORK_ORDER_FIELDS}
    row.update({k: str(data.get(k, "") or "") for k in WORK_ORDER_FIELDS if k not in ("created_at","updated_at","writer","output_xlsx","preview_png")})
    row["id"] = str(item_id)
    row["productGroup"] = "나인도어"
    row["templateType"] = "비단열"
    row["updated_at"] = now
    row["writer"] = session.get("name") or session.get("username") or ""
    if not row.get("status"):
        row["status"] = "작성중"
    if is_new:
        row["created_at"] = now
        rows.append(row)
    else:
        found = False
        for i, old in enumerate(rows):
            if int(old.get("id") or 0) == item_id:
                row["created_at"] = old.get("created_at") or now
                found = True
                rows[i] = row
                break
        if not found:
            row["created_at"] = now
            rows.append(row)
    output_name, preview_name, warning = create_work_order_xlsx(row)
    if output_name:
        row["output_xlsx"] = output_name
        row["preview_png"] = preview_name
        for i, old in enumerate(rows):
            if int(old.get("id") or 0) == item_id:
                rows[i] = row
                break
    write_work_order_rows(rows)
    return jsonify({"ok": True, "id": item_id, "output_xlsx": output_name, "preview_png": preview_name, "warning": ("" if preview_name else warning)})


@app.route("/api/work_order_csv/<int:item_id>", methods=["DELETE"])
def api_work_order_csv_delete(item_id):
    rows = read_work_order_rows()
    kept = []
    deleted = False
    output_name = ""
    for r in rows:
        if int(r.get("id") or 0) == item_id:
            deleted = True
            output_name = r.get("output_xlsx","preview_png") or ""
        else:
            kept.append(r)
    if not deleted:
        return jsonify({"ok": False, "error": "작업지시서를 찾을 수 없습니다."}), 404
    write_work_order_rows(kept)
    if output_name:
        try:
            p = WORK_ORDER_OUTPUT_DIR / output_name
            if p.exists():
                p.unlink()
        except Exception:
            pass
    return jsonify({"ok": True})


@app.route("/api/work_order_csv/mapping", methods=["GET"])
def api_work_order_csv_mapping():
    ensure_work_order_mapping_file()
    return jsonify({
        "template_exists": WORK_ORDER_TEMPLATE_PATH.exists(),
        "template_path": str(WORK_ORDER_TEMPLATE_PATH),
        "mapping_path": str(WORK_ORDER_MAPPING_PATH),
        "csv_path": str(WORK_ORDER_CSV_PATH),
        "output_dir": str(WORK_ORDER_OUTPUT_DIR)
    })





@app.route("/api/completed/<item_id>", methods=["POST", "PUT"])
def api_completed_update_fields(item_id):
    data = request.json or request.form or {}
    label_number = (data.get("label_number") or data.get("label") or data.get("라벨번호") or "").strip()
    completed_memo = (data.get("completed_memo") or data.get("complete_memo") or data.get("완료비고") or "").strip()
    memo = (data.get("memo") or data.get("비고") or "").strip()
    status = (data.get("status") or data.get("상태") or "").strip()

    # 제작완료 데이터 파일 후보를 최대한 호환되게 찾음
    candidates = [
        BASE_DIR / "completed.csv",
        DATA_DIR / "completed.csv",
        DATA_DIR / "completed_items.csv",
        DATA_DIR / "production_completed.csv",
        BASE_DIR / "completed_data.csv",
        BASE_DIR / "production_completed.csv",
    ]
    csv_path = None
    for p in candidates:
        if p.exists():
            csv_path = p
            break
    if csv_path is None:
        return jsonify({"ok": False, "error": "제작완료 CSV 파일을 찾지 못했습니다."}), 404

    import csv as _csv
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        rows = list(_csv.DictReader(f))
        fieldnames = list(rows[0].keys()) if rows else []

    if not rows:
        return jsonify({"ok": False, "error": "제작완료 데이터가 없습니다."}), 404

    def ensure_col(name):
        nonlocal fieldnames
        if name not in fieldnames:
            fieldnames.append(name)

    # 흔한 컬럼명 자동 탐색
    id_cols = ["id","ID","번호","no","No","idx"]
    label_cols = ["label_number","라벨번호","라벨","label","labelNo"]
    completed_memo_cols = ["completed_memo","완료비고","완료 비고","complete_memo"]
    memo_cols = ["memo","비고"]
    status_cols = ["status","상태"]

    def find_col(cands, default):
        for c in cands:
            if c in fieldnames:
                return c
        ensure_col(default)
        return default

    id_col = None
    for c in id_cols:
        if c in fieldnames:
            id_col = c
            break
    if id_col is None:
        id_col = fieldnames[0]

    label_col = find_col(label_cols, "라벨번호")
    completed_memo_col = find_col(completed_memo_cols, "완료비고")
    memo_col = find_col(memo_cols, "비고")
    status_col = find_col(status_cols, "상태")

    found = False
    for r in rows:
        if str(r.get(id_col, "")).strip() == str(item_id):
            r[label_col] = label_number
            r[completed_memo_col] = completed_memo
            if memo:
                r[memo_col] = memo
            if status:
                r[status_col] = status
            found = True
            break

    if not found:
        return jsonify({"ok": False, "error": "수정할 제작완료 항목을 찾지 못했습니다."}), 404

    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = _csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    return jsonify({"ok": True})




@app.route("/api/completed_edit/<item_id>", methods=["POST", "PUT"])
def api_completed_edit_v1_update(item_id):
    data = request.json or request.form or {}
    label_number = (data.get("label_number") or "").strip()
    completed_memo = (data.get("completed_memo") or "").strip()
    memo = (data.get("memo") or "").strip()

    # schedules 테이블 직접 업데이트
    conn = get_conn()
    try:
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cur = conn.execute(
            "UPDATE schedules SET completion_label=?, completion_memo=?, updated_at=? WHERE id=?",
            (label_number, completed_memo, now, item_id)
        )
        if cur.rowcount:
            conn.commit()
            conn.close()
            return jsonify({"ok": True})
    except Exception:
        pass
    finally:
        try:
            conn.close()
        except Exception:
            pass

    # 1) SQLite 테이블 후보 먼저 시도
    table_candidates = ["completed", "completed_items", "production_completed"]
    id_cols = ["id", "ID", "번호", "no", "idx"]
    label_cols = ["label_number", "label", "라벨번호", "라벨"]
    completed_memo_cols = ["completed_memo", "complete_memo", "완료비고", "완료 비고"]
    memo_cols = ["memo", "비고"]

    conn = get_conn()
    try:
        tables = [r["name"] for r in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        for table in table_candidates:
            if table not in tables:
                continue
            cols = [r["name"] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()]
            id_col = next((c for c in id_cols if c in cols), None)
            if not id_col:
                continue
            label_col = next((c for c in label_cols if c in cols), None)
            completed_memo_col = next((c for c in completed_memo_cols if c in cols), None)
            memo_col = next((c for c in memo_cols if c in cols), None)

            sets = []
            params = []
            if label_col:
                sets.append(f"{label_col}=?")
                params.append(label_number)
            if completed_memo_col:
                sets.append(f"{completed_memo_col}=?")
                params.append(completed_memo)
            if memo_col:
                sets.append(f"{memo_col}=?")
                params.append(memo)

            if not sets:
                continue

            params.append(item_id)
            cur = conn.execute(f"UPDATE {table} SET {', '.join(sets)} WHERE {id_col}=?", params)
            if cur.rowcount:
                conn.commit()
                conn.close()
                return jsonify({"ok": True, "storage": "sqlite", "table": table})
    except Exception:
        pass
    finally:
        try:
            conn.close()
        except Exception:
            pass

    # 2) CSV 후보 시도
    import csv as _csv
    path_candidates = [
        BASE_DIR / "completed.csv",
        BASE_DIR / "completed_data.csv",
        BASE_DIR / "production_completed.csv",
        DATA_DIR / "completed.csv",
        DATA_DIR / "completed_items.csv",
        DATA_DIR / "production_completed.csv",
        DATA_DIR / "completed_data.csv",
    ]
    csv_path = next((p for p in path_candidates if p.exists()), None)
    if not csv_path:
        return jsonify({"ok": False, "error": "제작완료 저장 파일을 찾지 못했습니다."}), 404

    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = _csv.DictReader(f)
        rows = list(reader)
        fieldnames = list(reader.fieldnames or [])

    if not fieldnames:
        return jsonify({"ok": False, "error": "제작완료 CSV 컬럼을 찾지 못했습니다."}), 404

    def ensure_col(name):
        if name not in fieldnames:
            fieldnames.append(name)

    id_col = next((c for c in id_cols if c in fieldnames), None) or fieldnames[0]
    label_col = next((c for c in label_cols if c in fieldnames), None) or "라벨번호"
    completed_memo_col = next((c for c in completed_memo_cols if c in fieldnames), None) or "완료비고"
    memo_col = next((c for c in memo_cols if c in fieldnames), None) or "비고"
    ensure_col(label_col)
    ensure_col(completed_memo_col)
    ensure_col(memo_col)

    found = False
    for r in rows:
        if str(r.get(id_col, "")).strip() == str(item_id):
            r[label_col] = label_number
            r[completed_memo_col] = completed_memo
            r[memo_col] = memo
            found = True
            break

    if not found:
        return jsonify({"ok": False, "error": "수정할 제작완료 항목을 찾지 못했습니다."}), 404

    with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = _csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    return jsonify({"ok": True, "storage": "csv"})




AS_UPLOAD_DIR = UPLOAD_DIR / "as"
AS_UPLOAD_DIR.mkdir(parents=True, exist_ok=True)



def as_region_group_from_text(*texts):
    text = " ".join([str(t or "") for t in texts])
    text = text.replace(" ", "")
    mapping = [
        ("서울", ["서울", "서울특별시"]),
        ("인천", ["인천", "인천광역시"]),
        ("경기", ["경기", "경기도"]),
        ("강원", ["강원", "강원도", "강원특별자치도"]),
        ("충북", ["충북", "충청북도"]),
        ("충남", ["충남", "충청남도"]),
        ("대전", ["대전", "대전광역시"]),
        ("세종", ["세종", "세종특별자치시"]),
        ("대구", ["대구", "대구광역시"]),
        ("경북", ["경북", "경상북도"]),
        ("경남", ["경남", "경상남도"]),
        ("부산", ["부산", "부산광역시"]),
        ("울산", ["울산", "울산광역시"]),
        ("광주", ["광주", "광주광역시"]),
        ("전북", ["전북", "전라북도", "전북특별자치도"]),
        ("전남", ["전남", "전라남도"]),
        ("제주", ["제주", "제주도", "제주특별자치도"]),
    ]
    for group, keys in mapping:
        if any(k in text for k in keys):
            return group
    return ""



def normalize_phone_v17(value):
    d = re.sub(r"\D", "", str(value or ""))
    if not d:
        return ""
    if len(d) == 8 and re.match(r"^(15|16|18)\d{2}", d):
        return f"{d[:4]}-{d[4:]}"
    if d.startswith("02"):
        if len(d) == 9:
            return f"02-{d[2:5]}-{d[5:]}"
        if len(d) >= 10:
            return f"02-{d[2:6]}-{d[6:10]}"
        return d
    if len(d) == 10:
        return f"{d[:3]}-{d[3:6]}-{d[6:]}"
    if len(d) == 11:
        return f"{d[:3]}-{d[3:7]}-{d[7:]}"
    if len(d) > 11:
        return f"{d[:3]}-{d[3:7]}-{d[7:11]}"
    return d

AS_FIELDS = [
    "region_group","status","receipt_date","request_date","receiver","customer","company_phone","consumer_phone",
    "region","address","product_group","model","detail_content","qty","label_no","request_content","memo","scheduled_date","assigned_to",
    "needed_parts","defect_cause","process_content","improvement","completed_date","completed_by","hold_reason"
]

def ensure_as_model_columns():
    conn = get_conn()
    existing = [r["name"] for r in conn.execute("PRAGMA table_info(as_requests)").fetchall()]
    for col in ["model", "detail_content", "qty", "label_no"]:
        if col not in existing:
            conn.execute(f"ALTER TABLE as_requests ADD COLUMN {col} TEXT DEFAULT ''")
    conn.commit()
    conn.close()

def save_as_photo_file(file_storage, as_id, photo_type):
    if not file_storage or not file_storage.filename:
        return None
    now = datetime.now()
    sub_dir = AS_UPLOAD_DIR / str(now.year) / f"{now.month:02d}" / str(as_id) / photo_type
    sub_dir.mkdir(parents=True, exist_ok=True)
    original = file_storage.filename
    ext = Path(original).suffix.lower()
    if ext not in [".jpg", ".jpeg", ".png", ".webp", ".gif", ".bmp"]:
        ext = ".png"
    filename = f"{uuid.uuid4().hex}{ext}"
    path = sub_dir / filename
    file_storage.save(path)

    try:
        if ext.lower() in [".jpg", ".jpeg", ".png", ".webp", ".bmp"]:
            img = Image.open(path)
            img = ImageOps.exif_transpose(img)
            img.thumbnail((1600, 1600))
            if img.mode not in ("RGB", "RGBA"):
                img = img.convert("RGB")
            img.save(path, quality=86)
    except Exception:
        pass

    return path.relative_to(UPLOAD_DIR).as_posix(), original

@app.route("/api/as/summary", methods=["GET"])
def api_as_summary():
    conn = get_conn()
    rows = conn.execute("SELECT status, COUNT(*) AS cnt FROM as_requests GROUP BY status").fetchall()
    conn.close()
    result = {"접수": 0, "예정": 0, "완료": 0, "보류": 0}
    for r in rows:
        result[r["status"] or "접수"] = r["cnt"]
    return jsonify(result)

@app.route("/api/as", methods=["GET"])
def api_as_list():
    status = (request.args.get("status") or "").strip()
    q = (request.args.get("q") or "").strip()
    region_group = (request.args.get("region_group") or "").strip()
    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()

    where = []
    params = []

    if status:
        where.append("status=?")
        params.append(status)

    if region_group:
        where.append("region_group=?")
        params.append(region_group)

    if q:
        like = f"%{q}%"
        where.append("""(
            customer LIKE ? OR region LIKE ? OR address LIKE ? OR product_group LIKE ?
            OR request_content LIKE ? OR memo LIKE ? OR company_phone LIKE ? OR consumer_phone LIKE ?
        )""")
        params += [like] * 8

    if status == "완료" and start and end:
        where.append("COALESCE(NULLIF(completed_date,''), substr(updated_at,1,10)) BETWEEN ? AND ?")
        params += [start, end]

    sql = "SELECT * FROM as_requests"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += """
        ORDER BY
        CASE status WHEN '접수' THEN 1 WHEN '예정' THEN 2 WHEN '보류' THEN 3 WHEN '완료' THEN 4 ELSE 9 END,
        COALESCE(NULLIF(request_date,''), receipt_date) DESC,
        id DESC
        LIMIT 800
    """

    conn = get_conn()
    rows = [row_to_dict(r) for r in conn.execute(sql, params).fetchall()]
    conn.close()
    return jsonify(rows)

@app.route("/api/as/<int:item_id>", methods=["GET"])
def api_as_detail(item_id):
    conn = get_conn()
    row = conn.execute("SELECT * FROM as_requests WHERE id=?", (item_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "A/S 항목을 찾을 수 없습니다."}), 404

    d = row_to_dict(row)
    photos = conn.execute("SELECT * FROM as_photos WHERE as_id=? ORDER BY id", (item_id,)).fetchall()
    d["receipt_photos"] = [row_to_dict(p) for p in photos if p["photo_type"] == "receipt"]
    d["completion_photos"] = [row_to_dict(p) for p in photos if p["photo_type"] == "completion"]
    conn.close()
    return jsonify(d)

@app.route("/api/as", methods=["POST"])
def api_as_create():
    ensure_as_model_columns()
    data = request.form if request.form else (request.json or {})
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    values = {k: (data.get(k) or "").strip() for k in AS_FIELDS}
    values["status"] = values.get("status") or "접수"
    values["company_phone"] = normalize_phone_v17(values.get("company_phone"))
    values["consumer_phone"] = normalize_phone_v17(values.get("consumer_phone"))
    values["region_group"] = values.get("region_group") or as_region_group_from_text(values.get("region"), values.get("address"))

    if not values.get("receipt_date"):
        values["receipt_date"] = datetime.now().strftime("%Y-%m-%d")
    if not values.get("customer"):
        return jsonify({"ok": False, "error": "거래처명을 입력하세요."}), 400

    conn = get_conn()
    cur = conn.execute("""
        INSERT INTO as_requests (
            status, receipt_date, request_date, receiver, customer, company_phone, consumer_phone,
            region, region_group, address, product_group, model, detail_content, qty, label_no,
            request_content, memo,
            scheduled_date, assigned_to, needed_parts, defect_cause, process_content, improvement,
            completed_date, completed_by, hold_reason, created_at, updated_at
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        values["status"], values["receipt_date"], values["request_date"], values["receiver"],
        values["customer"], values["company_phone"], values["consumer_phone"], values["region"],
        values["region_group"], values["address"], values["product_group"],
        values["model"], values["detail_content"], values["qty"], values["label_no"],
        values["request_content"], values["memo"],
        values["scheduled_date"], values["assigned_to"], values["needed_parts"], values["defect_cause"],
        values["process_content"], values["improvement"], values["completed_date"], values["completed_by"],
        values["hold_reason"], now, now
    ))
    item_id = cur.lastrowid

    for f in request.files.getlist("receipt_photos"):
        saved = save_as_photo_file(f, item_id, "receipt")
        if saved:
            filename, original = saved
            conn.execute("""
                INSERT INTO as_photos(as_id, photo_type, filename, original_name, created_at)
                VALUES (?,?,?,?,?)
            """, (item_id, "receipt", filename, original, now))

    # C/S관리부 직원들에게 A/S 접수 확인요청 자동 발송
    try:
        dept = conn.execute(
            "SELECT id FROM org_departments WHERE label LIKE '%C/S%' ORDER BY id LIMIT 1"
        ).fetchone()
        if dept:
            members = conn.execute(
                "SELECT DISTINCT name FROM org_members WHERE dept_id=?", (dept["id"],)
            ).fetchall()
            member_names = [m["name"] for m in members if m["name"]]
            if member_names:
                ph = ",".join("?" * len(member_names))
                targets = conn.execute(
                    f"SELECT id, name FROM users WHERE name IN ({ph}) AND active=1",
                    member_names
                ).fetchall()
                msg = f"[A/S 접수] {values['customer']}"
                if values.get("product_group"):
                    msg += f" - {values['product_group']}"
                if values.get("model"):
                    msg += f" ({values['model']})"
                from_id = session.get("user_id") or 0
                from_name = session.get("name") or session.get("username") or ""
                target_ids = []
                for t in targets:
                    conn.execute("""
                        INSERT INTO check_requests(
                            schedule_id, as_id, source_type,
                            from_user_id, from_user_name,
                            to_user_id, to_user_name, to_group,
                            message, status, created_at
                        ) VALUES (?,?,?,?,?,?,?,?,?,?,?)
                    """, (None, item_id, "as", from_id, from_name,
                          t["id"], t["name"], "", msg, "미확인", now))
                    target_ids.append(t["id"])
                if target_ids:
                    send_push_notifications(target_ids, "A/S 접수", msg, "/as")
    except Exception:
        pass

    conn.commit()
    conn.close()
    return jsonify({"ok": True, "id": item_id})

@app.route("/api/as/<int:item_id>", methods=["POST", "PUT"])
def api_as_update(item_id):
    ensure_as_model_columns()
    data = request.form if request.form else (request.json or {})
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = get_conn()
    row = conn.execute("SELECT * FROM as_requests WHERE id=?", (item_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "A/S 항목을 찾을 수 없습니다."}), 404

    values = {k: (data.get(k) if data.get(k) is not None else row[k]) for k in AS_FIELDS}
    values = {k: ("" if v is None else str(v).strip()) for k, v in values.items()}

    values["company_phone"] = normalize_phone_v17(values.get("company_phone"))
    values["consumer_phone"] = normalize_phone_v17(values.get("consumer_phone"))
    values["region_group"] = values.get("region_group") or as_region_group_from_text(values.get("region"), values.get("address"))

    if values.get("status") == "완료" and not values.get("completed_date"):
        values["completed_date"] = datetime.now().strftime("%Y-%m-%d")

    conn.execute("""
        UPDATE as_requests SET
            status=?, receipt_date=?, request_date=?, receiver=?, customer=?, company_phone=?, consumer_phone=?,
            region=?, region_group=?, address=?, product_group=?, model=?, detail_content=?, qty=?, label_no=?,
            request_content=?, memo=?, scheduled_date=?, assigned_to=?,
            needed_parts=?, defect_cause=?, process_content=?, improvement=?, completed_date=?, completed_by=?, hold_reason=?,
            updated_at=?
        WHERE id=?
    """, (
        values["status"], values["receipt_date"], values["request_date"], values["receiver"],
        values["customer"], values["company_phone"], values["consumer_phone"], values["region"],
        values["region_group"], values["address"], values["product_group"],
        values["model"], values["detail_content"], values["qty"], values["label_no"],
        values["request_content"], values["memo"],
        values["scheduled_date"], values["assigned_to"], values["needed_parts"], values["defect_cause"],
        values["process_content"], values["improvement"], values["completed_date"], values["completed_by"],
        values["hold_reason"], now, item_id
    ))

    for photo_field, photo_type in [("receipt_photos", "receipt"), ("completion_photos", "completion")]:
        for f in request.files.getlist(photo_field):
            saved = save_as_photo_file(f, item_id, photo_type)
            if saved:
                filename, original = saved
                conn.execute("""
                    INSERT INTO as_photos(as_id, photo_type, filename, original_name, created_at)
                    VALUES (?,?,?,?,?)
                """, (item_id, photo_type, filename, original, now))

    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/as/<int:item_id>/status", methods=["POST", "PUT"])
def api_as_status_update_v22(item_id):
    data = request.json or request.form or {}
    status = (data.get("status") or "").strip()
    if status not in ("접수", "예정", "완료", "보류"):
        return jsonify({"ok": False, "error": "상태값이 올바르지 않습니다."}), 400

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    row = conn.execute("SELECT id FROM as_requests WHERE id=?", (item_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "A/S 항목을 찾을 수 없습니다."}), 404

    fields = ["status=?", "updated_at=?"]
    params = [status, now]

    if status == "완료":
        fields.append("complete_date=COALESCE(NULLIF(complete_date,''), ?)")
        params.append(datetime.now().strftime("%Y-%m-%d"))
    elif status == "예정":
        fields.append("schedule_date=COALESCE(NULLIF(schedule_date,''), ?)")
        params.append(datetime.now().strftime("%Y-%m-%d"))

    params.append(item_id)
    conn.execute(f"UPDATE as_requests SET {', '.join(fields)} WHERE id=?", params)
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "status": status})

@app.route("/api/as/<int:item_id>", methods=["DELETE"])
def api_as_delete(item_id):
    conn = get_conn()
    photos = conn.execute("SELECT filename FROM as_photos WHERE as_id=?", (item_id,)).fetchall()
    for p in photos:
        try:
            path = UPLOAD_DIR / p["filename"]
            if path.exists():
                path.unlink()
        except Exception:
            pass

    conn.execute("DELETE FROM as_photos WHERE as_id=?", (item_id,))
    conn.execute("DELETE FROM as_requests WHERE id=?", (item_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/as/photos/<int:photo_id>", methods=["DELETE"])
def api_as_photo_delete(photo_id):
    conn = get_conn()
    row = conn.execute("SELECT * FROM as_photos WHERE id=?", (photo_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "사진을 찾을 수 없습니다."}), 404

    try:
        path = UPLOAD_DIR / row["filename"]
        if path.exists():
            path.unlink()
    except Exception:
        pass

    conn.execute("DELETE FROM as_photos WHERE id=?", (photo_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})




@app.route("/api/as_check_requests", methods=["POST"])
@require_login
def api_as_check_requests_create_v14():
    data = request.json or {}
    try:
        as_id = int(data.get("as_id") or 0)
    except Exception:
        as_id = 0
    message = (data.get("message") or "").strip()
    target_type = (data.get("target_type") or "group").strip()
    to_user_id = None
    to_user_name = ""
    to_group = ""

    if not as_id:
        return jsonify({"ok": False, "error": "A/S 항목을 선택하세요."}), 400
    if not message:
        return jsonify({"ok": False, "error": "확인요청 내용을 입력하세요."}), 400

    conn = get_conn()
    ar = conn.execute("SELECT id FROM as_requests WHERE id=?", (as_id,)).fetchone()
    if not ar:
        conn.close()
        return jsonify({"ok": False, "error": "A/S 항목을 찾을 수 없습니다."}), 404

    if target_type == "user":
        try:
            to_user_id = int(data.get("to_user_id") or 0)
        except Exception:
            to_user_id = None
        u = conn.execute("SELECT id, username, name FROM users WHERE id=? AND COALESCE(active,1)=1", (to_user_id,)).fetchone() if to_user_id else None
        if not u:
            conn.close()
            return jsonify({"ok": False, "error": "요청 받을 사용자를 선택하세요."}), 400
        to_user_name = u["name"] or u["username"]
    else:
        to_group = (data.get("to_group") or "").strip()
        if not to_group:
            conn.close()
            return jsonify({"ok": False, "error": "요청 받을 그룹을 선택하세요."}), 400

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute("""
        INSERT INTO check_requests(
            schedule_id, as_id, source_type, from_user_id, from_user_name,
            to_user_id, to_user_name, to_group, message, status, created_at
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """, (
        None, as_id, "as", session.get("user_id"),
        session.get("name") or session.get("username") or "",
        to_user_id, to_user_name, to_group, message, "미확인", now
    ))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})



@app.route("/api/construction", methods=["GET"])
@require_login
def api_construction_list():
    ctype = request.args.get("type", "")
    status = request.args.get("status", "")
    conn = get_conn()
    sql = "SELECT * FROM construction_items WHERE 1=1"
    params = []
    if ctype:
        sql += " AND type=?"
        params.append(ctype)
    if status:
        sql += " AND status=?"
        params.append(status)
    sql += " ORDER BY scheduled_date DESC, id DESC"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])

@app.route("/api/construction", methods=["POST"])
@require_login
def api_construction_create():
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    cur = conn.execute("""
        INSERT INTO construction_items
            (type, status, company, scheduled_date, customer, address, manager, content, memo, created_by, created_at, updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        data.get("type", "실측"), data.get("status", "예정"),
        data.get("company", ""), data.get("scheduled_date", ""),
        data.get("customer", ""), data.get("address", ""),
        data.get("manager", ""), data.get("content", ""),
        data.get("memo", ""),
        session.get("name") or session.get("username") or "", now, now
    ))
    conn.commit()
    conn.close()
    return jsonify({"ok": True, "id": cur.lastrowid})

@app.route("/api/construction/<int:item_id>", methods=["PUT"])
@require_login
def api_construction_update(item_id):
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    conn.execute("""
        UPDATE construction_items SET
            type=?, status=?, company=?, scheduled_date=?, customer=?, address=?,
            manager=?, content=?, memo=?, updated_at=?
        WHERE id=?
    """, (
        data.get("type", "실측"), data.get("status", "예정"),
        data.get("company", ""), data.get("scheduled_date", ""),
        data.get("customer", ""), data.get("address", ""),
        data.get("manager", ""), data.get("content", ""),
        data.get("memo", ""), now, item_id
    ))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/construction/<int:item_id>", methods=["DELETE"])
@require_login
def api_construction_delete(item_id):
    conn = get_conn()
    conn.execute("DELETE FROM construction_items WHERE id=?", (item_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/construction/docs", methods=["GET"])
@require_login
def api_construction_docs_list():
    conn = get_conn()
    rows = conn.execute("SELECT * FROM construction_docs ORDER BY id DESC").fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])

@app.route("/api/construction/docs", methods=["POST"])
@require_login
def api_construction_docs_upload():
    f = request.files.get("file")
    title = (request.form.get("title") or "").strip()
    if not f:
        return jsonify({"ok": False, "error": "파일 없음"}), 400
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ext = Path(f.filename).suffix.lower()
    fname = f"{uuid.uuid4().hex}{ext}"
    f.save(CONSTRUCTION_DOCS_DIR / fname)
    conn = get_conn()
    conn.execute("""
        INSERT INTO construction_docs (title, filename, original_name, uploaded_by, created_at)
        VALUES (?,?,?,?,?)
    """, (title or f.filename, fname, f.filename, session.get("name") or "", now))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/construction/docs/<int:doc_id>", methods=["DELETE"])
@require_login
def api_construction_docs_delete(doc_id):
    conn = get_conn()
    row = conn.execute("SELECT filename FROM construction_docs WHERE id=?", (doc_id,)).fetchone()
    if row:
        p = CONSTRUCTION_DOCS_DIR / row["filename"]
        if p.exists():
            p.unlink()
        conn.execute("DELETE FROM construction_docs WHERE id=?", (doc_id,))
        conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/construction_docs/<path:filename>")
@require_login
def construction_doc_file(filename):
    return send_from_directory(CONSTRUCTION_DOCS_DIR, filename)

@app.route("/api/as/videos", methods=["GET"])
def api_as_videos():
    try:
        wb = load_workbook(AS_VIDEO_EXCEL_PATH, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        data_start = 0
        for i, row in enumerate(rows):
            if row and row[0] == "제품군":
                data_start = i + 1
                break
        items = []
        for row in rows[data_start:]:
            if not row or not row[0]:
                continue
            items.append({
                "product_group": str(row[0] or "").strip(),
                "content": str(row[1] or "").strip(),
                "url": str(row[2] or "").strip(),
            })
        return jsonify(items)
    except Exception:
        return jsonify([])


@app.route("/api/as_ref", methods=["GET"])
def api_as_ref_list():
    product_group = (request.args.get("product_group") or "").strip()
    conn = get_conn()
    if product_group and product_group != "전체":
        rows = conn.execute(
            "SELECT * FROM as_ref_items WHERE product_group=? ORDER BY sort_order, id",
            (product_group,)
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM as_ref_items ORDER BY product_group, sort_order, id"
        ).fetchall()
    conn.close()
    return jsonify([row_to_dict(r) for r in rows])


@app.route("/api/as_ref", methods=["POST"])
def api_as_ref_create():
    data = request.json or {}
    title = (data.get("title") or "").strip()
    if not title:
        return jsonify({"ok": False, "error": "제목을 입력하세요."}), 400
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    cur = conn.execute("""
        INSERT INTO as_ref_items (product_group, title, content, url, url_type, sort_order, created_by, created_at, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        (data.get("product_group") or "공통").strip(),
        title,
        (data.get("content") or "").strip(),
        (data.get("url") or "").strip(),
        (data.get("url_type") or "").strip(),
        int(data.get("sort_order") or 0),
        session.get("name") or session.get("username") or "",
        now, now
    ))
    conn.commit()
    item_id = cur.lastrowid
    conn.close()
    return jsonify({"ok": True, "id": item_id})


@app.route("/api/as_ref/<int:item_id>", methods=["PUT"])
def api_as_ref_update(item_id):
    data = request.json or {}
    title = (data.get("title") or "").strip()
    if not title:
        return jsonify({"ok": False, "error": "제목을 입력하세요."}), 400
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    conn.execute("""
        UPDATE as_ref_items SET product_group=?, title=?, content=?, url=?, url_type=?, sort_order=?, updated_at=?
        WHERE id=?
    """, (
        (data.get("product_group") or "공통").strip(),
        title,
        (data.get("content") or "").strip(),
        (data.get("url") or "").strip(),
        (data.get("url_type") or "").strip(),
        int(data.get("sort_order") or 0),
        now, item_id
    ))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/as_ref/<int:item_id>", methods=["DELETE"])
def api_as_ref_delete(item_id):
    conn = get_conn()
    conn.execute("DELETE FROM as_ref_items WHERE id=?", (item_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


def ensure_work_orders_web_table_v29():
    conn = get_conn()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS work_orders_web (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            created_date TEXT,
            updated_at TEXT,
            order_date TEXT,
            production_date TEXT,
            due_date TEXT,
            customer TEXT,
            product_group TEXT,
            qty TEXT,
            model TEXT,
            detail TEXT,
            ship_type TEXT,
            site_name TEXT,
            memo TEXT,
            customer_manager TEXT,
            custom_1 TEXT,
            custom_2 TEXT,
            custom_note TEXT,
            status TEXT
        )
    """)
    conn.commit()
    conn.close()

def work_order_web_row_to_dict_v29(row):
    return {k: row[k] for k in row.keys()}

@app.route("/api/work_orders_web", methods=["GET"])
def api_work_orders_web_list_v29():
    ensure_work_orders_web_systemdoor_columns_v30()
    q = (request.args.get("q") or "").strip()
    status = (request.args.get("status") or "").strip()
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()

    where = []
    params = []
    if q:
        where.append("(customer LIKE ? OR site_name LIKE ? OR model LIKE ? OR detail LIKE ? OR memo LIKE ?)")
        like = f"%{q}%"
        params += [like, like, like, like, like]
    if status:
        where.append("status=?")
        params.append(status)
    if date_from:
        where.append("substr(created_date,1,10)>=?")
        params.append(date_from)
    if date_to:
        where.append("substr(created_date,1,10)<=?")
        params.append(date_to)

    sql = "SELECT * FROM work_orders_web"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY id DESC"

    conn = get_conn()
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([work_order_web_row_to_dict_v29(r) for r in rows])

@app.route("/api/work_orders_web", methods=["POST"])
def api_work_orders_web_create_v29():
    ensure_work_orders_web_systemdoor_columns_v30()
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fields = ["order_date","production_date","due_date","customer","product_group","qty","model","detail","ship_type","site_name","memo","customer_manager","custom_1","custom_2","custom_note","sd_color","sd_glass_thickness","sd_open_direction","sd_width","sd_height","sd_hinge_direction","sd_door_closer","sd_lock","sd_handle_height","sd_note","status"]
    values = {k: ("" if data.get(k) is None else str(data.get(k)).strip()) for k in fields}
    values["status"] = values.get("status") or "작성중"

    conn = get_conn()
    cols = ["created_date","updated_at"] + fields
    vals = [now, now] + [values[k] for k in fields]
    placeholders = ",".join(["?"] * len(cols))
    cur = conn.execute(f"INSERT INTO work_orders_web ({','.join(cols)}) VALUES ({placeholders})", vals)
    conn.commit()
    new_id = cur.lastrowid
    conn.close()
    return jsonify({"ok": True, "id": new_id})

@app.route("/api/work_orders_web/<int:item_id>", methods=["GET"])
def api_work_orders_web_get_v29(item_id):
    ensure_work_orders_web_systemdoor_columns_v30()
    conn = get_conn()
    row = conn.execute("SELECT * FROM work_orders_web WHERE id=?", (item_id,)).fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "not found"}), 404
    return jsonify(work_order_web_row_to_dict_v29(row))

@app.route("/api/work_orders_web/<int:item_id>", methods=["PUT"])
def api_work_orders_web_update_v29(item_id):
    ensure_work_orders_web_systemdoor_columns_v30()
    data = request.json or {}
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fields = ["order_date","production_date","due_date","customer","product_group","qty","model","detail","ship_type","site_name","memo","customer_manager","custom_1","custom_2","custom_note","sd_color","sd_glass_thickness","sd_open_direction","sd_width","sd_height","sd_hinge_direction","sd_door_closer","sd_lock","sd_handle_height","sd_note","status"]
    values = {k: ("" if data.get(k) is None else str(data.get(k)).strip()) for k in fields}
    values["status"] = values.get("status") or "작성중"

    conn = get_conn()
    row = conn.execute("SELECT id FROM work_orders_web WHERE id=?", (item_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"ok": False, "error": "작업지시서를 찾을 수 없습니다."}), 404
    set_sql = ",".join([f"{k}=?" for k in fields] + ["updated_at=?"])
    params = [values[k] for k in fields] + [now, item_id]
    conn.execute(f"UPDATE work_orders_web SET {set_sql} WHERE id=?", params)
    conn.commit()
    conn.close()
    return jsonify({"ok": True})

@app.route("/api/work_orders_web/<int:item_id>", methods=["DELETE"])
def api_work_orders_web_delete_v29(item_id):
    ensure_work_orders_web_systemdoor_columns_v30()
    conn = get_conn()
    # 연결된 생산스케쥴 조회 후 calendar_events·photos 포함 함께 삭제
    sch = conn.execute("SELECT id FROM schedules WHERE work_order_id=?", (str(item_id),)).fetchone()
    if sch:
        sch_id = sch["id"]
        conn.execute("DELETE FROM calendar_events WHERE schedule_id=?", (sch_id,))
        conn.execute("DELETE FROM schedule_photos WHERE schedule_id=?", (sch_id,))
        conn.execute("DELETE FROM schedules WHERE id=?", (sch_id,))
    conn.execute("DELETE FROM work_orders_web WHERE id=?", (item_id,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})



def ensure_work_orders_web_systemdoor_columns_v30():
    ensure_work_orders_web_table_v29()
    extra_cols = [
        "sd_color", "sd_glass_thickness", "sd_open_direction", "sd_width", "sd_height",
        "sd_hinge_direction", "sd_door_closer", "sd_lock", "sd_handle_height", "sd_note"
    ]
    conn = get_conn()
    existing = [r["name"] for r in conn.execute("PRAGMA table_info(work_orders_web)").fetchall()]
    for c in extra_cols:
        if c not in existing:
            conn.execute(f"ALTER TABLE work_orders_web ADD COLUMN {c} TEXT")
    conn.commit()
    conn.close()

def find_systemdoor_template_v30():
    candidates = [
        WORK_ORDER_TEMPLATE_DIR / "시스템도어.xlsm",
        WORK_ORDER_TEMPLATE_DIR / "시스템도어.xlsx",
        APP_DIR / "시스템도어.xlsm",
        APP_DIR / "시스템도어.xlsx",
    ]
    for p in candidates:
        if p.exists():
            return p
    return candidates[0]

def cell_value_to_text_v30(v):
    if v is None:
        return ""
    try:
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
    except Exception:
        pass
    return str(v)

@app.route("/api/work_order_systemdoor/calc", methods=["POST"])
def api_work_order_systemdoor_calc_v30():
    data = request.json or {}
    template = find_systemdoor_template_v30()
    if not template.exists():
        return jsonify({"ok": False, "error": f"시스템도어 템플릿 파일이 없습니다: {template}"}), 404

    try:
        from openpyxl import load_workbook
        import tempfile, shutil, subprocess, os, time
    except Exception as e:
        return jsonify({"ok": False, "error": f"필요 모듈을 불러오지 못했습니다: {e}"}), 500

    input_map = {
        "A11": data.get("sd_color", ""),
        "C11": data.get("sd_glass_thickness", ""),
        "E11": data.get("sd_open_direction", ""),
        "B15": data.get("sd_width", ""),
        "E15": data.get("sd_height", ""),
        "B17": data.get("sd_hinge_direction", ""),
        "B19": data.get("sd_door_closer", ""),
        "F19": data.get("sd_lock", ""),
        "F21": data.get("sd_handle_height", ""),
    }
    output_map = {
        "door_width": "B24",
        "door_height": "D24",
        "door_gasket": "F24",
        "osai_width": "B27",
        "osai_height": "D27",
        "osai_type": "F27",
        "glass_width": "B30",
        "glass_height": "D30",
    }

    tmpdir = Path(tempfile.mkdtemp(prefix="systemdoor_calc_"))
    try:
        tmpfile = tmpdir / template.name
        shutil.copy2(template, tmpfile)

        wb = load_workbook(tmpfile, keep_vba=tmpfile.suffix.lower()==".xlsm")
        ws = wb.active
        for cell, value in input_map.items():
            ws[cell] = value
        try:
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
        except Exception:
            pass
        wb.save(tmpfile)
        wb.close()

        # LibreOffice가 있으면 수식 재계산 시도
        lo_candidates = [
            str(SOFFICE_PATH),
            "soffice",
            r"C:\Program Files\LibreOffice\program\soffice.exe",
            r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
        ]
        for soffice in lo_candidates:
            try:
                subprocess.run(
                    [soffice, "--headless", "--convert-to", "xlsx", "--outdir", str(tmpdir), str(tmpfile)],
                    stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=45
                )
                converted = tmpdir / (tmpfile.stem + ".xlsx")
                if converted.exists():
                    tmpfile = converted
                    break
            except Exception:
                pass

        wb2 = load_workbook(tmpfile, data_only=True, read_only=True)
        ws2 = wb2.active
        outputs = {key: cell_value_to_text_v30(ws2[cell].value) for key, cell in output_map.items()}
        outputs["osai_gasket"] = ""
        wb2.close()

        return jsonify({"ok": True, "outputs": outputs})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
    finally:
        try:
            shutil.rmtree(tmpdir, ignore_errors=True)
        except Exception:
            pass


# ── REGEN XML 생성 ────────────────────────────────────────────────────────────
_RG_BAR_ENDS = 80
_RG_SAW = 8
_RG_REG05_OPTS = [4300, 5000]
_RG_REG06_LEN = 5700
_RG_DC_BASE = 245

def _rg_calc_reg06(frame, door, hinge, fix):
    reg06 = frame - 100
    door_sz = (frame - 80) if fix == '없음' else door
    fix1, fix2 = None, None
    if hinge == '좌측':
        if fix == '없음': dc = _RG_DC_BASE
        elif fix == '우측': dc = _RG_DC_BASE; fix1 = (door_sz - 20) + 17.5
        elif fix == '좌측': dc = reg06 - (door_sz - 20) + _RG_DC_BASE; fix1 = reg06 - (door_sz - 20) - 17.5
        elif fix == '양측': dc = (reg06/2)-((door_sz-20)/2)+_RG_DC_BASE; fix1=(reg06/2)-((door_sz-20)/2)-17.5; fix2=(reg06/2)+((door_sz-20)/2)+17.5
        else: dc = _RG_DC_BASE
    else:
        if fix == '없음': dc = reg06 - _RG_DC_BASE
        elif fix == '우측': dc = door_sz - 10 - _RG_DC_BASE; fix1 = (door_sz - 20) + 17.5
        elif fix == '좌측': dc = reg06 - _RG_DC_BASE; fix1 = reg06 - (door_sz - 20) - 17.5
        elif fix == '양측': dc = (reg06/2)+((door_sz-20)/2)-_RG_DC_BASE; fix1=(reg06/2)-((door_sz-20)/2)-17.5; fix2=(reg06/2)+((door_sz-20)/2)+17.5
        else: dc = reg06 - _RG_DC_BASE
    return dict(reg06=reg06, dc=dc, fix1=fix1, fix2=fix2)

def _rg_usable_2p(bl): return bl - _RG_BAR_ENDS - _RG_SAW*3 - _RG_BAR_ENDS
def _rg_usable_1p(bl): return bl - _RG_BAR_ENDS - _RG_SAW*2 - _RG_BAR_ENDS

def _rg_choose_bar(reg05):
    for bl in _RG_REG05_OPTS:
        if 2*reg05 <= _rg_usable_2p(bl): return bl, 2
    for bl in _RG_REG05_OPTS:
        if reg05 <= _rg_usable_1p(bl): return bl, 1
    return None, None

def _rg_pack(lengths, bar_len):
    limit = bar_len - _RG_BAR_ENDS - _RG_SAW - _RG_BAR_ENDS - _RG_SAW
    bars, cur, cur_sum = [], [], 0
    for ln in sorted(lengths, reverse=True):
        if cur_sum + ln + len(cur)*_RG_SAW <= limit: cur.append(ln); cur_sum += ln
        else:
            if cur: bars.append(cur)
            cur, cur_sum = [ln], ln
    if cur: bars.append(cur)
    return bars

def _rg_sz(w, h): return f"W{w}" if h <= 2000 else f"W{w}*H{h}"
def _rg_pid(seq, ds): return f"{seq:05d}{ds}"

def _rg_angles(pairs, ind):
    L = [f'{ind}<Angles>']
    for k, v in pairs: L.append(f'{ind}  <{k}>{v}</{k}>')
    L.append(f'{ind}</Angles>'); return L

def _rg_bar_ang(): return [('LeftAngle','90'),('RightAngle','90'),('LeftHeightAngle','0'),('RightHeightAngle','0'),('LeftWidthAngle','0'),('RightWidthAngle','0')]
def _rg_frame_ang(): return [('LeftAngle','0'),('RightAngle','0'),('LeftHeightAngle','90'),('RightHeightAngle','90'),('LeftWidthAngle','90'),('RightWidthAngle','90')]

def _rg_info_num(ind):
    L = [f'{ind}<InfoFields>']
    for i in range(1,5): L.append(f'{ind}  <Info id="{i}" />')
    L.append(f'{ind}</InfoFields>'); return L

def _rg_piece_info(label, sz, ind):
    return [f'{ind}<InfoFields>', f'{ind}  <Info id="" />', f'{ind}  <Info id="{label}" />', f'{ind}  <Info id="{sz}" />', f'{ind}  <Info id="Area00" />', f'{ind}</InfoFields>']

def _rg_piece(pid, length, label, sz, macro, pos, ind='      '):
    i2 = ind+'  '
    L = [f'{ind}<Piece ID="{pid}" Length="{length}" Quantity="1" enabled="true" mat_type="1" wash_loss="0" is_wash="3">']
    L += _rg_angles(_rg_frame_ang(), i2) + _rg_piece_info(label, sz, i2)
    L += [f'{i2}<Machinings>', f'{i2}  <Macro Type="MCR" Name="{macro}" PositionX="{pos}" Comment="0" />', f'{i2}</Machinings>', f'{ind}</Piece>']
    return L

def _rg_piece_mm(pid, length, label, sz, macros, ind='      '):
    i2 = ind+'  '
    L = [f'{ind}<Piece ID="{pid}" Length="{length}" Quantity="1" enabled="true" mat_type="1" wash_loss="0" is_wash="3">']
    L += _rg_angles(_rg_frame_ang(), i2) + _rg_piece_info(label, sz, i2)
    L.append(f'{i2}<Machinings>')
    for mn, mp in macros: L.append(f'{i2}  <Macro Type="MCR" Name="{mn}" PositionX="{mp}" Comment="0" />')
    L += [f'{i2}</Machinings>', f'{ind}</Piece>']
    return L

def _rg_bar(bid, blen, pieces, ind='    '):
    i2 = ind+'  '
    L = [f'{ind}<Bar ID="{bid}" Length="{blen}" Quantity="1" enabled="true">']
    L += _rg_angles(_rg_bar_ang(), i2) + _rg_info_num(i2) + pieces
    L.append(f'{ind}</Bar>'); return L

def _rg_build_xml(rows_data):
    import datetime as _dt
    today = _dt.date.today()
    date_str = f"{today.day:02d}.{today.month:02d}.{today.year}"
    ds = f"{today.year%100:02d}{today.month:02d}{today.day:02d}14"
    rs = sorted(rows_data, key=lambda x: x['height'], reverse=True)
    n = len(rs) - 1
    L = ['<?xml version="1.0" encoding="utf-8" standalone="no"?>', '<Unilink>', f'  <FileInfo CreatedBy="string" CreationTime="{date_str}" />', '']
    color_tag = '  ũ    '

    # REG-05
    L += ['','','', '  <Profile Serie="REGEN" Name="REG-05" Width="130" Height="70" enabled="1">',
          '    <Color>', f'      <Inside Color="{color_tag}" />', f'      <Outside Color="{color_tag}" />', '    </Color>'] + _rg_info_num('    ')
    bar_id = 1
    for idx, row in enumerate(rs):
        inv = n - idx; sz = _rg_sz(row['width'], row['height'])
        reg05 = float(row['height'] - 30)
        bl, per = _rg_choose_bar(int(reg05))
        if bl is None: raise ValueError(f"'{row['name']}' REG-05 {reg05}mm 원재료 범위 초과")
        rh = _rg_piece(_rg_pid(inv*32+8, ds), reg05, f'{row["name"]}-FrameR', sz, 'REG-05_RH', '16.1')
        for li, line in enumerate(rh):
            if 'REG-05_RH' in line: rh.insert(li+1, ''); break
        lh = _rg_piece(_rg_pid(inv*32+7, ds), reg05, f'{row["name"]}-FrameL', sz, 'REG-05_LH', '16.1')
        if per == 2:
            L += _rg_bar(bar_id, bl, rh+lh); bar_id += 1
        else:
            L += _rg_bar(bar_id, bl, rh); bar_id += 1
            L += _rg_bar(bar_id, bl, lh); bar_id += 1
    L += ['  </Profile>', '']

    # REG-08
    L += ['','','', '  ', '','','','',
          '  <Profile Serie="REGEN" Name="REG-08" Width="130" Height="15" enabled="1">',
          '    <Color>', f'      <Inside Color="{color_tag}" />', f'      <Outside Color="{color_tag}" />', '    </Color>'] + _rg_info_num('    ')
    for ft, mn, px in [('T','REG-08_T','5.5'),('B','REG-08_B','5.5')]:
        widths = sorted([r['width'] for r in rows_data], reverse=True)
        for bw in _rg_pack(widths, 5700):
            plines = []
            for w in bw:
                row = next(r for r in rows_data if r['width']==w)
                oi = next(i for i,r in enumerate(rs) if r['name']==row['name'])
                seq = (n-oi)*32 + (2 if ft=='T' else 5)
                plines += _rg_piece(_rg_pid(seq,ds), float(w), f'{row["name"]}-Frame{ft}', _rg_sz(row['width'],row['height']), mn, px)
            L += _rg_bar(bar_id, 5700, plines); bar_id += 1
    L += ['', '  </Profile>', '']

    # REG-06
    L += ['','','', '  <Profile Serie="REGEN" Name="REG-06" Width="130" Height="20" enabled="1">',
          '    <Color>', f'      <Inside Color="{color_tag}" />', f'      <Outside Color="{color_tag}" />', '    </Color>'] + _rg_info_num('    ')
    reg06_lens = [r['lk']['reg06'] for r in rs]
    for blens in _rg_pack(reg06_lens, _RG_REG06_LEN):
        plines = []
        for ln in blens:
            row = next(r for r in rs if r['lk']['reg06']==ln)
            oi = next(i for i,r in enumerate(rs) if r['name']==row['name'])
            seq = (n-oi)*32 + 6; lk = row['lk']
            macros = [('REG-06_DC', str(lk['dc']))]
            if lk['fix1'] is not None: macros.append(('REG-06_FIX', str(lk['fix1'])))
            if lk['fix2'] is not None: macros.append(('REG-06_FIX', str(lk['fix2'])))
            macros.append(('REG-06-20','0'))
            plines += _rg_piece_mm(_rg_pid(seq,ds), float(ln), f'{row["name"]}-REG06', _rg_sz(row['width'],row['height']), macros)
        L += _rg_bar(bar_id, _RG_REG06_LEN, plines); bar_id += 1
    L += ['', '  </Profile>', '', '</Unilink>']
    return '\r\n'.join(L)


@app.route("/api/regen/generate_xml", methods=["POST"])
@require_perm("work_order")
def regen_generate_xml_api():
    import datetime as _dt
    from flask import Response as _Resp
    data = request.json or {}
    rows_input = data.get("rows", [])
    if not rows_input:
        return jsonify({"ok": False, "error": "데이터가 없습니다."})
    processed = []
    for r in rows_input:
        try:
            w, h, d = int(r['width']), int(r['height']), int(r['door'])
            hinge = r.get('hinge', '좌측')
            fix = r.get('fix', '없음')
            lk = _rg_calc_reg06(w, d, hinge, fix)
            processed.append({'name': r['name'], 'width': w, 'height': h, 'door': d, 'hinge': hinge, 'fix': fix, 'lk': lk})
        except Exception as e:
            return jsonify({"ok": False, "error": f"데이터 오류: {e}"})
    try:
        content = _rg_build_xml(processed)
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})
    today = _dt.date.today()
    filename = f"regen_{today.strftime('%Y%m%d')}.XML"
    return _Resp('﻿' + content,
        mimetype="application/xml; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.route("/api/att/save", methods=["POST"])
@require_login
def att_save():
    body = request.json or {}
    month = body.get("month", "")
    title = body.get("title") or month
    saved_id = body.get("id")
    data = json.dumps(body, ensure_ascii=False)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn = get_conn()
    if saved_id:
        conn.execute("UPDATE att_saved SET title=?, month=?, data=?, updated_at=? WHERE id=?",
                     (title, month, data, now, saved_id))
    else:
        conn.execute("INSERT INTO att_saved (title, month, data, created_at, updated_at) VALUES (?,?,?,?,?)",
                     (title, month, data, now, now))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


@app.route("/api/att/list", methods=["GET"])
@require_login
def att_list():
    conn = get_conn()
    rows = conn.execute("""
        SELECT id, title, month, created_at, updated_at FROM att_saved
        ORDER BY COALESCE(NULLIF(updated_at,''), created_at) DESC
    """).fetchall()
    conn.close()
    return jsonify([{"id": r[0], "title": r[1], "month": r[2], "created_at": r[3], "updated_at": r[4]} for r in rows])


@app.route("/api/att/load/<int:rid>", methods=["GET"])
@require_login
def att_load(rid):
    conn = get_conn()
    row = conn.execute("SELECT data FROM att_saved WHERE id=?", (rid,)).fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "not found"}), 404
    return jsonify(json.loads(row[0]))


@app.route("/api/att/delete/<int:rid>", methods=["DELETE"])
@require_login
def att_delete(rid):
    conn = get_conn()
    conn.execute("DELETE FROM att_saved WHERE id=?", (rid,))
    conn.commit()
    conn.close()
    return jsonify({"ok": True})


init_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=False, threaded=True)
